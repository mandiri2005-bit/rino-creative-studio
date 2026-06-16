# -*- coding: utf-8 -*-
"""
LaoZhang API - FastAPI Backend
Run with: python laozhang_api.py
"""
import os
import uuid
import asyncio
from pathlib import Path
import base64
import json
import time
import threading
from contextvars import ContextVar
from typing import Iterator, Optional
from openai import OpenAI
from datetime import datetime
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Header, Request, Depends
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, validator
import uvicorn
import requests as _requests
import re as _re

# ---------------------------------------------------------------------------
# Logging — must be configured before any library import touches the root
# logger.  qdrant_index used to call basicConfig(level=ERROR) which silenced
# httpx INFO lines (Qdrant HTTP requests) and rag_narration _log.info lines.
# We configure a clean root handler here so those lines reach stdout.
# ---------------------------------------------------------------------------
import logging as _logging
if not _logging.root.handlers:                       # only if nothing set yet
    _logging.basicConfig(
        level=_logging.INFO,
        format="%(asctime)s  %(levelname)s  %(name)s  %(message)s",
        datefmt="%H:%M:%S",
    )
_logging.getLogger("httpx").setLevel(_logging.DEBUG)        # Qdrant HTTP requests
_logging.getLogger("httpcore").setLevel(_logging.DEBUG)     # underlying transport
_logging.getLogger("rag_narration").setLevel(_logging.INFO) # RAG pipeline steps

try:
    from moat.gutenberg.rag_narration import generate_rag_narration as _rag_generate
    RAG_AVAILABLE = True
except ImportError:
    RAG_AVAILABLE = False

# document parsing
import io

try:
    import pdfplumber

    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    import docx

    DOCX_OK = True
except ImportError:
    DOCX_OK = False

try:
    import openpyxl

    XLSX_OK = True
except ImportError:
    XLSX_OK = False

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
API_KEY = os.environ.get("LAOZHANG_API_KEY", "")
if not API_KEY:
    import warnings

    warnings.warn("LAOZHANG_API_KEY not set -- clients must provide X-LaoZhang-API-Key header")

# Separate key for image generation (can be same or different)
IMAGE_API_KEY = os.environ.get("LAOZHANG_IMAGE_API_KEY", API_KEY)

# API key for deepseek-v4-pro and deepseek-r1 — same BASE_URL (LaoZhang), different key
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
DEEPSEEK_DIRECT_MODELS = {"deepseek-v4-pro", "deepseek-r1"}

# Direct Google key for Nusantara corpus text-enhance + Qdrant embedding
GEMINI_API_KEY   = os.environ.get("GEMINI_API_KEY", "")
# Qdrant cloud cluster for nusantara_visual_v1 (separate from legacy local qdrant:6333)
QDRANT_CLOUD_URL = os.environ.get("QDRANT_CLOUD_URL", "")
QDRANT_CLOUD_KEY = os.environ.get("QDRANT_CLOUD_KEY", "")

BASE_URL = "https://api.laozhang.ai/v1"
MCP_API_URL = os.environ.get("MCP_API_URL", "http://127.0.0.1:8001")  # mcp_files.py sidecar
IMAGE_URL = "https://api.laozhang.ai/v1"
GOOGLE_IMAGE_BASE = "https://api.laozhang.ai/v1beta/models"

# ── Vertex AI / OAuth credentials (no API key) ───────────────────────────────
GCP_PROJECT_ID     = os.environ.get("GCP_PROJECT_ID", "")
GCP_REFRESH_TOKEN  = os.environ.get("GCP_REFRESH_TOKEN", "")
GCP_CLIENT_ID      = os.environ.get("GCP_CLIENT_ID", "")
GCP_CLIENT_SECRET  = os.environ.get("GCP_CLIENT_SECRET", "")

GCP_LOCATION       = os.environ.get("GCP_LOCATION", "global")

_vertex_ready = False
_gcp_creds = None  # OAuth Credentials, reused by both Imagen (vertexai) and Gemini (google.genai) paths

def _ensure_vertex():
    global _vertex_ready, _gcp_creds
    if _vertex_ready:
        return True
    if not all([GCP_PROJECT_ID, GCP_REFRESH_TOKEN, GCP_CLIENT_ID, GCP_CLIENT_SECRET]):
        return False
    try:
        from google.oauth2.credentials import Credentials as _GCreds
        import vertexai as _vertexai
        _creds = _GCreds(
            token=None,
            refresh_token=GCP_REFRESH_TOKEN,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=GCP_CLIENT_ID,
            client_secret=GCP_CLIENT_SECRET,
        )
        _vertexai.init(project=GCP_PROJECT_ID, credentials=_creds)
        _gcp_creds = _creds
        _vertex_ready = True
        return True
    except Exception as e:
        import warnings
        warnings.warn(f"Vertex AI init failed: {e}")
        return False

def _is_gemini_image_model(m: str) -> bool:
    """Nano Banana lineup (gemini-*-image) goes through the Gemini API, NOT ImageGenerationModel."""
    m = (m or "").lower()
    return m.startswith("gemini-") and "image" in m

_ensure_vertex()

# -- Best balance -- reliable + affordable --------------------------------
# -- Power ---------------------------------------------------------------
# -- Ultra-cheap -- high volume / simple tasks ----------------------------
MODELS = {
    # Best balance
    "gemini-2.5-flash": "gemini-2.5-flash",
    "deepseek-v3": "deepseek-chat",
    "gpt-4o-mini": "gpt-4o-mini",
    "qwen-max": "qwen-max",
    "gemini-2.5-flash-lite": "gemini-2.5-flash-lite",
    # Power
    "gemini-2.5-pro": "gemini-2.5-pro",
    "claude-sonnet": "claude-sonnet-4-6-thinking",
    "gpt-4o": "gpt-4o",
    "grok-4": "grok-4-latest",
    "claude-opus-4-6": "claude-opus-4-6",
    "claude-opus-4-7": "claude-opus-4-7",
    "claude-opus-4-7-thinking": "claude-opus-4-7-thinking",
    # Ultra-cheap
    "glm": "glm-4.5-flash",
    "gpt-5-nano": "gpt-5-nano",
    "deepseek-v3-0324": "deepseek-v3-250324",
    "deepseek-v4-pro": "deepseek-v4-pro",
    "deepseek-r1": "deepseek-r1",
    "grok-4-fast": "grok-4-fast",
    "gemini-3-flash": "gemini-3-flash-preview",
}

# Models that support tool/function calling via OpenAI-compatible endpoint
TOOL_CAPABLE_MODELS = {
    "claude-sonnet-4-6", "claude-sonnet-4-6-thinking", "claude-opus-4-6", "claude-opus-4-7",
    "claude-opus-4-7-thinking",
    "gpt-4o", "gpt-4o-mini", "gpt-5-nano",
    "gemini-2.5-pro", "gemini-2.5-flash", "gemini-2.5-flash-lite",
    "deepseek-chat", "deepseek-v3-250324", "deepseek-v4-pro", "deepseek-r1",
    "grok-4-latest", "grok-4-fast",
}

# Models confirmed to accept OpenAI-style multimodal image_url payloads via LaoZhang.
# Checked against upstream model identifiers (post MODELS.get(...) resolution).
# If your model isn't here but should support vision, add it AND test with /upload + chat.
VISION_CAPABLE_MODELS = {
    # OpenAI vision
    "gpt-4o", "gpt-4o-mini", "gpt-4-turbo",
    # Anthropic vision
    "claude-sonnet-4-6", "claude-sonnet-4-6-thinking",
    "claude-opus-4-6", "claude-opus-4-7", "claude-opus-4-7-thinking",
    "claude-3-5-sonnet", "claude-3-5-haiku", "claude-3-opus",
    # Gemini vision (all 2.5+ and 3.x)
    "gemini-2.5-pro", "gemini-2.5-flash", "gemini-2.5-flash-lite",
    "gemini-3-flash-preview", "gemini-3.1-flash", "gemini-3.1-pro-preview",
    # Grok vision
    "grok-4-latest", "grok-4-fast",
    # Qwen vision
    "qwen-max", "qwen-vl-max",
}

# MCP tool definitions sent to models that support tool calling
MCP_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "search_files",
            "description": (
                "Search the user's local folder for relevant content using semantic+keyword hybrid search. "
                "Call this when the user's question relates to their files, documents, code, or notes. "
                "Returns the most relevant excerpts."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {
                        "type": "string",
                        "description": "The search query -- what to look for in the files"
                    },
                    "paths": {
                        "type": "string",
                        "description": "Optional comma-separated list of file paths to restrict search to. Leave empty to search all files.",
                        "default": ""
                    }
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": "Read the complete content of a specific file from the user's local folder.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Relative path to the file within the indexed folder"
                    }
                },
                "required": ["path"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_files",
            "description": "List all available files in the user's indexed folder.",
            "parameters": {
                "type": "object",
                "properties": {}
            }
        }
    }
]

# Safe output-token ceiling per resolved model name
MODEL_MAX_TOKENS: dict[str, int] = {
    # Best balance
    "gemini-2.5-flash": 16384,
    "deepseek-chat": 8192,
    "gpt-4o-mini": 16384,
    "qwen-max": 8192,
    "gemini-2.5-flash-lite": 8192,
    # Power
    "gemini-2.5-pro": 65536,
    "claude-sonnet-4-6": 8192,
    "claude-sonnet-4-6-thinking": 64000,
    "gpt-4o": 16384,
    "grok-4-latest": 32000,
    "claude-opus-4-6": 32000,
    "claude-opus-4-7": 32000,
    "claude-opus-4-7-thinking": 32000,
    # Ultra-cheap
    "glm-4.5-flash": 4096,
    "gpt-5-nano": 16384,
    "gpt-5.4-nano": 16384,
    "gpt-5-mini": 16384,
    "gpt-5.4-mini": 16384,
    "gpt-5.1": 32000,
    "gpt-5": 32000,
    "gpt-5.2": 32000,
    "gpt-5.4": 32000,
    "gpt-5.5": 32000,
    "gpt-5-pro": 32000,
    "deepseek-v3-250324": 8192,
    "deepseek-v4-pro": 65536,
    "deepseek-r1": 65536,
    "grok-4-fast": 8192,
    "gemini-3-flash-preview": 8192,
}
DEFAULT_MAX_TOKENS = 16384

# MAX_SESSIONS removed — session persistence is in PostgreSQL, no eviction needed.

# ── Cost estimation (USD) — used by log_usage() after each stream ────────────
# Keys match either the user-facing alias OR the resolved upstream name.
# Prices are best-effort $/1M tokens (input, output); update as rates change.
_MODEL_COSTS_PER_M: dict[str, tuple[float, float]] = {
    # OpenAI
    "gpt-4o-mini":            (0.15,   0.60),
    "gpt-4o":                 (5.00,  15.00),
    "gpt-4.1-mini":           (0.40,   1.60),
    "gpt-4.1":                (2.00,   8.00),
    "gpt-5-nano":             (0.05,   0.40),
    "o3-mini":                (1.10,   4.40),
    "o3":                    (10.00,  40.00),
    # Anthropic
    "claude-haiku":           (0.80,   4.00),
    "claude-sonnet":          (3.00,  15.00),
    "claude-opus":           (15.00,  75.00),
    # DeepSeek (alias + upstream)
    "deepseek-chat":          (0.27,   1.10),
    "deepseek-v3":            (0.27,   1.10),
    "deepseek-v4-pro":        (0.55,   2.19),
    "deepseek-r1":            (0.55,   2.19),
    # Gemini (longer prefixes first so -lite/-pro win over -flash)
    "gemini-2.5-flash-lite":  (0.075,  0.30),
    "gemini-2.5-flash":       (0.15,   0.60),
    "gemini-2.5-pro":         (1.25,  10.00),
    "gemini-3-flash":         (0.15,   0.60),
    "gemini-1.5-flash":       (0.075,  0.30),
    "gemini-1.5-pro":         (1.25,   5.00),
    # Others (best-effort estimates)
    "qwen-max":               (1.60,   6.40),
    "grok-4-fast":            (0.20,   0.50),
    "grok-4":                 (3.00,  15.00),
    "glm":                    (0.10,   0.10),
}

def _calc_cost(model: str, tokens_in: int, tokens_out: int) -> float:
    """Estimate USD cost. Checks the alias and the resolved upstream name,
    longest-matching prefix wins. Returns 0.0 only if nothing matches."""
    names = [model.lower(), str(MODELS.get(model, "")).lower()]
    for name in names:
        if not name:
            continue
        key = max((k for k in _MODEL_COSTS_PER_M if name.startswith(k)),
                  key=len, default=None)
        if key:
            in_p, out_p = _MODEL_COSTS_PER_M[key]
            return round((tokens_in * in_p + tokens_out * out_p) / 1_000_000, 8)
    return 0.0


async def _log_narasi_usage(tenant_id, user_id, model, resp, *, job_id=None, session_id=None):
    """Best-effort usage logging for narasi LLM endpoints — writes to usage_logs
    with endpoint='narasi'. Never raises: cost tracking must not break generation.
    `user_id` MUST be the resolved users.id UUID (not the raw Clerk id).
    `job_id` MUST be the internal jobs.id UUID (not the external 8-char id)."""
    try:
        usage = getattr(resp, "usage", None)
        tok_in  = int(getattr(usage, "prompt_tokens",     0) or 0) if usage else 0
        tok_out = int(getattr(usage, "completion_tokens", 0) or 0) if usage else 0
        cost = _calc_cost(model, tok_in, tok_out)
        _ml = (model or "").lower()
        if   _ml.startswith("gemini"):            _provider = "gemini"
        elif _ml.startswith("deepseek"):          _provider = "deepseek"
        elif _ml.startswith(("gpt", "o3", "o1")): _provider = "openai"
        else:                                     _provider = "laozhang"
        await db.log_usage(tenant_id, user_id, model, "narasi",
                           tok_in, tok_out, cost,
                           job_id=job_id, session_id=session_id, provider=_provider)
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("log_usage (narasi) failed (non-fatal): %s", _e)

# ---------------------------------------------------------------------------
# FastAPI app  — with DB lifespan (Phase 1 migration)
# ---------------------------------------------------------------------------
from contextlib import asynccontextmanager
import database as db
import redis_client as rc
import nusantara_corpus as _nc
from auth_middleware import (get_current_user, CurrentUser,
                             _tenant_id as _ctx_tenant_id, _user_id as _ctx_user_id)

@asynccontextmanager
async def lifespan(application):
    await db.init_db()
    await rc.init_redis()
    yield
    await rc.close_redis()
    await db.close_db()

app = FastAPI(title="LaoZhang Chat API", lifespan=lifespan)

@app.get("/health")
def health():
    return {"status": "ok"}

# CORS origins from env (comma-separated). Defaults to local dev + Railway staging.
_cors_env = os.getenv("CORS_ORIGINS", "")
_cors_origins = [o.strip() for o in _cors_env.split(",") if o.strip()] or [
    "http://localhost:8080",
    "https://ravishing-miracle-production-01b2.up.railway.app",
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# Per-request API key override  (set via X-LaoZhang-API-Key header)
# _tenant_ctx / TenantContext / get_tenant_ctx are defined in auth_middleware
# and set by get_current_user() on every authenticated request.
# ---------------------------------------------------------------------------
from auth_middleware import TenantContext, _tenant_ctx, get_tenant_ctx

# Raw ContextVar for the X-LaoZhang-API-Key header value (WS1 behaviour)
_req_key_raw: ContextVar[str] = ContextVar("_req_key_raw", default="")

class _ReqKeyCompat:
    """
    Backward-compat shim: all existing _req_key.get() call sites continue to
    work — header override takes precedence, then tenant's stored key, then env.
    """
    def get(self) -> str:
        raw = _req_key_raw.get()
        if raw:
            return raw
        ctx = _tenant_ctx.get()
        return ctx.api_key or API_KEY

    def set(self, v: str):
        return _req_key_raw.set(v)

    def reset(self, token) -> None:
        _req_key_raw.reset(token)

_req_key = _ReqKeyCompat()

# "deepseek" = DEEPSEEK_API_KEY, "laozhang" = LAOZHANG_API_KEY
_deepseek_route: ContextVar[str] = ContextVar("_deepseek_route", default="deepseek")


@app.middleware("http")
async def key_override_middleware(request: Request, call_next):
    key = request.headers.get("X-LaoZhang-API-Key", "").strip()
    token = _req_key.set(key if key else API_KEY)
    # Take first value only — browser may send duplicate headers merged as "a, a"
    route = request.headers.get("X-DeepSeek-Route", "deepseek").split(",")[0].strip().lower()
    token_route = _deepseek_route.set(route if route in ("deepseek", "laozhang") else "deepseek")
    try:
        return await call_next(request)
    finally:
        _req_key.reset(token)
        _deepseek_route.reset(token_route)


# ---------------------------------------------------------------------------
# Cancel flags — now fully in Redis (rc.set_cancel / rc.is_cancelled /
# rc.clear_cancel). No in-process dict; cancel works across containers.
# The sync chat_stream generator still takes a threading.Event, which the
# async generate() loop sets when it observes the Redis flag (bridge below).
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Client factory  (uses per-request key if provided)
# ---------------------------------------------------------------------------
def make_client(model: str = "") -> OpenAI:
    """Return the right OpenAI-compatible client for the given model.
    For deepseek-v4-pro / deepseek-r1:
      X-DeepSeek-Route: deepseek  (default) -> DEEPSEEK_API_KEY + BASE_URL
      X-DeepSeek-Route: laozhang            -> LAOZHANG_API_KEY + BASE_URL
    All other models always use LAOZHANG_API_KEY.
    """
    resolved = MODELS.get(model, model)
    if resolved in DEEPSEEK_DIRECT_MODELS or model in DEEPSEEK_DIRECT_MODELS:
        if _deepseek_route.get() == "laozhang":
            return OpenAI(api_key=_req_key.get() or API_KEY, base_url=BASE_URL)
        key = DEEPSEEK_API_KEY
        if not key:
            raise ValueError(
                "DEEPSEEK_API_KEY is not set. Add DEEPSEEK_API_KEY to your .env file."
            )
        return OpenAI(api_key=key, base_url=BASE_URL)
    return OpenAI(api_key=_req_key.get() or API_KEY, base_url=BASE_URL)


# ── Review personas (rules) live SERVER-SIDE — never shipped to the client ────
import os as _os_rp, json as _json_rp
_REVIEW_PERSONAS = {}
try:
    with open(_os_rp.path.join(_os_rp.path.dirname(__file__), "review_personas.json"), encoding="utf-8") as _pf:
        _REVIEW_PERSONAS = _json_rp.load(_pf)
except Exception as _pe:
    import logging as _lg_rp; _lg_rp.getLogger("narasi").warning("review_personas.json load failed: %s", _pe)

def _review_persona_for(style):
    """Map a style id to a server-side review persona (rules). Keeps the rule text
    out of the client entirely."""
    s = (style or "").lower()
    if "harari" in s or "diamond" in s or "big history" in s or "academic popular" in s:
        key = "harari"
    elif "non-fiction" in s or "narrative" in s or "literary" in s or "journalistic" in s:
        key = "narrative"
    else:
        key = "default"
    return _REVIEW_PERSONAS.get(key) or _REVIEW_PERSONAS.get("default") or {}


# Maps laozhang_api style names -> style_rag_config keys
# style_rag_config._ALIASES handles all translation internally
# Keeping a minimal map here only for the rare styles not in style_rag_config
_RAG_STYLE_LEGACY = {
    "biography":            "pov_first_person",
    "documentary":          "natgeo",
    "science":              "youtube_popular_science",
    "finance":              "academic_popular",
    "economics":            "academic_popular",
    "business":             "academic_popular",
    "philosophical":        "literary_essay",
    # --- styles that don't map to a single Qdrant label → use None (no filter) ---
    "narrative non-fiction": None,   # broad genre — let semantic search pick best passages
    "narrative nonfiction":  None,
    "narrative_nonfiction":  None,
    "creative nonfiction":   None,
    "creative_nonfiction":   None,
}

def _rag_style(style: str) -> str | None:
    """Pass style key to style_rag_config — it handles all 13 style aliases.
    Returns None for broad genres (narrative non-fiction etc.) to skip Qdrant
    style filtering and rely on semantic search only.
    Falls back to legacy map, then returns style as-is as last resort."""
    s = style.lower().strip()
    if s in _RAG_STYLE_LEGACY:
        return _RAG_STYLE_LEGACY[s]  # may be None for broad genres
    return s


# ---------------------------------------------------------------------------
# MCP tool executor -- calls mcp_files.py sidecar
# ---------------------------------------------------------------------------
def execute_mcp_tool(tool_name: str, tool_args: dict) -> str:
    """Execute a tool call by hitting the mcp_files.py REST sidecar."""
    try:
        if tool_name == "search_files":
            query = tool_args.get("query", "")
            paths = tool_args.get("paths", "")
            url = f"{MCP_API_URL}/context?q={_urlencode(query)}&paths={_urlencode(paths)}"
            resp = _requests.get(url, timeout=15)
            if resp.ok:
                data = resp.json()
                ctx = data.get("context", "")
                mode = data.get("search_mode", "unknown")
                if ctx:
                    return f"[Search mode: {mode}]\n\n{ctx}"
                return "[No relevant content found in files]"
            return f"[MCP search error: HTTP {resp.status_code}]"

        elif tool_name == "read_file":
            path = tool_args.get("path", "")
            url = f"{MCP_API_URL}/file?path={_urlencode(path)}"
            resp = _requests.get(url, timeout=15)
            if resp.ok:
                data = resp.json()
                return data.get("content", "[Empty file]")
            return f"[MCP read error: HTTP {resp.status_code}]"

        elif tool_name == "list_files":
            resp = _requests.get(f"{MCP_API_URL}/files", timeout=10)
            if resp.ok:
                data = resp.json()
                files = data.get("files", [])
                if not files:
                    return "[No files indexed]"
                lines = [f"Folder: {data.get('folder', '?')} ({len(files)} files)"]
                for f in files:
                    lines.append(f"  {f['path']}  ({f['size_kb']} KB)")
                return "\n".join(lines)
            return f"[MCP list error: HTTP {resp.status_code}]"

        return f"[Unknown tool: {tool_name}]"

    except _requests.exceptions.ConnectionError:
        return "[MCP server offline -- run: python mcp_files.py --folder <path>]"
    except Exception as e:
        return f"[MCP tool error: {e}]"


def _urlencode(s: str) -> str:
    from urllib.parse import quote
    return quote(str(s), safe="")


def is_mcp_available() -> bool:
    """Quick health-check for mcp_files.py sidecar."""
    try:
        r = _requests.get(f"{MCP_API_URL}/", timeout=2)
        return r.ok
    except Exception:
        return False


# ---------------------------------------------------------------------------
# File parser
# ---------------------------------------------------------------------------
IMAGE_EXTS = {"png", "jpg", "jpeg", "gif", "webp", "bmp", "heic", "heif"}
IMAGE_MIME = {
    "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
    "gif": "image/gif", "webp": "image/webp", "bmp": "image/bmp",
    "heic": "image/heic", "heif": "image/heif",
}


def parse_uploaded_file(filename: str, content: bytes) -> str:
    """Legacy text-only parser. Returns extracted text or raises HTTPException.
    For images, use parse_uploaded_file_v2 which returns a dict with kind='image'.
    """
    ext = filename.lower().rsplit(".", 1)[-1]

    # All plain-text and code file types -- decoded as UTF-8
    TEXT_EXTS = {
        "txt", "md", "csv", "json", "xml", "html", "htm", "css",
        "js", "ts", "jsx", "tsx", "py", "sql", "yaml", "yml",
        "sh", "bash", "java", "cpp", "c", "h", "go", "rs", "rb",
        "php", "swift", "kt", "r", "toml", "ini", "env", "log",
        "conf", "cfg", "tf", "proto",
        "srt", "vtt", "ass", "ssa", "sub",  # subtitles
        "graphql", "gql", "vue", "svelte", "astro", "mdx",
        "tsv", "diff", "patch", "tex", "rst", "org",
    }

    if ext in TEXT_EXTS:
        return content.decode("utf-8", errors="replace")

    # Images are NOT handled here — caller should detect and route via v2 path.
    if ext in IMAGE_EXTS:
        raise HTTPException(400, "Image file — use /upload v2 path (returns inline data).")

    # Unknown extension: try UTF-8; if it decodes cleanly, accept it
    if ext not in {"pdf", "docx", "xlsx", "xls"}:
        try:
            return content.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(
                400,
                f"Cannot read .{ext} as text. Supported binary formats: PDF, DOCX, XLSX. "
                "Images: PNG, JPG, JPEG, GIF, WEBP, BMP, HEIC. "
                "For other files, make sure they are UTF-8 encoded text."
            )

    elif ext == "pdf":
        if not PDF_OK:
            raise HTTPException(400, "pdfplumber not installed. Run: pip install pdfplumber")
        text_parts = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    text_parts.append(t)
        return "\n".join(text_parts)

    elif ext == "docx":
        if not DOCX_OK:
            raise HTTPException(400, "python-docx not installed. Run: pip install python-docx")
        doc = docx.Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    elif ext in ("xlsx", "xls"):
        if not XLSX_OK:
            raise HTTPException(400, "openpyxl not installed. Run: pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                row_str = "\t".join(str(c) if c is not None else "" for c in row)
                if row_str.strip():
                    parts.append(row_str)
        return "\n".join(parts)

    elif ext == "csv":
        return content.decode("utf-8", errors="replace")

    else:
        raise HTTPException(400, f"Unsupported file type: .{ext}")


# ---------------------------------------------------------------------------
# File-output instruction — appended to every system prompt so AI can emit
# downloadable files inline using <file> tags.
# ---------------------------------------------------------------------------
FILE_OUTPUT_INSTRUCTION = """

## File Output

When the user asks you to generate a file (code, config, document, CSV, JSON,
script, etc.), wrap the file content in a `<file>` XML tag so the frontend can
offer it as a download:

```
<file name="example.py" mime="text/x-python">
print("hello world")
</file>
```

Rules:
- `name` = suggested filename (with extension).
- `mime` = MIME type (text/plain, application/json, text/csv, text/html,
  text/x-python, application/javascript, text/markdown, etc.).
- Content inside the tag is the raw file body — no extra markdown fences.
- You may emit multiple `<file>` blocks in one response.
- You can still include normal explanation text outside the tags.
- Only use `<file>` when the user wants a downloadable artifact; short inline
  code snippets shown for explanation do NOT need the tag.
"""


# ---------------------------------------------------------------------------
# Streaming helper
# ---------------------------------------------------------------------------
def chat_stream(
        prompt: str,
        model: str,
        system: str,
        temperature: float,
        max_tokens: int,
        history: list[dict],
        cancel_event: threading.Event,
        use_tools: bool = False,
        mcp_paths: str = "",
        images: list[dict] | None = None,
) -> Iterator[str]:
    """
    Agentic chat stream with optional MCP tool calling loop.

    When use_tools=True and the model supports tool calling:
      1. Send request with MCP tool definitions
      2. If model emits tool_use -> execute tool -> feed result back -> repeat
      3. Stream final text response to caller

    SSE protocol:
      data: <text chunk>          -- normal text token
      data: [TOOL_CALL:<json>]    -- model called a tool (UI can show it)
      data: [TOOL_RESULT:<json>]  -- tool result (UI can show it)
      data: [CANCELLED]           -- user cancelled
      data: [DONE]                -- stream finished
    """
    client = make_client(model)
    model_key = model
    model = MODELS.get(model, model)
    ceiling = MODEL_MAX_TOKENS.get(model, DEFAULT_MAX_TOKENS)
    max_tokens = min(max_tokens, ceiling)

    messages: list[dict] = [{"role": "system", "content": system + FILE_OUTPUT_INSTRUCTION}]
    messages.extend(history)

    # Build user message: if images present, use OpenAI-style multimodal content
    if images:
        # Pre-check: warn if selected model is not known to support vision via LZ.
        # This produces a clear in-chat warning instead of "I can't see images" from upstream.
        if model not in VISION_CAPABLE_MODELS and model_key not in VISION_CAPABLE_MODELS:
            yield (
                f"⚠ Model `{model_key}` is not in the vision-capable whitelist. "
                f"Images attached will likely be ignored.\n"
                f"Try: gpt-4o, claude-sonnet, gemini-2.5-flash, gemini-2.5-pro.\n\n"
            )
        content_parts: list[dict] = [{"type": "text", "text": prompt}]
        for img in images:
            b64 = img.get("b64", "")
            mime = img.get("mime", "image/png")
            if b64:
                content_parts.append({
                    "type": "image_url",
                    "image_url": {"url": f"data:{mime};base64,{b64}"}
                })
        messages.append({"role": "user", "content": content_parts})
        # Diagnostic: print which model is receiving the vision payload
        print(
            f"[VISION] model={model_key} → upstream={model} | "
            f"images={len(images)} | "
            f"sizes_kb=[{', '.join(str(len(i.get('b64', ''))*3//4//1024) for i in images)}]",
            flush=True,
        )
    else:
        messages.append({"role": "user", "content": prompt})

    total_chars = sum(
        len(m.get("content", "") if isinstance(m.get("content"), str) else "")
        for m in messages
    )
    if total_chars > 1600000:
        yield f"[WARNING: prompt is {total_chars:,} chars -- may exceed model context limit]\n"

    # Decide whether to use tool calling
    can_use_tools = (
            use_tools
            and model in TOOL_CAPABLE_MODELS
            and is_mcp_available()
    )

    # -- Opsi A/C: Agentic tool-calling loop ---------------------------------
    if can_use_tools:
        MAX_TOOL_ROUNDS = 6  # safety ceiling
        for _round in range(MAX_TOOL_ROUNDS):
            if cancel_event.is_set():
                yield "[CANCELLED]"
                return

            try:
                # Non-streaming for tool rounds (need full response to inspect)
                response = client.chat.completions.create(
                    model=model,
                    messages=messages,
                    temperature=temperature,
                    max_tokens=max_tokens,
                    tools=MCP_TOOLS,
                    tool_choice="auto",
                    stream=False,
                )
            except Exception as e:
                yield f"[ERROR: {e}]"
                return

            choice = response.choices[0]
            finish = choice.finish_reason
            msg = choice.message

            # -- Model finished -- stream the text --------------------------
            if finish in ("stop", "end_turn", "length") or not msg.tool_calls:
                content = msg.content or ""
                words = content.split(" ")
                for i, word in enumerate(words):
                    if cancel_event.is_set():
                        yield "[CANCELLED]"
                        return
                    yield (word + " ") if i < len(words) - 1 else word
                # Emit usage for cost tracking
                if hasattr(response, "usage") and response.usage:
                    import json as _json
                    yield f"[USAGE:{_json.dumps({'input': response.usage.prompt_tokens or 0, 'output': response.usage.completion_tokens or 0})}]"
                return

            # -- Model called tools -- execute each one ---------------------
            if msg.tool_calls:
                # Add assistant message with tool calls to history
                messages.append({
                    "role": "assistant",
                    "content": msg.content or "",
                    "tool_calls": [
                        {
                            "id": tc.id,
                            "type": "function",
                            "function": {"name": tc.function.name, "arguments": tc.function.arguments}
                        }
                        for tc in msg.tool_calls
                    ]
                })

                for tc in msg.tool_calls:
                    if cancel_event.is_set():
                        yield "[CANCELLED]"
                        return

                    fn_name = tc.function.name
                    try:
                        fn_args = json.loads(tc.function.arguments or "{}")
                    except Exception:
                        fn_args = {}

                    # Signal to UI that a tool is being called
                    yield f"[TOOL_CALL:{json.dumps({'tool': fn_name, 'args': fn_args})}]"

                    # Inject mcp_paths into search_files if provided
                    if fn_name == "search_files" and mcp_paths:
                        fn_args.setdefault("paths", mcp_paths)

                    result = execute_mcp_tool(fn_name, fn_args)

                    # Signal result to UI
                    yield f"[TOOL_RESULT:{json.dumps({'tool': fn_name, 'result': result[:500]})}]"

                    # Add tool result to messages
                    messages.append({
                        "role": "tool",
                        "tool_call_id": tc.id,
                        "name": fn_name,
                        "content": result,
                    })

                continue  # loop back -- model will now answer with tool results

        # Fallback if loop exhausted
        yield "[ERROR: Tool calling loop exceeded max rounds]"
        return

    # -- Standard streaming (no tools) ----------------------------------------
    try:
        stream = client.chat.completions.create(
            model=model,
            messages=messages,
            temperature=temperature,
            max_tokens=max_tokens,
            stream=True,
            stream_options={"include_usage": True},
        )
        _input_tokens = 0
        _output_tokens = 0
        _finish_reason = None
        for chunk in stream:
            if cancel_event.is_set():
                stream.close()
                yield "[CANCELLED]"
                return
            # Capture usage from final chunk
            if hasattr(chunk, "usage") and chunk.usage:
                _input_tokens = getattr(chunk.usage, "prompt_tokens", 0) or _input_tokens
                _output_tokens = getattr(chunk.usage, "completion_tokens", 0) or _output_tokens
            if not chunk.choices:
                continue
            choice = chunk.choices[0]
            if choice.finish_reason:
                _finish_reason = choice.finish_reason
            delta = choice.delta.content if choice.delta else None
            if delta:
                yield delta
        # Log finish reason
        print(f"[stream] model={model} finish_reason={_finish_reason} in={_input_tokens} out={_output_tokens} max_tokens={max_tokens}", flush=True)
        # Emit usage event for frontend cost tracking
        if _input_tokens or _output_tokens:
            import json as _json
            yield f"[USAGE:{_json.dumps({'input': _input_tokens, 'output': _output_tokens, 'finish': _finish_reason})}]"
    except Exception as api_err:
        yield f"[ERROR: {api_err}]"


# ---------------------------------------------------------------------------
# Conversation class — REMOVED (Phase 1 WS3 migration)
# History is now persisted in PostgreSQL (chat_sessions + chat_messages).
# The generate() async generator calls chat_stream() directly.
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Request models
# ---------------------------------------------------------------------------
class ChatRequest(BaseModel):
    session_id: str
    message: str
    model: str = "gemini-2.5-pro"
    system: str = "You are a helpful assistant."
    temperature: float = 1.0
    max_tokens: int = 8192
    use_tools: bool = False  # Opsi A/C: enable agentic MCP tool calling
    mcp_paths: str = ""  # comma-separated paths to restrict file search
    # NEW: optional inline images attached to this user turn (vision capable models)
    # each item: {"b64": "<base64>", "mime": "image/png", "name": "foo.png"}
    images: list[dict] = []

    @validator("message")
    def message_not_empty(cls, v):
        v = v.strip()
        if not v:
            raise ValueError("Message cannot be empty")
        if len(v) > 2000000:
            raise ValueError("Message too long (max 2,000,000 characters)")
        return v

    @validator("max_tokens")
    def cap_tokens(cls, v):
        return min(v, 100000)

    @validator("temperature")
    def clamp_temp(cls, v):
        return max(0.0, min(2.0, v))


class SaveRequest(BaseModel):
    session_id: str
    filename: str = ""


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@app.get("/")
async def root():
    return {"status": "ok"}


@app.get("/models")
async def list_models():
    return {"models": list(MODELS.keys())}


# -- Upload file -> returns extracted text ----------------------------------
@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    content = await file.read()
    if len(content) > 20 * 1024 * 1024:  # 20 MB limit
        raise HTTPException(400, "File too large (max 20 MB)")
    fname = file.filename or "upload"
    ext = fname.lower().rsplit(".", 1)[-1]

    # IMAGE PATH: return inline base64 + mime so chat can send as multimodal part
    if ext in IMAGE_EXTS:
        import base64 as _b64
        b64 = _b64.b64encode(content).decode("ascii")
        mime = IMAGE_MIME.get(ext, "image/png")
        return {
            "kind": "image",
            "filename": fname,
            "mime": mime,
            "b64": b64,
            "size_bytes": len(content),
            # legacy fields so old frontend doesn't break
            "chars": 0,
            "preview": f"[Image: {fname}, {len(content)} bytes]",
            "text": "",
        }

    # TEXT PATH (default): parse and return extracted text
    text = parse_uploaded_file(fname, content)
    preview = text[:300] + ("..." if len(text) > 300 else "")
    return {
        "kind": "text",
        "filename": fname,
        "chars": len(text),
        "preview": preview,
        "text": text,
    }


# -- Cancel a running stream -----------------------------------------------
@app.post("/cancel/{session_id}")
async def cancel_stream(session_id: str):
    await rc.set_cancel(session_id)
    # The running stream sees the flag on its next poll, wherever it runs.
    return {"status": "cancel_requested", "session_id": session_id}


# -- One-shot non-streaming chat (for auto-pick video feature) -------------
class OnceRequest(BaseModel):
    message: str
    model: str = "gemini-2.5-flash"
    system: str = "You are a helpful assistant."
    max_tokens: int = 12000  # high enough for thinking/reasoning models

@app.post("/chat/once")
async def chat_once(req: OnceRequest):
    FALLBACK_MODEL = "gemini-2.5-flash"

    def _try_model(model_name: str) -> str:
        c = make_client(model_name)
        # Try with system prompt
        r = c.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "system", "content": req.system},
                {"role": "user", "content": req.message},
            ],
            max_tokens=req.max_tokens,
            stream=False,
        )
        txt = (r.choices[0].message.content or "").strip()
        if txt:
            return txt
        # Retry with merged system+user (some models ignore system)
        r2 = c.chat.completions.create(
            model=model_name,
            messages=[
                {"role": "user", "content": f"{req.system}\n\n{req.message}"},
            ],
            max_tokens=req.max_tokens,
            stream=False,
        )
        return (r2.choices[0].message.content or "").strip()

    # Try requested model first
    text = _try_model(req.model)
    # If still empty and not already fallback model, try gemini-2.5-flash
    if not text and req.model != FALLBACK_MODEL:
        print(f"[chat/once] {req.model} returned empty, falling back to {FALLBACK_MODEL}")
        text = _try_model(FALLBACK_MODEL)
    return {"text": text}


# -- Main chat stream ------------------------------------------------------
def _to_uuid(s: str) -> str:
    """Convert any string to a deterministic UUID v5 (idempotent)."""
    try:
        uuid.UUID(s)
        return s  # already valid UUID
    except ValueError:
        return str(uuid.uuid5(uuid.NAMESPACE_DNS, s))

async def _resolve_user_uuid(tenant_id: str, clerk_user_id: str) -> str:
    """
    Convert a Clerk user ID (user_xxx) to the PostgreSQL UUID in the users table.
    Looks up by external_id WITH the RLS tenant context set (otherwise row-level
    security hides the row and the lookup always misses). If the user isn't found
    — e.g. the Clerk webhook hasn't fired in local dev — provision a minimal row
    just-in-time so chat_sessions/usage_logs foreign keys resolve.
    """
    if not clerk_user_id:
        return str(uuid.uuid5(uuid.NAMESPACE_DNS, "clerk-user-anon"))
    try:
        row = await db._q_fetchrow(
            "SELECT id FROM users WHERE tenant_id=$1 AND external_id=$2",
            db._uid(tenant_id), clerk_user_id, tenant=str(tenant_id)
        )
        if row:
            return str(row["id"])
        # Just-in-time provisioning (webhook likely didn't reach local dev)
        new_id = await db.upsert_user(tenant_id, clerk_user_id)
        if new_id:
            return str(new_id)
    except Exception as e:
        print(f"[_resolve_user_uuid] {e}", flush=True)
    # Last-resort fallback: derive deterministic UUID (FK guard will null it)
    return str(uuid.uuid5(uuid.NAMESPACE_DNS, f"clerk-user-{clerk_user_id}"))


@app.post("/chat/stream")
async def stream_chat(req: ChatRequest,
                      user: CurrentUser = Depends(get_current_user)):
    # ── Phase 1 WS3: tenant_id from JWT, user UUID resolved from DB ─────
    _TENANT_ID = user.tenant_id
    _USER_ID   = await _resolve_user_uuid(user.tenant_id, user.user_id)
    _session_id = _to_uuid(req.session_id)   # normalise any string → valid UUID

    # ── PostgreSQL session + history ──────────────────────────────────────
    try:
        await db.get_or_create_session(
            _TENANT_ID, _USER_ID, _session_id,
            req.model, req.system,
            temperature=req.temperature,
            max_tokens=req.max_tokens,
            use_tools=req.use_tools,
            mcp_paths=req.mcp_paths,
        )
        history = await db.get_session_history(_TENANT_ID, _session_id)
    except Exception as db_err:
        print(f"[stream_chat] DB session error: {db_err}", flush=True)
        raise HTTPException(status_code=503, detail="Database unavailable")

    # Build OpenAI-style history list (role + content only)
    history_msgs = [{"role": r["role"], "content": r["content"]} for r in history]

    # Local Event bridges the async loop → the sync chat_stream generator.
    # Cross-container cancel state lives in Redis (cancel:{session_id}).
    cancel_event = threading.Event()

    # Capture values for async DB writes inside the sync generator
    _tenant_id = _TENANT_ID
    _user_id   = _USER_ID
    _model = req.model

    async def generate():
        chunks: list[str] = []
        cancelled = False
        usage_data: dict = {}   # populated from [USAGE:{...}] chunk
        _SENTINEL = object()

        def _next_chunk(gen):
            """Wrap next() so StopIteration becomes sentinel — safe for executor."""
            try:
                return next(gen)
            except StopIteration:
                return _SENTINEL

        try:
            import asyncio as _asyncio
            loop = _asyncio.get_event_loop()
            # Call chat_stream directly — no in-memory Conversation wrapper
            sync_gen = chat_stream(
                prompt=req.message,
                model=req.model,
                system=req.system,
                temperature=req.temperature,
                max_tokens=req.max_tokens,
                history=history_msgs,
                cancel_event=cancel_event,
                use_tools=req.use_tools,
                mcp_paths=req.mcp_paths,
                images=req.images or [],
            )
            while True:
                # Poll cross-container cancel flag; signal the sync generator.
                if await rc.is_cancelled(_session_id):
                    cancel_event.set()
                chunk = await loop.run_in_executor(None, _next_chunk, sync_gen)
                if chunk is _SENTINEL:
                    break
                if chunk == "[CANCELLED]":
                    cancelled = True
                    yield "data: [CANCELLED]\n\n"
                    return
                if chunk.startswith("[ERROR"):
                    yield f"data: {chunk}\n\n"
                    yield "data: [DONE]\n\n"
                    return
                if chunk.startswith("[USAGE:"):
                    try:
                        # chunk is "[USAGE:{...}]" — strip 7-char prefix AND trailing ]
                        usage_data.update(json.loads(chunk[7:].rstrip("]")))
                    except Exception:
                        pass
                    yield f"data: {chunk}\n\n"
                    continue
                chunks.append(chunk)
                # SSE: encode newlines as \n so the JS side can restore them
                encoded = chunk.replace("\\", "\\\\").replace("\n", "\\n")
                yield f"data: {encoded}\n\n"
        except Exception as e:
            yield f"data: [ERROR: {e}]\n\n"
        finally:
            await rc.clear_cancel(_session_id)
            if not cancelled and chunks:
                # Persist user turn + assistant reply to PostgreSQL
                stored_user = req.message
                if req.images:
                    stored_user += (
                        f"\n\n[Attached {len(req.images)} image(s): "
                        + ", ".join(i.get("name", "img") for i in req.images)
                        + "]"
                    )
                reply = "".join(chunks)
                tok_in  = int(usage_data.get("input",  0))
                tok_out = int(usage_data.get("output", 0))
                cost    = _calc_cost(_model, tok_in, tok_out)
                # provider must satisfy usage_logs CHECK: laozhang|deepseek|gemini|openai|other
                _ml = _model.lower()
                if   _ml.startswith("gemini"):   _provider = "gemini"
                elif _ml.startswith("deepseek"): _provider = "deepseek"
                elif _ml.startswith(("gpt", "o3", "o1")): _provider = "openai"
                else: _provider = "laozhang"
                try:
                    await db.append_message(
                        _tenant_id, _session_id, "user", stored_user, _model)
                    await db.append_message(
                        _tenant_id, _session_id, "assistant", reply, _model,
                        tokens_in=tok_in, tokens_out=tok_out, cost_usd=cost)
                    # endpoint must be one of: chat|image|tts|video|embedding|batch|other
                    await db.log_usage(
                        _tenant_id, _user_id, _model, "chat",
                        tok_in, tok_out, cost,
                        session_id=str(_session_id), provider=_provider)
                except Exception as db_err:
                    print(f"[stream_chat] DB append/usage error: {db_err}", flush=True)
        yield "data: [DONE]\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@app.get("/history/{session_id}")
async def get_history(session_id: str,
                      user: CurrentUser = Depends(get_current_user)):
    _TENANT_ID = user.tenant_id
    try:
        history = await db.get_session_history(_TENANT_ID, _to_uuid(session_id))
    except Exception as db_err:
        print(f"[get_history] DB error: {db_err}", flush=True)
        raise HTTPException(status_code=503, detail="Database unavailable")
    if history is None:
        raise HTTPException(status_code=404, detail="Session not found")
    return {"history": history}


@app.post("/save")
async def save_conversation(req: SaveRequest,
                            user: CurrentUser = Depends(get_current_user)):
    _TENANT_ID = user.tenant_id
    _session_id = _to_uuid(req.session_id)
    try:
        history = await db.get_session_history(_TENANT_ID, _session_id)
    except Exception as db_err:
        print(f"[save] DB error: {db_err}", flush=True)
        raise HTTPException(status_code=503, detail="Database unavailable")
    if not history:
        raise HTTPException(status_code=404, detail="Session not found")

    filename = req.filename or f"conversation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"

    with open(filename, "w", encoding="utf-8") as f:
        f.write(f"Session: {req.session_id}\n")
        f.write("=" * 50 + "\n")
        for msg in history:
            role = "You" if msg["role"] == "user" else "AI"
            f.write(f"{role}: {msg['content']}\n\n")

    return {"saved": filename}


@app.delete("/session/{session_id}")
async def clear_session(session_id: str,
                        user: CurrentUser = Depends(get_current_user)):
    tid = user.tenant_id
    sid = _to_uuid(session_id)
    try:
        await db.delete_session(tid, sid)
    except Exception as db_err:
        print(f"[clear_session] DB error: {db_err}", flush=True)
        raise HTTPException(status_code=503, detail="Database unavailable")
    return {"status": "cleared", "session_id": session_id}


@app.delete("/sessions")
async def clear_all(user: CurrentUser = Depends(get_current_user)):
    tid = user.tenant_id
    try:
        await db._q_exec(
            "DELETE FROM chat_sessions WHERE tenant_id=$1", db._uid(tid))
    except Exception as db_err:
        print(f"[clear_all] DB error: {db_err}", flush=True)
        raise HTTPException(status_code=503, detail="Database unavailable")
    return {"status": "all sessions cleared"}


# ---------------------------------------------------------------------------
# MCP status endpoint
# ---------------------------------------------------------------------------
@app.get("/mcp/status")
async def mcp_status():
    """Check if mcp_files.py sidecar is running and return its info."""
    available = is_mcp_available()
    if not available:
        return {"available": False, "url": MCP_API_URL}
    try:
        r = _requests.get(f"{MCP_API_URL}/", timeout=3)
        info = r.json()
        return {"available": True, "url": MCP_API_URL, **info}
    except Exception as e:
        return {"available": False, "url": MCP_API_URL, "error": str(e)}


@app.get("/mcp/tool-capable-models")
async def tool_capable_models():
    """Return list of models that support agentic tool calling."""
    return {
        "tool_capable": [k for k, v in MODELS.items() if v in TOOL_CAPABLE_MODELS],
        "total_models": len(MODELS),
    }


# ---------------------------------------------------------------------------
# Image generation helpers
# ---------------------------------------------------------------------------
import base64 as _b64

# -----------------------------------------------------------------------------
# Image model registry
# api:
#   "chat-image-b64"  -> /v1/chat/completions, response has base64 in markdown
#   "chat-image-url"  -> /v1/chat/completions, response has image URL in markdown
#   "google"          -> /v1beta/models/{model}:generateContent (native Google)
#   "openai-image"    -> /v1/images/generations (OpenAI images endpoint)
#
# token_group: the token group required on laozhang.ai (empty = default group)
# -----------------------------------------------------------------------------
IMAGE_MODELS = {
    # -----------------------------------------------------------------------
    # SORA IMAGE
    # endpoint : /v1/chat/completions
    # response : image URL in markdown  ->  we download + return base64
    # ratios   : append 【x:x】 to prompt
    # token    : default group
    # -----------------------------------------------------------------------
    "sora-image": {
        "api": "chat-image-url", "model": "sora_image",
        "price": "$0.01/img", "token_group": "default",
        "ratios": ["1:1", "2:3", "3:2"],
        "resolutions": [],
    },
    # -----------------------------------------------------------------------
    # GPT-IMAGE-2  (three variants -- different token groups / model names)
    # endpoint : /v1/images/generations  (OpenAI images API)
    # token    : default group  ->  $0.03/call, no size/quality params
    # -----------------------------------------------------------------------
    "gpt-image-2": {
        "api": "openai-image", "model": "gpt-image-2",
        "price": "$0.03/call", "token_group": "default",
        "ratios": ["1:1"],
        "resolutions": [],
        "extra_params": {},  # no size, no quality
    },
    # default group, reverse Codex route -- supports explicit pixel sizes
    "gpt-image-2-vip": {
        "api": "openai-image", "model": "gpt-image-2-vip",
        "price": "$0.03/call", "token_group": "default",
        "ratios": ["1:1", "16:9", "9:16", "4:3", "3:4", "2:3", "3:2", "21:9", "4:5", "5:4"],
        "resolutions": ["1K", "2K", "4K"],
        "size_map_vip": {
            # ratio -> {1K, 2K, 4K}
            "1:1": {"1K": "1280x1280", "2K": "2048x2048", "4K": "2880x2880"},
            "16:9": {"1K": "1280x720", "2K": "2048x1152", "4K": "3840x2160"},
            "9:16": {"1K": "720x1280", "2K": "1152x2048", "4K": "2160x3840"},
            "4:3": {"1K": "1280x960", "2K": "2048x1536", "4K": "3312x2480"},
            "3:4": {"1K": "960x1280", "2K": "1536x2048", "4K": "2480x3312"},
            "2:3": {"1K": "848x1280", "2K": "1360x2048", "4K": "2336x3520"},
            "3:2": {"1K": "1280x848", "2K": "2048x1360", "4K": "3520x2336"},
            "21:9": {"1K": "1280x544", "2K": "2048x864", "4K": "3840x1632"},
            "4:5": {"1K": "1024x1280", "2K": "1632x2048", "4K": "2560x3216"},
            "5:4": {"1K": "1280x1024", "2K": "2048x1632", "4K": "3216x2560"},
        },
    },
    # Sora2Official token group -- official API transit, full params
    "gpt-image-2-official": {
        "api": "openai-image", "model": "gpt-image-2",
        "price": "official pricing", "token_group": "Sora2Official",
        "ratios": ["1:1", "16:9", "9:16"],
        "resolutions": [],
        "extra_params": {"quality": "auto"},  # supports quality param
    },
    # -----------------------------------------------------------------------
    # NANO BANANA (Standard)
    # Old model gemini-2.5-flash-image-preview = 1:1/1K only via chat/completions
    # New official gemini-2.5-flash-image = 10 ratios + 1K via Google native
    # Both $0.025/img, default group, pay-per-use token required
    # -----------------------------------------------------------------------
    "nano-banana": {
        "api": "chat-image-b64", "model": "gemini-2.5-flash-image",
        "price": "$0.025/img", "token_group": "default",
        "ratios": ["1:1"],
        "resolutions": ["1K"],
    },
    "nano-banana-hd": {
        "api": "google", "model": "gemini-2.5-flash-image",
        "price": "$0.025/img", "token_group": "default",
        "ratios": ["1:1", "16:9", "9:16", "4:3", "3:4", "21:9", "3:2", "2:3", "5:4", "4:5"],
        "resolutions": ["1K"],
    },
    # -----------------------------------------------------------------------
    # NANO BANANA 2  (Gemini 3.1 Flash)
    # mode A: /v1/chat/completions  ->  1:1, 1K, base64
    # mode B: /v1beta/models/...:generateContent  ->  10 ratios, 1K/2K/4K
    # token    : default group, pay-per-use token required
    # -----------------------------------------------------------------------
    "nano-banana-2": {
        "api": "chat-image-b64", "model": "gemini-3.1-flash-image-preview",
        "price": "$0.055/img", "token_group": "default",
        "ratios": ["1:1"],
        "resolutions": ["1K"],
    },
    "nano-banana-2-hd": {
        "api": "google", "model": "gemini-3.1-flash-image-preview",
        "price": "$0.055/img", "token_group": "default",
        "ratios": ["1:1", "16:9", "9:16", "4:3", "3:4", "21:9", "3:2", "2:3", "5:4", "4:5"],
        "resolutions": ["1K", "2K", "4K"],
    },
    # -----------------------------------------------------------------------
    # NANO BANANA PRO  (Gemini 3 Pro)
    # mode A: /v1/chat/completions  ->  1:1, 1K, base64
    # mode B: /v1beta/models/...:generateContent  ->  10 ratios, 1K/2K/4K
    # token    : default group, pay-per-use token required
    # -----------------------------------------------------------------------
    "nano-banana-pro": {
        "api": "chat-image-b64", "model": "gemini-3-pro-image-preview",
        "price": "$0.09/img", "token_group": "default",
        "ratios": ["1:1"],
        "resolutions": ["1K"],
    },
    "nano-banana-pro-hd": {
        "api": "google", "model": "gemini-3-pro-image-preview",
        "price": "$0.09/img", "token_group": "default",
        "ratios": ["1:1", "16:9", "9:16", "4:3", "3:4", "21:9", "3:2", "2:3", "5:4", "4:5"],
        "resolutions": ["1K", "2K", "4K"],
    },
    # -----------------------------------------------------------------------
    # FLUX KONTEXT  (images/generations)
    # token    : default group
    # -----------------------------------------------------------------------
    "flux-kontext-pro": {
        "api": "openai-image", "model": "flux-kontext-pro",
        "price": "$0.035/img", "token_group": "default",
        "ratios": ["1:1", "16:9", "9:16", "4:3", "3:4", "3:7", "7:3"],
        "resolutions": [],
    },
    "flux-kontext-max": {
        "api": "openai-image", "model": "flux-kontext-max",
        "price": "$0.07/img", "token_group": "default",
        "ratios": ["1:1", "16:9", "9:16", "4:3", "3:4", "3:7", "7:3"],
        "resolutions": [],
    },
    # -----------------------------------------------------------------------
    # GPT-IMAGE-1  (images/generations, token-based billing)
    # token    : default group
    # -----------------------------------------------------------------------
    "gpt-image-1": {
        "api": "openai-image", "model": "gpt-image-1",
        "price": "per token", "token_group": "default",
        "ratios": ["1:1", "16:9", "9:16"],
        "resolutions": [],
    },
    # -----------------------------------------------------------------------
    # SEEDREAM  (images/generations, returns URL)
    # token    : default group
    # -----------------------------------------------------------------------
    # Seedream: size is always "2K" string, response_format="url", no n param
    "seedream-4-0": {
        "api": "seedream", "model": "seedream-4-0-250828",
        "price": "$0.035/img", "token_group": "default",
        "ratios": ["1:1", "16:9", "9:16", "4:3", "3:4"],
        "resolutions": [],
    },
    "seedream-4-5": {
        "api": "seedream", "model": "seedream-4-5-251128",
        "price": "$0.045/img", "token_group": "default",
        "ratios": ["1:1", "16:9", "9:16", "4:3", "3:4"],
        "resolutions": [],
    },
}


# -----------------------------------------------------------------------------
# Image generation helpers
# -----------------------------------------------------------------------------

def _img_headers(key: str | None = None):
    return {"Authorization": f"Bearer {key or IMAGE_API_KEY}",
            "Content-Type": "application/json"}


def _generate_chat_image(prompt: str, model: str, aspect_ratio: str,
                         ref_b64: str = "", returns_url: bool = False,
                         key: str | None = None) -> str:
    """
    /v1/chat/completions -- Sora Image (URL) and Nano Banana family (base64).
    Sora uses 【x:x】 ratio markers; Nano Banana is 1:1 only.
    """
    import re
    full_prompt = prompt
    if returns_url and aspect_ratio in ("2:3", "3:2", "1:1"):
        full_prompt = f"{prompt}【{aspect_ratio}】"

    if ref_b64 and not returns_url:
        content = [
            {"type": "text", "text": full_prompt},
            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{ref_b64}"}},
        ]
    else:
        content = full_prompt

    payload = {"model": model, "stream": False,
               "messages": [{"role": "user", "content": content}]}
    r = _requests.post(f"{IMAGE_URL}/chat/completions",
                       headers=_img_headers(key), json=payload, timeout=180)
    r.raise_for_status()
    body = r.json()["choices"][0]["message"]["content"]

    if returns_url:
        m = re.search(r'!\[.*?\]\((https?://[^)]+)\)', body)
        if not m:
            raise HTTPException(500, f"No image URL in response: {body[:300]}")
        img_r = _requests.get(m.group(1), timeout=60)
        img_r.raise_for_status()
        return _b64.b64encode(img_r.content).decode()

    m = re.search(r'data:image/[^;]+;base64,([A-Za-z0-9+/=]+)', body)
    if m:
        return m.group(1)
    m = re.search(r'!\[.*?\]\(data:image/[^;]+;base64,([A-Za-z0-9+/=]+)\)', body)
    if m:
        return m.group(1)
    raise HTTPException(500, f"No base64 image in response: {body[:300]}")


def _generate_google(prompt: str, model: str, aspect_ratio: str,
                     image_size: str, ref_b64: str = "",
                     key: str | None = None) -> str:
    """/v1beta/models/{model}:generateContent -- Nano Banana 2/Pro HD mode."""
    url = f"{GOOGLE_IMAGE_BASE}/{model}:generateContent"
    parts: list = [{"text": prompt}]
    if ref_b64:
        parts.append({"inline_data": {"mime_type": "image/png", "data": ref_b64}})
    payload = {
        "contents": [{"parts": parts}],
        "generationConfig": {
            "responseModalities": ["IMAGE"],
            "imageConfig": {"aspectRatio": aspect_ratio, "imageSize": image_size},
        },
    }
    r = _requests.post(url, headers=_img_headers(key), json=payload, timeout=180)
    r.raise_for_status()
    return r.json()["candidates"][0]["content"]["parts"][0]["inlineData"]["data"]


def _generate_openai_image(prompt: str, model: str, aspect_ratio: str,
                           image_size: str, ref_b64: str = "",
                           extra_params: dict | None = None,
                           size_map_vip: dict | None = None,
                           returns_url: bool = False,
                           key: str | None = None) -> str:
    """/v1/images/generations -- Flux, GPT-Image-1, GPT-Image-2, Seedream."""
    # Standard aspect-ratio -> pixel size map
    std_size_map = {
        "1:1": "1024x1024", "16:9": "1792x1024", "9:16": "1024x1792",
        "4:3": "1365x1024", "3:4": "1024x1365",
        "3:2": "1248x832", "2:3": "832x1248",
        "21:9": "1584x672", "3:7": "832x1904", "7:3": "1904x832",
        "5:4": "1152x896", "4:5": "896x1152",
    }
    payload: dict = {"model": model, "prompt": prompt, "n": 1}

    if size_map_vip:
        # gpt-image-2-vip: use exact pixel dimensions from the VIP size map
        res = image_size if image_size in ("1K", "2K", "4K") else "2K"
        size = size_map_vip.get(aspect_ratio, {}).get(res, "2048x2048")
        payload["size"] = size
    elif model not in ("gpt-image-2",):
        # For most models pass a size; skip for plain gpt-image-2 default
        payload["size"] = std_size_map.get(aspect_ratio, "1024x1024")

    payload["response_format"] = "url" if returns_url else "b64_json"

    if extra_params:
        payload.update(extra_params)

    r = _requests.post(f"{IMAGE_URL}/images/generations",
                       headers=_img_headers(key), json=payload, timeout=180)
    r.raise_for_status()
    item = r.json()["data"][0]

    if returns_url:
        img_r = _requests.get(item["url"], timeout=60)
        img_r.raise_for_status()
        return _b64.b64encode(img_r.content).decode()

    if item.get("b64_json"):
        val = item["b64_json"]
        if val.startswith("data:"):
            val = val.split(",", 1)[1]
        val += "=" * ((4 - len(val) % 4) % 4)
        return val

    # Fallback: URL returned
    img_r = _requests.get(item["url"], timeout=60)
    img_r.raise_for_status()
    return _b64.b64encode(img_r.content).decode()


class ImageRequest(BaseModel):
    prompt: str
    model: str = "nano-banana"
    aspect_ratio: str = "1:1"
    image_size: str = "1K"
    ref_image: str = ""
    nusantara_corpus: bool = False

    @validator("model")
    def valid_model(cls, v):
        if v not in IMAGE_MODELS:
            raise ValueError(f"Unknown image model: {v}. Available: {list(IMAGE_MODELS.keys())}")
        return v


class ImageResponse(BaseModel):
    image_b64: str
    model: str
    width: int = 0
    height: int = 0


def _img_headers(key: str | None = None):
    return {
        "Authorization": f"Bearer {key or IMAGE_API_KEY}",
        "Content-Type": "application/json",
    }


@app.get("/image-models")
async def list_image_models():
    return {"models": list(IMAGE_MODELS.keys())}


def _generate_seedream(prompt: str, model: str, ref_b64: str = "") -> str:
    """
    Seedream via /v1/images/generations.
    size must be "2K" string. response_format="url". No n param. No pixel dimensions.
    """
    payload = {
        "model": model,
        "prompt": prompt,
        "response_format": "url",
        "size": "2K",
        "watermark": False,
        "sequential_image_generation": "disabled",
        "stream": False,
    }
    if ref_b64:
        payload["image"] = f"data:image/png;base64,{ref_b64}"

    r = _requests.post(f"{IMAGE_URL}/images/generations",
                       headers=_img_headers(IMAGE_API_KEY), json=payload, timeout=180)
    r.raise_for_status()
    image_url = r.json()["data"][0]["url"]
    img_r = _requests.get(image_url, timeout=60)
    img_r.raise_for_status()
    return _b64.b64encode(img_r.content).decode()


class VertexImageRequest(BaseModel):
    prompt: str
    model: str = "imagegeneration@006"
    aspect_ratio: str = "1:1"
    nusantara_corpus: bool = False

@app.post("/generate-image/vertex")
async def generate_image_vertex(req: VertexImageRequest):
    if not _ensure_vertex():
        raise HTTPException(503, "Vertex AI not configured — set GCP_PROJECT_ID, GCP_REFRESH_TOKEN, GCP_CLIENT_ID, GCP_CLIENT_SECRET")
    prompt = req.prompt
    if req.nusantara_corpus:
        prompt, _, _ = _nc.enhance_prompt(
            prompt,
            gemini_api_key=GEMINI_API_KEY or None,
            qdrant_url=QDRANT_CLOUD_URL or None,
            qdrant_api_key=QDRANT_CLOUD_KEY or None,
        )
    # ── Nano Banana (gemini-*-image) → Gemini API on Vertex, not ImageGenerationModel ──
    if _is_gemini_image_model(req.model):
        try:
            from google import genai as _genai
            from google.genai import types as _gtypes
            client = _genai.Client(
                vertexai=True,
                project=GCP_PROJECT_ID,
                location=GCP_LOCATION,
                credentials=_gcp_creds,
            )
            resp = client.models.generate_content(
                model=req.model,
                contents=prompt,
                config=_gtypes.GenerateContentConfig(response_modalities=["IMAGE"]),
            )
            img_bytes = None
            for cand in (resp.candidates or []):
                for part in (getattr(cand.content, "parts", None) or []):
                    inline = getattr(part, "inline_data", None)
                    if inline and getattr(inline, "data", None):
                        img_bytes = inline.data
                        break
                if img_bytes:
                    break
            if not img_bytes:
                raise HTTPException(502, f"{req.model} returned no image (check model availability in project/location {GCP_LOCATION})")
            b64 = _b64.b64encode(img_bytes).decode()
            return {"image_b64": b64, "model": req.model}
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(500, str(e))

    # ── Imagen (imagegeneration@*, imagen-*) → ImageGenerationModel ──
    try:
        from vertexai.preview.vision_models import ImageGenerationModel as _IGen
        import io as _io
        mdl = _IGen.from_pretrained(req.model)
        images = mdl.generate_images(
            prompt=prompt,
            number_of_images=1,
            aspect_ratio=req.aspect_ratio,
        )
        buf = _io.BytesIO()
        images[0]._pil_image.save(buf, format="JPEG", quality=92)
        b64 = _b64.b64encode(buf.getvalue()).decode()
        return {"image_b64": b64, "model": req.model}
    except Exception as e:
        raise HTTPException(500, str(e))


class EnhancePromptRequest(BaseModel):
    prompt: str

@app.post("/enhance-prompt")
async def enhance_prompt_endpoint(req: EnhancePromptRequest):
    enhanced, hits, ref_b64 = _nc.enhance_prompt(
        req.prompt,
        gemini_api_key=GEMINI_API_KEY or None,
        qdrant_url=QDRANT_CLOUD_URL or None,
        qdrant_api_key=QDRANT_CLOUD_KEY or None,
    )
    return {"enhanced_prompt": enhanced, "ref_b64": ref_b64 or "", "hits": len(hits)}


@app.post("/generate-image")
async def generate_image(req: ImageRequest,
                         x_image_api_key: str = Header(None, alias="X-Image-API-Key")):
    cfg = IMAGE_MODELS.get(req.model)
    if not cfg:
        raise HTTPException(400, f"Unknown image model: {req.model}")

    # Always use IMAGE_API_KEY from env (LAOZHANG_IMAGE_API_KEY)
    key = IMAGE_API_KEY

    # Nusantara corpus: enrich prompt + optionally supply ref image for conditioning
    _nc_ref_b64: str | None = None
    if req.nusantara_corpus:
        req.prompt, _nc_hits, _nc_ref_b64 = _nc.enhance_prompt(
            req.prompt,
            gemini_api_key=GEMINI_API_KEY or None,
            qdrant_url=QDRANT_CLOUD_URL or None,
            qdrant_api_key=QDRANT_CLOUD_KEY or None,
        )
        # Use corpus ref image for conditioning only when caller provided none
        if _nc_ref_b64 and not req.ref_image:
            req.ref_image = _nc_ref_b64

    try:
        api = cfg["api"]
        mdl = cfg["model"]
        ep = cfg.get("extra_params") or {}
        smap = cfg.get("size_map_vip")

        if api == "chat-image-url":
            b64 = _generate_chat_image(req.prompt, mdl, req.aspect_ratio,
                                       req.ref_image, returns_url=True, key=key)
        elif api == "chat-image-b64":
            b64 = _generate_chat_image(req.prompt, mdl, req.aspect_ratio,
                                       req.ref_image, returns_url=False, key=key)
        elif api == "google":
            b64 = _generate_google(req.prompt, mdl, req.aspect_ratio,
                                   req.image_size, req.ref_image, key=key)
        elif api == "seedream":
            b64 = _generate_seedream(req.prompt, mdl, req.ref_image)
        elif api in ("openai-image", "openai-image-url"):
            b64 = _generate_openai_image(
                req.prompt, mdl, req.aspect_ratio, req.image_size,
                req.ref_image, extra_params=ep, size_map_vip=smap,
                returns_url=(api == "openai-image-url"), key=key,
            )
        else:
            raise HTTPException(400, f"Unknown API type: {api}")

        return {
            "image_b64": b64,
            "model": req.model,
            "api_type": api,
            "token_group": cfg.get("token_group", "default"),
        }

    except _requests.HTTPError as e:
        raise HTTPException(e.response.status_code,
                            f"API error: {e.response.text[:400]}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        pass  # IMAGE_API_KEY from env, no restore needed


# VEO 3.1 ROUTES
# ==================================================================

VEO_BASE_URL = "https://api.laozhang.ai"
VEO_API_URL = f"{VEO_BASE_URL}/v1/videos"

VEO_PRESETS = {
    "720p_landscape": {"seconds": "8", "size": "1280x720", "resolution": "720p", "aspectRatio": "16:9"},
    "1080p_landscape": {"seconds": "8", "size": "1920x1080", "resolution": "1080p", "aspectRatio": "16:9"},
    "1080p_portrait": {"seconds": "8", "size": "1080x1920", "resolution": "1080p", "aspectRatio": "9:16"},
    "4k_landscape": {"seconds": "8", "size": "3840x2160", "resolution": "4k", "aspectRatio": "16:9"},
}


def _veo_headers(override_key: Optional[str] = None) -> dict:
    key = IMAGE_API_KEY or API_KEY  # use image key (LAOZHANG_IMAGE_API_KEY) same as working script
    return {"Authorization": f"Bearer {key}"}


def _veo_metadata(preset: dict) -> str:
    return json.dumps({
        "durationSeconds": int(preset["seconds"]),
        "resolution": preset["resolution"],
        "aspectRatio": preset["aspectRatio"],
    })


class VeoSubmitRequest(BaseModel):
    prompt: str
    model: str = "veo-3.1-generate-preview"
    preset: dict = None  # VEO_PRESETS value dict
    negative_prompt: str = "blurry, watermark, distorted, low quality"
    seed: str = ""
    ref_image_b64: str = ""  # base64-encoded reference image
    ref_image_mime: str = "image/jpeg"


@app.post("/veo/submit")
async def veo_submit(req: VeoSubmitRequest, x_veo_api_key: Optional[str] = Header(default=None)):
    """Submit a Veo 3.1 image-to-video or text-to-video task."""
    preset = req.preset or VEO_PRESETS["1080p_landscape"]
    headers = _veo_headers(x_veo_api_key)

    fields = {
        "model": req.model,
        "prompt": req.prompt,
        "seconds": preset["seconds"],
        "duration": preset["seconds"],
        "size": preset["size"],
        "resolution": preset["resolution"],
        "aspectRatio": preset["aspectRatio"],
        "metadata": _veo_metadata(preset),
        "negativePrompt": req.negative_prompt,
    }
    if req.seed:
        fields["seed"] = req.seed

    files = {k: (None, v) for k, v in fields.items()}

    if req.ref_image_b64:
        # Decode base64 -> bytes for multipart upload
        img_bytes = base64.b64decode(req.ref_image_b64)
        ext_map = {"image/jpeg": "reference.jpg", "image/png": "reference.png", "image/webp": "reference.webp"}
        fname = ext_map.get(req.ref_image_mime, "reference.jpg")
        files["input_reference"] = (fname, img_bytes, req.ref_image_mime)

    try:
        res = _requests.post(VEO_API_URL, headers=headers, files=files, timeout=180)
        res.raise_for_status()
        data = res.json()
        task_id = data.get("id") or data.get("task_id")
        return {"task_id": task_id, "status": data.get("status", "queued"), "raw": data}
    except _requests.HTTPError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/veo/status/{task_id}")
async def veo_status(task_id: str, x_veo_api_key: Optional[str] = Header(default=None)):
    """Poll Veo task status."""
    headers = _veo_headers(x_veo_api_key)
    try:
        res = _requests.get(f"{VEO_API_URL}/{task_id}", headers=headers, timeout=60)
        res.raise_for_status()
        data = res.json()
        return {
            "task_id": task_id,
            "status": data.get("status", "unknown"),
            "progress": data.get("progress", 0),
            "raw": data,
        }
    except _requests.HTTPError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/veo/download/{task_id}")
async def veo_download(task_id: str, x_veo_api_key: Optional[str] = Header(default=None)):
    """
    Return a JSON {url} pointing to the stream endpoint.
    Browser uses this URL as <video src> -- no large payload in this response.
    """
    # Build stream URL -- server.js will proxy it
    return {"url": f"/api/veo/stream/{task_id}"}


VEO_OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Veo_outputs")
os.makedirs(VEO_OUTPUT_DIR, exist_ok=True)


@app.get("/veo/stream/{task_id}")
async def veo_stream(task_id: str, x_veo_api_key: Optional[str] = Header(default=None)):
    """
    Proxy MP4 bytes from laozhang /v1/videos/{id}/content.
    Saves a copy to Veo_outputs/{task_id}.mp4 before streaming to browser.
    Retries up to 6x if content is still IN_PROGRESS.
    """
    from fastapi.responses import Response as FResponse
    import time as _time

    headers = _veo_headers(x_veo_api_key)
    last_err = None
    content_url = f"{VEO_API_URL}/{task_id}/content"
    safe_id = task_id.replace("/", "_").replace("\\", "_")
    save_path = os.path.join(VEO_OUTPUT_DIR, f"{safe_id}.mp4")

    # Return cached file immediately if already saved
    if os.path.exists(save_path) and os.path.getsize(save_path) > 1000:
        print(f"[veo/stream] ✓ Serving cached: {save_path}")
        with open(save_path, "rb") as f:
            return FResponse(
                content=f.read(),
                media_type="video/mp4",
                headers={
                    "Content-Disposition": f'inline; filename="{safe_id[:24]}.mp4"',
                    "Cache-Control": "no-store",
                    "X-Veo-Cached": "true",
                },
            )

    print(f"[veo/stream] Fetching: {content_url}")

    for attempt in range(6):
        try:
            res = _requests.get(content_url, headers=headers, timeout=180)
            print(f"[veo/stream] attempt={attempt + 1} status={res.status_code} "
                  f"content-type={res.headers.get('content-type', '')} "
                  f"size={len(res.content)} bytes")

            if res.status_code == 200 and len(res.content) > 1000:
                # -- Save to Veo_outputs/ ----------------------------------
                with open(save_path, "wb") as f:
                    f.write(res.content)
                size_mb = len(res.content) / 1_048_576
                print(f"[veo/stream] ✓ Saved {size_mb:.1f} MB -> {save_path}")

                return FResponse(
                    content=res.content,
                    media_type="video/mp4",
                    headers={
                        "Content-Disposition": f'inline; filename="{safe_id[:24]}.mp4"',
                        "Cache-Control": "no-store",
                        "X-Veo-Saved-Path": save_path,
                    },
                )

            last_err = res.text[:500]
            print(f"[veo/stream] Not ready: {last_err}")

            if "IN_PROGRESS" in res.text or "in_progress" in res.text or res.status_code == 404:
                _time.sleep(12)
            else:
                raise HTTPException(status_code=res.status_code, detail=res.text[:300])

        except HTTPException:
            raise
        except Exception as e:
            last_err = str(e)
            print(f"[veo/stream] Exception attempt {attempt + 1}: {e}")
            _time.sleep(12)

    raise HTTPException(status_code=503, detail=f"Video not available after retries. Last error: {last_err}")


# ==================================================================
# SORA 2 ROUTES (Official Forward -- Sora2Official group)
# ==================================================================

SORA_API_URL = "https://api.laozhang.ai/v1/videos"
SORA_OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Sora_outputs")
os.makedirs(SORA_OUTPUT_DIR, exist_ok=True)


def _sora_headers(override_key: Optional[str] = None) -> dict:
    key = override_key or API_KEY
    return {"Authorization": f"Bearer {key}"}


class SoraSubmitRequest(BaseModel):
    prompt: str
    model: str = "sora-2"  # sora-2 | sora-2-pro
    size: str = "1280x720"  # see docs for valid combos
    seconds: str = "8"  # "4" | "8" | "12"
    ref_image_b64: str = ""
    ref_image_mime: str = "image/jpeg"


@app.post("/sora/submit")
async def sora_submit(req: SoraSubmitRequest, x_sora_api_key: Optional[str] = Header(default=None)):
    """Submit a Sora 2 text-to-video or image-to-video task."""
    headers = _sora_headers(x_sora_api_key)

    form_data = {
        "model": req.model,
        "prompt": req.prompt,
        "size": req.size,
        "seconds": req.seconds,
    }

    files = {k: (None, v) for k, v in form_data.items()}

    if req.ref_image_b64:
        img_bytes = base64.b64decode(req.ref_image_b64)
        ext_map = {"image/jpeg": "reference.jpg", "image/png": "reference.png", "image/webp": "reference.webp"}
        fname = ext_map.get(req.ref_image_mime, "reference.jpg")
        files["input_reference"] = (fname, img_bytes, req.ref_image_mime)
        print(f"[sora/submit] image-to-video: {fname} ({len(img_bytes)} bytes)")

    print(f"[sora/submit] model={req.model} size={req.size} seconds={req.seconds}s")

    try:
        res = _requests.post(SORA_API_URL, headers=headers, files=files, timeout=120)
        res.raise_for_status()
        data = res.json()
        task_id = data.get("id") or data.get("task_id")
        print(f"[sora/submit] task_id={task_id} status={data.get('status')}")
        return {"task_id": task_id, "status": data.get("status", "queued"), "raw": data}
    except _requests.HTTPError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sora/status/{task_id}")
async def sora_status(task_id: str, x_sora_api_key: Optional[str] = Header(default=None)):
    """Poll Sora task status."""
    headers = _sora_headers(x_sora_api_key)
    try:
        res = _requests.get(f"{SORA_API_URL}/{task_id}", headers=headers, timeout=60)
        res.raise_for_status()
        data = res.json()
        return {
            "task_id": task_id,
            "status": data.get("status", "unknown"),
            "progress": data.get("progress", 0),
            "raw": data,
        }
    except _requests.HTTPError as e:
        raise HTTPException(status_code=e.response.status_code, detail=e.response.text)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sora/stream/{task_id}")
async def sora_stream(task_id: str, x_sora_api_key: Optional[str] = Header(default=None)):
    """
    Fetch MP4 from /v1/videos/{id}/content, save to Sora_outputs/, stream to browser.
    Retries on 404/IN_PROGRESS (can lag after status=completed).
    """
    from fastapi.responses import Response as FResponse
    import time as _time

    headers = _sora_headers(x_sora_api_key)
    content_url = f"{SORA_API_URL}/{task_id}/content"
    safe_id = task_id.replace("/", "_").replace("\\", "_")
    save_path = os.path.join(SORA_OUTPUT_DIR, f"{safe_id}.mp4")

    # Serve from cache if already downloaded
    if os.path.exists(save_path) and os.path.getsize(save_path) > 1000:
        print(f"[sora/stream] ✓ Serving cached: {save_path}")
        with open(save_path, "rb") as f:
            return FResponse(content=f.read(), media_type="video/mp4",
                             headers={"Content-Disposition": f'inline; filename="{safe_id[:24]}.mp4"',
                                      "Cache-Control": "no-store", "X-Sora-Cached": "true"})

    print(f"[sora/stream] Fetching: {content_url}")
    last_err = None

    for attempt in range(6):
        try:
            res = _requests.get(content_url, headers=headers, timeout=180)
            print(f"[sora/stream] attempt={attempt + 1} status={res.status_code} size={len(res.content)}")

            if res.status_code == 200 and len(res.content) > 1000:
                with open(save_path, "wb") as f:
                    f.write(res.content)
                size_mb = len(res.content) / 1_048_576
                print(f"[sora/stream] ✓ Saved {size_mb:.1f} MB -> {save_path}")
                return FResponse(content=res.content, media_type="video/mp4",
                                 headers={"Content-Disposition": f'inline; filename="{safe_id[:24]}.mp4"',
                                          "Cache-Control": "no-store",
                                          "X-Sora-Saved-Path": save_path})

            last_err = res.text[:400]
            print(f"[sora/stream] Not ready: {last_err}")

            if "IN_PROGRESS" in res.text or "in_progress" in res.text or res.status_code == 404:
                _time.sleep(12)
            else:
                raise HTTPException(status_code=res.status_code, detail=res.text[:300])

        except HTTPException:
            raise
        except Exception as e:
            last_err = str(e)
            print(f"[sora/stream] Exception {attempt + 1}: {e}")
            _time.sleep(12)

    raise HTTPException(status_code=503, detail=f"Video not available after retries: {last_err}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
# ===============================================================================
# WHISK + FLOW STORYBOARD  (from v2.5)
# ===============================================================================

class WhiskRequest(BaseModel):
    # support both naming conventions: subject_b64 and subject_image_b64
    subject_b64: str = ""
    subject_image_b64: str = ""
    subject_mime: str = "image/jpeg"
    subject_image_mime: str = ""
    subject_desc: str = ""
    subject_description: str = ""
    scene_b64: str = ""
    scene_image_b64: str = ""
    scene_mime: str = "image/jpeg"
    scene_image_mime: str = ""
    scene_desc: str = ""
    scene_description: str = ""
    style_b64: str = ""
    style_image_b64: str = ""
    style_mime: str = "image/jpeg"
    style_image_mime: str = ""
    style_desc: str = ""
    style_description: str = ""
    model: str = "flux-kontext-max"
    aspect_ratio: str = "1:1"
    image_size: str = "1K"

    def effective_subject_b64(self):  return self.subject_b64 or self.subject_image_b64

    def effective_subject_mime(self): return self.subject_image_mime or self.subject_mime or "image/jpeg"

    def effective_subject_desc(self): return self.subject_desc or self.subject_description

    def effective_scene_b64(self):    return self.scene_b64 or self.scene_image_b64

    def effective_scene_mime(self):   return self.scene_image_mime or self.scene_mime or "image/jpeg"

    def effective_scene_desc(self):   return self.scene_desc or self.scene_description

    def effective_style_b64(self):    return self.style_b64 or self.style_image_b64

    def effective_style_mime(self):   return self.style_image_mime or self.style_mime or "image/jpeg"

    def effective_style_desc(self):   return self.style_desc or self.style_description


def _describe_via_vision(b64: str, mime: str, slot_hint: str) -> str:
    """Call gemini-2.5-flash with vision to get a concise image description."""
    client = make_client(model)  # use chat key; deepseek direct models use DEEPSEEK_API_KEY
    resp = client.chat.completions.create(
        model="gemini-2.5-flash",
        messages=[{
            "role": "user",
            "content": [
                {"type": "text", "text": (
                    f"Describe this image in 1-2 concise sentences for an image generation prompt. "
                    f"Slot: '{slot_hint}'. Focus on the most visually distinctive elements. "
                    f"Do NOT use phrases like 'This image shows' -- describe directly."
                )},
                {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
            ],
        }],
        max_tokens=120,
        stream=False,
    )
    return resp.choices[0].message.content.strip()


class FlowStoryboardRequest(BaseModel):
    script: str
    style: str = "cinematic"
    scene_count: int = 4   # 0 = let AI decide
    model: str = "nano-banana-hd"  # for storyboard images
    chat_model: str = "gemini-2.5-flash"  # for scene text generation
    aspect_ratio: str = "16:9"
    generate_images: bool = False
    image_style: str = ""  # visual render style (e.g. "Studio Ghibli", "Rembrandt painting")
    auto_scene_count: bool = False  # True = AI decides how many scenes


# Max scenes generated per single AI text call. Higher scene counts are split
# into multiple parallel calls so each batch gets its own fresh token budget
# (avoids the single-call max_tokens ceiling truncating the JSON array).
MAX_SCENES_PER_BATCH = 8


def _split_scene_batches(total: int, max_per: int = MAX_SCENES_PER_BATCH):
    """Split `total` scenes into evenly-sized batches of at most `max_per`.

    The number of batches is the minimum needed, and scenes are distributed as
    evenly as possible across them (front-loaded), e.g.:
        10 -> [5, 5]            (2 batches, not [8, 2])
        16 -> [8, 8]
        17 -> [6, 6, 5]
        30 -> [8, 8, 7, 7]      (4 batches, not [8, 8, 8, 6])
    Returns a list of (count, offset) tuples where offset is the 0-based index
    of the batch's first scene within the full sequence.
    """
    if total <= 0:
        return []
    n_batches = -(-total // max_per)  # ceil division
    base, rem = divmod(total, n_batches)
    sizes = [base + 1 if i < rem else base for i in range(n_batches)]
    out, offset = [], 0
    for c in sizes:
        out.append((c, offset))
        offset += c
    return out


@app.post("/whisk")
async def whisk_generate(
        req: WhiskRequest,
        x_image_api_key: Optional[str] = Header(None, alias="X-Image-API-Key"),
):
    """
    Whisk: combine Subject + Scene + Style into one generated image.
    For each slot, uses provided text description or calls vision model on the image.
    """
    from concurrent.futures import ThreadPoolExecutor

    def resolve(b64: str, mime: str, desc: str, hint: str) -> str:
        if desc.strip():
            return desc.strip()
        if b64:
            try:
                return _describe_via_vision(b64, mime, hint)
            except Exception:
                return ""
        return ""

    with ThreadPoolExecutor(max_workers=3) as pool:
        fut_subject = pool.submit(resolve, req.effective_subject_b64(), req.effective_subject_mime(),
                                  req.effective_subject_desc(), "subject/character")
        fut_scene = pool.submit(resolve, req.effective_scene_b64(), req.effective_scene_mime(),
                                req.effective_scene_desc(), "background/environment/scene")
        fut_style = pool.submit(resolve, req.effective_style_b64(), req.effective_style_mime(),
                                req.effective_style_desc(), "artistic style/visual aesthetic")
        subject_txt = fut_subject.result(timeout=40)
        scene_txt = fut_scene.result(timeout=40)
        style_txt = fut_style.result(timeout=40)

    parts = []
    if subject_txt: parts.append(subject_txt)
    if scene_txt:   parts.append(f"in {scene_txt}")
    if style_txt:   parts.append(f"rendered in the style of {style_txt}")

    if not parts:
        raise HTTPException(400, "At least one slot (subject, scene, or style) must have an image or description.")

    combined_prompt = ", ".join(parts)

    cfg = IMAGE_MODELS.get(req.model)
    if not cfg:
        raise HTTPException(400, f"Unknown image model: {req.model}")

    # Always use IMAGE_API_KEY from env
    key = IMAGE_API_KEY
    ref_b64 = req.effective_subject_b64() if cfg["api"] == "openai-image" and req.effective_subject_b64() else ""

    try:
        api = cfg["api"]
        mdl = cfg["model"]
        ep = cfg.get("extra_params") or {}
        smap = cfg.get("size_map_vip")

        if api == "chat-image-url":
            b64 = _generate_chat_image(combined_prompt, mdl, req.aspect_ratio, ref_b64, returns_url=True, key=key)
        elif api == "chat-image-b64":
            b64 = _generate_chat_image(combined_prompt, mdl, req.aspect_ratio, ref_b64, returns_url=False, key=key)
        elif api == "google":
            b64 = _generate_google(combined_prompt, mdl, req.aspect_ratio, req.image_size, ref_b64, key=key)
        elif api == "seedream":
            b64 = _generate_seedream(combined_prompt, mdl, ref_b64)
        elif api in ("openai-image", "openai-image-url"):
            b64 = _generate_openai_image(
                combined_prompt, mdl, req.aspect_ratio, req.image_size,
                ref_b64, extra_params=ep, size_map_vip=smap,
                returns_url=(api == "openai-image-url"), key=key,
            )
        else:
            raise HTTPException(400, f"Unknown API type: {api}")

        return {
            "image_b64": b64,
            "model": req.model,
            "combined_prompt": combined_prompt,
            "subject_desc": subject_txt,
            "scene_desc": scene_txt,
            "style_desc": style_txt,
        }
    except _requests.HTTPError as e:
        raise HTTPException(e.response.status_code, f"API error: {e.response.text[:400]}")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        pass  # IMAGE_API_KEY from env, no restore needed


# ---------------------------------------------------------------------------


class FlowImagesRequest(BaseModel):
    scenes: list[dict]
    model: str = "nano-banana-hd"
    aspect_ratio: str = "16:9"
    image_style: str = ""


@app.post("/flow/images")
async def flow_images_only(
        req: FlowImagesRequest,
        x_image_api_key: Optional[str] = Header(None, alias="X-Image-API-Key"),
):
    """Generate storyboard images for already-generated scenes (no text generation).
    Supports all IMAGE_MODELS -- used by frontend Google mode with nano-banana models.
    """
    cfg = IMAGE_MODELS.get(req.model)
    if not cfg:
        raise HTTPException(400, f"Unknown image model: {req.model}. Available: {list(IMAGE_MODELS.keys())}")
    if not req.scenes:
        raise HTTPException(400, "scenes required")

    global IMAGE_API_KEY
    effective_key = IMAGE_API_KEY or x_image_api_key
    original_key = IMAGE_API_KEY
    images: list[dict] = []

    try:
        IMAGE_API_KEY = effective_key

        def _gen_frame(args):
            idx, scene = args
            style_suffix = f" {req.image_style} style." if req.image_style else ""
            prompt = (
                f"{scene.get('description', '')}. "
                f"Camera: {scene.get('camera', '')}."
                f"{style_suffix} Cinematic still frame."
            )
            try:
                api = cfg["api"]
                mdl = cfg["model"]
                ep = cfg.get("extra_params") or {}
                if api == "google":
                    r = _requests.post(
                        f"{GOOGLE_IMAGE_BASE}/{mdl}:generateContent",
                        headers={"Authorization": f"Bearer {IMAGE_API_KEY}", "Content-Type": "application/json"},
                        json={"contents": [{"parts": [{"text": prompt}]}],
                              "generationConfig": {"responseModalities": ["IMAGE"],
                                                   "imageConfig": {"aspectRatio": req.aspect_ratio,
                                                                   "imageSize": "1K"}}},
                        timeout=180)
                    r.raise_for_status()
                    return r.json()["candidates"][0]["content"]["parts"][0]["inlineData"]["data"]
                elif api in ("chat-image-b64", "chat-image-url"):
                    return _generate_chat_image(prompt, mdl, req.aspect_ratio, "",
                                                returns_url=(api == "chat-image-url"),
                                                key=IMAGE_API_KEY)
                elif api in ("openai-image", "openai-image-url"):
                    return _generate_openai_image(prompt, mdl, req.aspect_ratio, "1K", "",
                                                  extra_params=ep,
                                                  returns_url=(api == "openai-image-url"),
                                                  key=IMAGE_API_KEY)
                return ""
            except Exception as _err:
                print(f"[FLOW IMAGES ERROR] model={req.model} scene={idx} err={_err}")
                return ""

        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=4) as pool:
            results = list(pool.map(_gen_frame, enumerate(req.scenes), timeout=120))
        images = [{"index": i, "image_b64": b64} for i, b64 in enumerate(results)]

    finally:
        IMAGE_API_KEY = original_key

    return {"images": images}


WORDS_TO_TOKENS_NARASI = 1.5

# Models that route through LaoZhang's OpenAI-compatible relay AND consume
# thinking tokens from the same max_tokens budget (unlike the Google SDK which
# keeps thinking tokens separate).  Add a generous overhead so thinking doesn't
# eat into the actual text output budget.
THINKING_MODELS_NARASI = {
    "gemini-2.5-pro",
    "gemini-2.5-flash",
    "claude-sonnet-4-6-thinking",
    "claude-opus-4-7-thinking",
    "deepseek-v4-pro",    # ← tambah ini
    "deepseek-r1",        # ← dan ini
}
THINKING_TOKEN_OVERHEAD = 32000  # conservative buffer for thinking tokens

STYLE_RULES = {
    "creative non-fiction": """
STYLE: Creative Non-Fiction
= Techniques of fiction (concrete scenes, specific POV, sensory detail) applied to REAL FACTS.

STRUCTURE PER CHAPTER:
1. COLD OPEN -- One specific cinematic scene. Specific object, person, moment -- NOT abstract.
   BAD: "Para leluhur membawa harapan ke cakrawala."
   GOOD: "Di geladak sempit: benih padi dibungkus daun pisang, seekor babi betina bunting diikat di tiang."
2. UNTOLD STORY -- The fact most people don't know. Specific data: %, dates, species names, site names.
3. SUDUT PANDANG -- At least one scene from a specific character's human POV.

FORBIDDEN: "harapan", "keberanian", "gema purba", "kita adalah kelanjutan mereka", "cakrawala yang menari", "penjelajah tak gentar"
REQUIRED: Min 2 specific facts with numbers/dates per section. 1 concrete object/sensory detail per paragraph.
""",

    "storytelling": """
STYLE: Storytelling -- Narrative Drama
= Story-first. Every historical fact must be delivered through SCENE and CHARACTER, not exposition.

STRUCTURE PER CHAPTER:
1. SCENE OPENER -- Drop into the middle of a moment. In medias res. Who, what, where -- in the first sentence.
   BAD: "Pada masa itu, perdagangan rempah sangat berkembang."
   GOOD: "Tangannya gemetar ketika menyerahkan ikat cengkih terakhir kepada nahkoda asing itu."
2. CONFLICT/TENSION -- Every chapter needs a problem or stakes. What does someone want? What stands in the way?
3. DIALOGUE -- At least 2 lines of spoken dialogue per chapter. Ground it in specific context.
4. TURN -- A moment where something changes: a realization, a surprise, a decision.

FORBIDDEN: Passive summary of events. Telling emotion instead of showing. Generic descriptions.
REQUIRED: Named or clearly characterized figures. Cause-and-effect within scenes. Physical action.
""",

    "bedtime story": """
STYLE: Bedtime Story -- Gentle, Soothing
= Warm narrator voice, gentle wonder, age-appropriate vocabulary. History as a lullaby.

STRUCTURE PER CHAPTER:
1. SOFT OPENING -- Begin with a peaceful image or a gentle question. No drama, no conflict.
   GOOD: "Bayangkan kamu duduk di tepi pantai, ribuan tahun yang lalu, melihat perahu pertama muncul di cakrawala."
2. SENSE OF WONDER -- Each chapter reveals one amazing thing in a way that feels like a gift, not a lesson.
3. COMFORTING CLOSE -- End each chapter with warmth. A sense that things turned out okay.

FORBIDDEN: Violence, conflict, darkness. Complex syntax. Academic jargon.
REQUIRED: Short sentences. Soft vocabulary. Metaphors from nature and everyday life. Second-person ("kamu") or inclusive "kita".
""",

     "harari": """
STYLE: Harari / Jared Diamond -- Big History (Sapiens-style)
= Claim -> Evidence -> Implication. Zoom from the specific to the cosmic. Ask: what does this mean for ALL of humanity?

# HARARI / DIAMOND BIG-HISTORY VOICEOVER SYSTEM — v2
## Tuned for: reduced rule saturation, added anti-repetition discipline, separated editor pass

---

## DEPLOYMENT NOTE

This file contains **two prompts**. Use them at different stages.

- **PART A — GENERATION PROMPT.** Feed to the model when writing a chapter.
- **PART B — EDITOR PASS PROMPT.** Run on a finished chapter or full manuscript. **Do not include during generation.** Loading it during generation creates self-conscious prose.

Before starting, decide deliverable mode:

- `[SCRIPT_MODE]` — working VO script handed to an editor/VO talent. Keep `[pause]` / `[beat]` / `[silence]` markers as production direction. Density rules apply.
- `[PROSE_MODE]` — literary nonfiction styled as VO, read on the page. **Omit all markers.** Use paragraph breaks and short landing sentences instead.

Pick one. The rest of the system adapts.

For multi-chapter projects, maintain a **Pattern Ledger** — a short running list of rhetorical figures, signature lines, and cinematic objects used so far. Pass it as context when generating chapter N. The anti-repetition rules below cannot be enforced without it.

---
---

# PART A — GENERATION PROMPT

---

## PRIME DIRECTIVE

You are writing **documentary narration** in the tradition of Yuval Noah Harari and Jared Diamond.

Every sentence will be spoken aloud, heard once, and felt immediately. Write each sentence as spoken language at formal/literary register from the first word. Do not write prose and convert later.

Three commands govern everything:

1. **Claim → Evidence → Implication.** Zoom from the specific to the cosmic.
2. **One breath per idea.** Around 25 words per sentence, ceiling not target.
3. **Do not sound profound all the time.** Sound observant, precise, restrained, occasionally devastating. *If every sentence feels monumental, nothing feels monumental.*

The narrator must sound like *an intelligent observer speaking carefully about deep time* — not a philosophical trailer, not a brilliant essayist, not a novelist controlling destiny.

---

## PRIORITY WHEN RULES CONFLICT

When the system's rules pull in different directions, follow this order:

1. **Prime Directive**
2. **Layer 4 — Cinematic Restraint** (the load-bearing discipline)
3. **Anti-repetition discipline** (global, applies across chapters)
4. **Layer 1 content requirements** (substance the chapter must contain)
5. **Layer 2 delivery preferences** (preferred shapes; flexible)

Worked example: if a chapter cannot land its scholarly counter-theory without exceeding the aphorism cap, *Layer 4 wins.* Re-shape the counter-theory; do not force the aphorism.

Worked example: if including the per-book cognitive implication this chapter would create a third trailer-worthy paragraph in a row, *defer it to the next chapter.* Layer 4 wins.

---

## ANTI-REPETITION DISCIPLINE (global)

Across the manuscript, no rhetorical figure should appear more than twice. On the third occurrence, break the pattern: same content, different shape.

Track at minimum:

- Negation tricolon ("Tidak ada X. Tidak ada Y. Tidak ada Z.")
- Repeated absence framing ("Nelayan tanpa X. Pedagang tanpa Y.")
- One-word reveal paragraph
- Specific epistemic-distance formula ("Ia tidak tahu — tidak mungkin tahu")
- Specific chapter-opening grammar
- Specific chapter-closing grammar (poetic detonator vs procedural vs question)
- Anaphoric repetition at paragraph head
- Any signature aphorism shape

**Chapter ending diversity:** of 10 chapters, no more than 6 may close on a poetic detonator. At least 3 must close on procedural fact, unresolved scholarly question, or physical residue stated plainly.

**The general rule:** if you notice yourself reaching for a shape because it worked before, that is the moment to break it.

---

# LAYER 1 — CONTENT

---

## 1.1 Opening orientation

Four working modes. Vary across chapters. Don't repeat the same mode in consecutive chapters.

- **Reversal** — a received belief, named, then turned over.
- **Specific surprise** — a concrete dated, located, quantified fact that recontextualises the chapter.
- **Scene** — body, weather, object, or action, no interpretation yet.
- **Stated unknown** — a paradox or open question the chapter will work through.

The strongest specific-surprise opening should be reserved for the chapter with the most unfamiliar evidence. The most reversal-heavy opening should be the prologue.

---

## 1.2 Evidence requirements

Per chapter:
- One named scholar with their role (2-4 words). Dates only when the date matters.
- At least one quantified data point — site, artifact, percentage, kilometres, dated range.
- **One named scholarly debate or counter-theory.** This is mandatory and load-bearing. Counter-theories must be stated at full formal register, never collapsed to "ada yang berpendapat beda."
- One acknowledgment of what the evidence **cannot** show.

---

## 1.3 Comparative lens

When Subject A is positioned over Subject B, explain B's **ecological or geographic constraints**, never their character or courage.

- WRONG: *"Mesir takut pada laut."*
- RIGHT: *"Sementara peradaban Nil mengoptimalkan untuk banjir tahunan yang bisa diprediksi, pelaut Austronesia mengoptimalkan untuk ketidakpastian."*

This is one of the hardest rules to follow. It is also one of the most important.

---

## 1.4 Epistemic distance

The narrator does not know the inner states of historical actors. Available hedges: *"kemungkinan besar," "mungkin," "yang baru kita pahami sekarang," "yang tampaknya..."*

**Hindsight prophecy is banned outright.** "Tanpa mereka sadari..." / "Little did they know..." — never. There are no exceptions.

**Note on the "ia tidak tahu — tidak mungkin tahu" formula:** this is one valid path to epistemic distance, but **not the only one.** Other paths: scholarly voice ("Bukti genetik mencatat hasilnya, bukan kehendaknya"), structural silence ("Apa yang terjadi setelah itu, tidak ada catatan"), or simply leaving the moment uninterpreted. **Use the "ia tidak tahu" hedge at most twice across the manuscript.**

---

## 1.5 Historical friction

Each chapter contains one moment of friction — a known failure, a competing faction, an evidential silence, or an unintended consequence. Choose what fits the chapter's evidence. Don't taxonomise.

---

## 1.6 Required per chapter (reduced)

Exactly four things must appear:

1. One named scholarly debate
2. One quantified data point
3. One historical friction moment
4. One sensory or material moment that grounds the chapter's macro claim in a body, an object, or a weather condition

---

## 1.7 Required per book (NOT per chapter)

These were per-chapter in v1 and produced over-density. Spread them across the manuscript instead:

- **Two to three explicit cognitive/evolutionary implications** total. Not one per chapter. Most chapters do not need one.
- **One cinematic object tracked across chapters** in transformed state. Introduce in early chapter, return in late chapter, transformed by intermediate generations.
- **Two to three understated wonder moments** total. Not more.
- **Two to three linguistic-authenticity moments** where a regional term carries unpacked cognitive/ecological weight.

This frees most chapters to be quieter and more procedural.

---

## 1.8 Hard bans (non-negotiable)

These kill prestige tone on contact:

- *"Bayangkan seorang..."* as a chapter opener.
- *"Inilah pelajaran besar bagi seluruh umat manusia."*
- *"Para leluhur yang gagah berani / tak gentar / penuh semangat."*
- Hindsight prophecy in any form.
- Character-degradation comparative framing.
- Over-confident claims on genuinely contested hypotheses (use hedged language: *"kandidat terkuat," "lebih berpihak,"* not *"sudah pasti"*).
- Anachronistic political or psychological language imposed on historical actors.

---

## 1.9 Strongly avoid (probabilistic, not absolute)

Use sparingly, ideally less than the named threshold:

- *"Temuan ini mengungkap..."* — twice across the manuscript at most.
- *"Luar biasa," "menakjubkan," "mengherankan"* and other wonder-inflation language.
- Harari-branded terminology (*"fiksi kolektif," "revolusi kognitif"*) — find your own formulations from the evidence on the page.
- More than two aphorisms in a single chapter. More than 14 aphorisms across the manuscript.

---

# LAYER 2 — VO DELIVERY

---

## 2.1 Breath economy

Sentences run up to about 25 words. The ceiling is a guide, not a target — many sentences will be shorter. When splitting a long sentence, **never drop the register.** Shorten the breath units, keep the vocabulary.

- WRONG: *"Ibu petani Hemudu? Dua tahun sudah cukup."* (conversational)
- RIGHT: *"Ibu petani Hemudu menyapih di usia dua tahun. Jarak kelahiran memendek."* (formal, two breaths)

---

## 2.2 Sentence variety

Vary sentence length organically. Avoid extended rhythmic uniformity. Short sentences should feel earned by the content, not placed for cadence.

If you find yourself writing patterned rhythms — long-short-long-short, three shorts in a row, then a long — you have automated cadence. Break the pattern. Real documentary speech is **slightly irregular**.

---

## 2.3 Paragraph architecture

- Target 3-6 sentences per paragraph. Soft target, not a cap.
- After a data-dense paragraph, ground in sensory or material detail.
- Never two aphorisms in the same paragraph. Keep the stronger; convert the weaker to plain observation.

---

## 2.4 Proper noun gloss

Mandatory on first mention — 2 to 5 words inline. A listener cannot pause and search.

| Wrong | Right |
|-------|-------|
| *"situs Hemudu, delta Sungai Yangtze"* | *"situs Hemudu — desa kuno di delta Sungai Yangtze"* |
| *"haplogroup Q dan P"* | *"penanda genetik yang disebut haplogroup Q dan P"* |
| *"nekara tipe Heger I"* | *"nekara perunggu bergaya Heger I — klasifikasi standar arkeologis"* |

Persons: first mention `[Name], [role in 2-4 words]`. Subsequent mentions: last name only.

---

## 2.5 Citation integration

Academic citation must speak naturally, not parenthetically.

- BANNED: *"...sebagaimana dikemukakan Blust (1999, AO 34:2)..."*
- RIGHT: *"...Robert Blust, dalam empat dekade rekonstruksi leksikalnya, menemukan..."*

Counter-theories at full formal register, always. See 1.2.

---

## 2.6 Anchor lines (probabilistic guidance)

**Target range:** 2 to 4 anchor-quality lines per chapter. **Some chapters will naturally have fewer.** Do not manufacture anchors to hit a count. A manufactured anchor is worse than no anchor.

An anchor, when one emerges:
- Under 12 words.
- Standalone meaning.
- Paradox or reversal preferred but not required.
- Does **not** restate the surrounding paragraph.

---

## 2.7 Markers — conditional on deliverable mode

**`[SCRIPT_MODE]`:** Use `[pause]`, `[beat]`, `[silence]` as production direction. Maximum one marker per 3-4 paragraphs. Never two markers in adjacent paragraphs. Markers are functional metadata for the editor, not literary devices.

**`[PROSE_MODE]`:** Omit all markers. Use paragraph breaks and short landing sentences. Cadence is implied by line length and white space.

---

## 2.8 Chapter closings

End each chapter on a beat that opens forward. Don't echo the paragraph above.

**Vary the closing register across the manuscript:**
- Some chapters close on a poetic detonator.
- Some close on procedural fact.
- Some close on an unresolved scholarly question.
- Some close on physical residue — an object that outlasts the story.

**At least 3 of 10 chapter endings must be procedural, factual, or quiet.** Not every chapter needs a mic-drop. Sequential mic-drops produce fatigue and signal automation.

---

## 2.9 Banned in VO

- Complex nested clauses with three or more subordinate levels.
- Two aphorisms in the same paragraph.
- Anchor that echoes the paragraph before it.
- Register drop to conversational when splitting long sentences.
- Abstract tangents the listener cannot visualise.
- Factual claims not present in or inferable from source material.
- Hindsight prophecy in any form.

---

# LAYER 3 — INTEGRATION

Where Layers 1 and 2 must hold simultaneously:

1. **Opening + VO-readiness:** the chapter's first sentence must be both content-effective and speakable in one breath. Statistical-paradox openings need a breath break between the statistic and its implication.

2. **Evidence + gloss:** every named scholar follows person-name protocol on first mention; every site or technical term gets micro-gloss; lists of more than two data points break into separate breath units.

3. **Scholarly tension + register:** counter-theories always at full formal register. Never collapsed.

4. **Implication + anchor (when both present):** if the chapter contains its share of the per-book cognitive implication, the implication should crystallise into an anchor — or set one up. Don't bolt anchor and implication together arbitrarily; the anchor is the *earned form* of the implication.

5. **Friction + landing:** moments of uncertainty, failure, or conflict are natural landing places. In `[SCRIPT_MODE]`, place `[beat]` after the friction sentence, not before. In `[PROSE_MODE]`, let paragraph break do the work.

---

# LAYER 4 — CINEMATIC RESTRAINT
*The load-bearing discipline. When in doubt, this wins.*

---

## Core principle

A documentary narrator does not explain every meaning. Sometimes the image carries it. Sometimes silence carries it better.

**Your job is not to sound intelligent every sentence. Your job is to control cognitive pressure over time.**

---

## 4.1 Emotional modulation

Never sustain maximum intensity for more than 2 consecutive paragraphs. After any high-intensity sequence, insert one low-intensity factual or observational paragraph.

| High-intensity (don't stack) | Low-intensity reset |
|------------------------------|---------------------|
| Existential / philosophical | Archaeological description |
| Civilisation-scale implication | Logistics |
| Elegiac reflection | Material detail |
| Aphoristic line | Chronology |
| Emotional compression | Environmental observation |
| Mortality framing | Calm factual narration |

**Test:** if three paragraphs in a row read trailer-worthy, flatten one completely.

---

## 4.2 Factual breathing paragraphs

Every chapter contains at least one paragraph with:
- Zero metaphor.
- Zero existential framing.
- Plain explanation of evidence.

These are not filler. They restore credibility and listener trust between intensity peaks.

---

## 4.3 Aphorism control

Aphoristic lines are rare weapons.

- Maximum 1 aphorism per paragraph.
- Maximum 2 major quotables per chapter.
- Maximum 14 aphorisms across the manuscript.

If three consecutive sentences could appear as standalone quotes, rewrite two of them into plain observation.

**Caution:** aphorisms that crystallise the same thesis ("hierarchy moved medium," "knowledge moved medium," "power moved medium") count as one repeated figure under the anti-repetition discipline. Use the thesis-shape **once across the manuscript.**

---

## 4.4 Anti-imitation filter

Intellectual atmosphere may resemble Harari. Cadence must not.

**Banned pattern (when repeated):**
> short sentence
> short sentence
> philosophical reversal
> existential punchline

After every philosophical statement, insert one of:
- Material detail
- Procedural explanation
- Environmental observation
- Ordinary human action

**Documentary writing trusts reality.** The facts carry their own weight.

---

## 4.5 Documentary realism

History must feel **discovered, not pre-written.**

Avoid sounding like an omniscient philosopher, a novelist controlling destiny, or a civilisation poet composing quotes.

Prefer observational authority, restrained intelligence, evidentiary humility.

**Earned philosophy:** philosophical lines emerge *from* evidence — they do not descend onto it. Sequence: object → observation → implication.

**Invisible author:** the narrator may sound wise. The writer must remain invisible. If a sentence sounds like *"this was written to be quoted,"* rewrite simpler.

---

## 4.6 Camera mode

At least once every 3-4 paragraphs, switch into camera mode: describe only movement, texture, sound, weather, physical action, visible material — **without interpretation.**

| Camera mode | Interpretation (avoid here) |
|-------------|----------------------------|
| *"Air menetes dari ujung cadik. Tali rotan mengencang setiap kali lambung menghantam gelombang."* | *"Laut menguji keberanian mereka."* |

Image first, meaning second. If the audience can see a cracked stone, smoke, jungle rain, or abandoned foundations, do not explain the symbolism. Delay interpretation.

---

## 4.7 Atmospheric reset

After heavy exposition, scholarly debate, abstract implication, or comparison, insert weather, texture, sound, or movement. These are not decoration. They regulate listener fatigue.

---

## 4.8 Human scale

Civilisational narration must repeatedly return to **bodies, labour, hunger, weather, fatigue, tools, smell, sound, repetitive work.** Every macro claim periodically reconnects to *a hand, a rope, a wound, a bowl, a paddle, a wall, a child, a shoreline.*

**History happens through bodies first.**

---

## 4.9 No continuous grand meaning

After every major implication, decompression — a sentence that returns to object, body, weather, sound, or silence.

The structure is implication → decompression. Not implication → implication → implication.

---

## 4.10 Controlled wonder

Wonder works understated. Avoid awe-language. Prefer quiet declarative:

> *"Batunya kecil. Tetapi ia bertahan lebih lama dari kerajaannya."*

---

## 4.11 Respect the audience

PBS/Netflix-grade narration assumes intelligence. **Sometimes stop one sentence earlier.**

Remove redundant explanation. Reduce repeated meaning. Avoid paragraph-level thesis restatement. Trust montage and editing.

**If a paragraph works perfectly without visuals, it may be too literary.**

---

## 4.12 Ending control

Strong documentary endings **arrive, land, stop.** No philosophical spiralling afterward. If the final paragraph contains more than one big idea, cut the weaker.

---

## 4.13 Historical humility

Narrator stays aware of uncertainty. Avoid omniscient emotional framing, retroactive destiny, mythic inevitability.

Preferred tone: careful reconstruction under incomplete evidence. The audience should feel history is fragile, partial, difficult to recover.

---

## 4.14 Tone balance

Final output must balance two registers:

| Cinematic mode | Scholarly mode |
|----------------|----------------|
| Immersion | Restraint |
| Emotional intimacy | Calm |
| Sensory immediacy | Evidentiary precision |
| Human-scale tension | Measured uncertainty |

If narration goes too lyrical, inject restraint. If too academic, inject sensory immediacy.

Ideal: *an intelligent observer speaking carefully about deep time.*

---

# SAMPLE PASSAGE
*A worked example showing the rules in concert. Use it to calibrate, not to imitate.*

> Di pesisir timur Sulawesi, sekitar dua ribu tiga ratus sebelum masehi, seorang pembuat perahu memilih batang bambu yang sudah dikeringkan selama dua musim. Bambu yang lebih muda akan retak di bawah tekanan rotan. Bambu yang lebih tua menjadi rapuh.
>
> Ia mengukur jarak antara lambung dan pelampung dengan rentang lengan. Bukan dengan satuan yang bisa dicatat — dengan tubuhnya sendiri.
>
> Arkeolog Atholl Anderson, dari Universitas Nasional Australia, berargumen bahwa standardisasi cadik Austronesia tidak dapat dijelaskan oleh difusi tunggal. Pola yang ia petakan di lebih dari empat puluh situs menunjukkan konvergensi independen dalam batas-batas yang terlalu sempit untuk kebetulan. Geoffrey Irwin tidak sependapat: menurut hitungannya, satu populasi inti dengan migrasi cepat cukup untuk menjelaskan keseragamannya. Bukti morfologis belum bisa memisahkan kedua model.
>
> Yang dapat kita pastikan lebih sedikit: jarak antara lambung dan pelampung, di hampir setiap perahu yang masih dibangun hari ini dari Madagaskar hingga Hawaii, berada dalam rentang satu meter dari rasio yang sama.
>
> Pembuat perahu di pesisir Sulawesi tidak tahu sebuah pulau di mana cadik akan dirakit dengan rasio yang nyaris identik, ribuan tahun setelah ia mati. Yang ia tahu hanya bahwa pelampungnya tidak boleh terlalu jauh, agar perahu tidak berputar; dan tidak terlalu dekat, agar perahu tidak terbalik.
>
> Sore itu hujan turun sebentar. Rotan yang basah lebih lentur, tetapi juga lebih sulit dikencangkan.
>
> Ia menunggu hujan berhenti.

What the passage does:
- Opens with **scene + body**, not interpretation.
- Introduces named scholars with role; states the debate at full formal register.
- Acknowledges what is **not** known (which model fits the evidence).
- Uses epistemic distance through **scholarly voice**, not the "ia tidak tahu" formula. The formula appears once, later, in a softer form.
- Camera mode in the final two paragraphs — weather, rotan, waiting. No symbolism.
- No aphorism. No anchor. Some chapters look like this.

---

# FINAL RULE

**Do not optimise every paragraph equally.**

Some paragraphs may be quieter, plainer, rougher, more procedural, less quotable. Contrast creates memorability.

*Documenter terbaik tidak terdengar seperti mencoba menjadi abadi. Mereka hanya terdengar benar.*

---
---

# PART B — EDITOR PASS PROMPT
*Run on a finished chapter or full manuscript. Do not load during generation.*

---

## ROLE

You are an editor reading a completed manuscript for prestige documentary VO. Your job is to find where the writing has automated — where rules became patterns, where good moves became signatures.

You are not adding. You are auditing and proposing breaks.

---

## STEP 1 — SIGNATURE HUNT

Read the full manuscript. Find every rhetorical figure used three or more times.

Hunt at minimum:
- Negation tricolon ("Tidak ada X. Tidak ada Y. Tidak ada Z.")
- Absence anaphora ("X tanpa Y. P tanpa Q.")
- One-word reveal paragraphs
- "Ia tidak tahu — tidak mungkin tahu" formula
- Repeated chapter-opening grammar
- Repeated chapter-closing grammar
- Anaphoric paragraph heads
- Any aphorism shape that recurs (e.g., "X belum hilang. Ia hanya berpindah Y.")

For each pattern, output:
- The figure
- All locations (chapter and approximate position)
- A proposed rewrite for the third (and later) uses that keeps the content but changes the shape

---

## STEP 2 — CHAPTER ENDING AUDIT

List every chapter's closing line. Categorise each as: **poetic detonator / procedural / unresolved question / physical residue.**

If more than 6 of 10 endings are poetic detonators, propose which 2-3 to rewrite as procedural, factual, or quietly questioning.

---

## STEP 3 — APHORISM COUNT

Count aphoristic sentences per chapter and across the manuscript. Per-chapter max: 2. Manuscript max: 14.

For any over-count, identify which aphorisms can be demoted to plain observation. Keep the strongest; flatten the rest.

---

## STEP 4 — RHYTHM AUDIT

Scan for places where sentence-length rhythm has become patterned. Three short sentences in a row; three long in a row; long-short-long-short for more than four beats; repeated negation patterns.

Propose one break per chapter where automation is most visible.

---

## STEP 5 — IMITATION FLAG

Flag any sentence that sounds *written to be quoted.* Flag any paragraph that admires its own intelligence. Flag any moment of writerly cleverness that does not earn its place.

For each flag, propose the simpler version.

---

## STEP 6 — TRUST AUDIT

Find paragraphs that explain implications the audience could infer from the previous paragraph plus imagined visuals. Mark for trimming. Stopping one sentence earlier is almost always an improvement.

---

## STEP 7 — CONTINUITY CHECK

List the cinematic objects introduced. Confirm at least one returns in transformed state in a later chapter. If the transformed return is also a verbatim scene callback (same character, same gesture), propose a transformation that changes the form, not just the location.

---

## OUTPUT

Return findings as a structured edit list. Each item:
- Location
- The pattern flagged
- Why it matters
- A proposed rewrite, or a question for the writer if the call requires judgment

Do **not** rewrite the manuscript yourself. The editor pass produces an edit list. The writer (or a separate revision pass) applies it.

---

## EDITOR'S FINAL RULE

The system this manuscript was generated under is mostly working. Your job is not to dismantle it. Your job is to find the moves that worked the first time and were reused until they became signatures — and to break the pattern on the second or third recurrence.

A manuscript that passes every rule and still sounds like itself across ten chapters is the failure mode. Diversity of shape, across chapters, is the prestige signal.
""",


    "youtube": """
STYLE: YouTube -- Popular Science
= Hook in first sentence. Curiosity loops. Reframe what viewer thinks they know.

STRUCTURE PER CHAPTER:
1. HOOK -- First sentence must be a question, surprising fact, or counterintuitive claim.
   GOOD: "Apa yang kamu anggap 'makanan Indonesia' sebenarnya datang dari perahu yang berlayar 4.000 tahun lalu."
2. SETUP THE MYSTERY -- What's the weird thing we're about to explain? Why should they keep watching?
3. EXPLAIN WITH ANALOGY -- One modern analogy per complex concept. Make the ancient feel familiar.
4. PAYOFF + REFRAME -- Answer the question, then add "...and here's what that means for you today."

FORBIDDEN: Academic tone. Passive voice. Long blocks without a hook or punchline.
REQUIRED: Short punchy sentences mixed with longer ones. Direct address ("kamu", "kalian"). At least one modern analogy.
""",

    "journalistic": """
STYLE: Journalistic -- Long Form
= Report the past like a journalist covering a breaking story. Sources, scenes, quotes, stakes.

STRUCTURE PER CHAPTER:
1. LEAD -- The most important/surprising fact first. Then context.
2. NUT GRAF -- What is this chapter really about? Why does it matter?
3. SCENE + VOICE -- At least one reconstructed scene + one "quoted" source (archaeologist, historical record, oral tradition).
4. MULTIPLE ANGLES -- Show competing interpretations. What do scholars disagree about?

FORBIDDEN: Single narrative voice without tension. Unverified claims presented as fact.
REQUIRED: Attribution language ("menurut penelitian...", "arkeolog menemukan..."). Present tense for dramatic reconstruction. Specific numbers and sources.
""",

    "literary essay": """
STYLE: Literary Essay
= Personal intellectual voice. Digressive. Thinking on the page, not presenting conclusions.

STRUCTURE PER CHAPTER:
1. PERSONAL/ASSOCIATIVE OPENING -- Start with an observation, memory, or cultural reference that connects to the topic obliquely.
2. DIGRESSION -- Follow one idea sideways before returning to the main thread.
3. COMPLEXITY -- Resist simple conclusions. Show what we don't know. Sit with the ambiguity.
4. RESONANT CLOSE -- End not with a conclusion but with a lingering image or open question.

FORBIDDEN: Thesis statements. Bullet-point logic. Authoritative declarations.
REQUIRED: First-person or intimate narrator voice. Cultural and literary references. Sentences that think out loud.
""",

    "podcast narrative": """
STYLE: Podcast Narrative
= Written for the ear, not the eye. Conversational, signposted, built on spoken rhythm.

STRUCTURE PER CHAPTER:
1. CONVERSATIONAL HOOK -- Address the listener directly. Short sentence to catch attention.
   GOOD: "Coba bayangkan ini."
2. SCENE -- Tell a short story in present tense, as if recounting to a friend.
3. EXPLANATION -- "Nah, inilah yang menarik..." -- signpost the insight clearly.
4. LISTENER TAKEAWAY -- End with "apa artinya ini?" for the listener's life or worldview.

FORBIDDEN: Complex nested sentences. Dense data without analogies. Visual-only descriptions.
REQUIRED: Short sentences (max 20 words each for key points). Signpost phrases. Rhythm that works read aloud.
""",

    "academic popular": """
STYLE: Academic Popular (like Sapiens)
= Big claim -> evidence -> implication. Accessible language for complex ideas. Thought experiments.

STRUCTURE PER CHAPTER:
1. BOLD OPENING CLAIM -- State the argument plainly. No hedging.
   GOOD: "Nusantara bukan korban sejarah -- ia adalah salah satu laboratorium evolusi budaya terbesar yang pernah ada."
2. EVIDENCE STACK -- 3-4 specific data points that support the claim. Studies, sites, percentages.
3. THOUGHT EXPERIMENT -- "Bayangkan jika..." -- use hypothetical to make abstract concrete.
4. IMPLICATION FOR TODAY -- Connect past to present human behavior, society, or culture.

FORBIDDEN: Jargon without definition. Evidence without interpretation. Hedging that kills momentum.
REQUIRED: Footnote-worthy specifics in accessible language. Comparative lens. One thought experiment per chapter.
""",

    "cinematic voiceover": """
STYLE: Cinematic Voiceover
= Written for a narrator's voice over moving images. Short. Punchy. Visual. Rhythmic.

STRUCTURE PER CHAPTER:
1. VISUAL ESTABLISHING LINE -- One sentence, one image. What is the camera seeing?
   GOOD: "Empat ribu tahun yang lalu. Laut Sulawesi. Sebuah perahu bercadik membelah kabut pagi."
2. NARRATION IN SHORT BURSTS -- 2-4 sentence paragraphs max. Pause between images.
3. EMOTIONAL BEAT -- One moment of human connection. Keep it brief.
4. TITLE CARD CLOSE -- End with a short, quotable line. One sentence. Strikes like a title card.

FORBIDDEN: Long complex sentences. Explanatory exposition. Anything that can't be spoken in one breath.
REQUIRED: Present tense. Fragments allowed for rhythm. Powerful monosyllabic words where possible. Visual-first, emotion-second.
""",

        "narrative non-fiction": """
STYLE: Narrative Non-Fiction (Cinematic History)
= Sources: Erik Larson / Robert Caro / Sebastian Junger / Jon Krakauer / Hampton Sides
= Audience: Netflix prestige-doc / PBS NOVA / History Channel premium tier
= Output target: BROADCAST-FINAL on first generation. No revision pass assumed.
= Output format: plain prose, no production markers. No [ANCHOR], [BEAT], or bracketed markup in output.

The audience must feel history unfolding physically in front of them.
Do not explain history from above. Enter from ground level.

The narration must feel: observed, lived-in, materially specific, temporally immersive,
emotionally restrained, confident without sounding performative.

THE WRITER MUST DISAPPEAR BEHIND THE REALITY.

=== LAYER 0 -- NARRATIVE NONFICTION ENGINE ===

# Section 1 -- ENTRY

1.1 SCENE BEFORE THESIS
Never begin with abstraction if a physical scene can carry the idea.
BAD:  "Perdagangan mengubah struktur kekuasaan Nusantara."
GOOD: "Nekara perunggu itu tiba melalui laut. Suaranya terdengar sampai ujung desa."
Every major historical argument must first appear as an object, body, landscape, ritual,
weather condition, practical problem, or physical action.

1.2 PEOPLE ARE NOT METAPHORS
Historical figures are not symbolic delivery systems for arguments.
BAD:  "Seorang anak Atayal sedang mengikat awal diaspora Austronesia."
GOOD: "Tangannya lengket oleh getah kapuk. Ia menarik rotan lebih keras agar bambu pelampung tidak bergeser."
Humans act within immediate realities -- not historical destiny. Meaning comes later.

1.3 MATERIAL REALITY FIRST
Prioritize: weight, distance, hunger, moisture, fatigue, wood, salt, smoke, mud, sound, labor, weather.
Civilizations emerge from physical constraints.

1.4 VARY CHAPTER ENTRY MODES
Rotate openings: object / action / landscape / sound / physical process / archival fragment /
body movement / environmental condition. Never repeat the same opening mode in adjacent chapters.

# Section 2 -- STANCE

2.1 RESTRAIN THE WRITING
Avoid stacked aphorisms, theatrical paradoxes, constant quotable lines, repeated rhetorical reversals.
If a sentence sounds written to impress -> simplify it. The prose should become invisible.

2.2 UNDERWRITE THE IMPORTANT MOMENTS
The more important the historical implication -> the calmer the prose should become.
BAD:    "Bukan kepahlawanan. Bukan takdir."
BETTER: "Sebagian keputusan mereka mungkin terasa biasa pada zamannya."
BAD:    "Namanya hilang selamanya."
BETTER: "Namanya tidak muncul lagi dalam catatan berikutnya."
Rule: If a sentence feels designed to land dramatically -> reduce emotional pressure by 30-50%.

2.3 INVISIBLE AUTHORITY
Sound like someone deeply familiar with the material -- not someone performing expertise.
Avoid over-explaining implications, announcing significance, guiding audience emotion explicitly.
GOOD: "Papirus itu berisi jadwal kerja, distribusi roti, dan pengiriman batu kapur." -- no explanation needed.

2.4 THE WRITER MUST SOMETIMES DISAPPEAR COMPLETELY
Some paragraphs must contain NO thesis, NO philosophy, NO symbolic framing, NO overt implication.
Only: observation, action, sequence, environment. These paragraphs create trust.

2.5 THE NARRATOR MUST NOT SOUND SELF-AWARE
The narrator must never sound like delivering revelations, constructing philosophy,
performing intelligence, building social-media quotes. The narrator simply knows where to look.

2.6 UNDERDIRECT THE EMOTION
Never sound emotionally ahead of the audience. Avoid telling viewers what is moving,
announcing awe, escalating tension artificially, narrating emotional conclusions.
Emotion should arrive late and quietly.

# Section 3 -- INFORMATION FLOW

3.1 INFORMATION MUST FEEL DISCOVERED
Avoid visible exposition. Replace "Penelitian menunjukkan...", "Ini membuktikan...", "Temuan ini mengungkap..."
Embed evidence naturally inside narrative flow.
GOOD: "Di situs Hemudu, arkeolog menemukan bulir padi yang bentuknya sudah membulat. Tangkainya tidak mudah rontok."

3.2 EXPLANATION COMES LATE
Right pattern: SCENE -> detail -> consequence -> historical meaning
Wrong pattern: thesis -> explanation -> illustrative anecdote

3.3 DO NOT COMPLETE THE READER'S THOUGHT
BAD:    "Dan inilah yang menunjukkan kecenderungan universal manusia."
BETTER: "Praktik serupa kemudian muncul di tempat-tempat lain yang terpisah sangat jauh."

3.4 PRESERVE UNCERTAINTY NATURALLY
BAD:  "Para ahli masih memperdebatkan hipotesis ini."
GOOD: "Jejak genetiknya tidak sepenuhnya cocok dengan peta bahasa. Sampai sekarang, para peneliti belum benar-benar sepakat mengapa."

3.5 DO NOT OVER-SUSTAIN SPECULATIVE MATERIAL
Less evidentiary support -> less total narrative time it should dominate.

3.6 EVIDENCE TIERING (inline, never explained to audience)
Mark claims inline without breaking flow:
  Tier 1 -- direct material evidence, contemporary
  Tier 2 -- primary text, contemporary or near-contemporary
  Tier 3 -- later chronicle, hearsay, distorted by time
  Tier 4 -- mythological tradition
  Tier 5 -- accepted scholarly reconstruction
  Tier 6 -- modern hypothesis, not consensus
Narrator epistemic confidence must shift to match tier:
  Tier 1-2 -- declarative, calm authority
  Tier 3-4 -- distanced, reported speech ("Herodotus menulis...")
  Tier 5   -- present scholarly reading ("rekonstruksi yang paling banyak diterima...")
  Tier 6   -- name the proposer, mark as hypothesis, do not adopt as fact

# Section 4 -- PROSE CRAFT

4.1 AVOID RHETORICAL MIRROR STRUCTURES
Avoid: "bukan X, melainkan Y" / "tidak hanya..., tetapi..." / "yang berubah bukan..., melainkan..."
These reveal authorial design. The narration must feel discovered, not engineered.

4.2 DETONATOR SENTENCES ARE RARE
At most ONE isolated emphasis sentence every 700-1,200 words. Most prose flows continuously.

4.3 REDUCE SIGNALING LANGUAGE
Avoid: "yang mengejutkan", "yang luar biasa", "yang paling penting", "yang lebih menarik",
"secara revolusioner", "secara fundamental".
Importance emerges from accumulation, contrast, consequence -- not narrator emphasis.

4.4 REDUCE QUOTABLE-LINE DENSITY
Only 1-2 truly quotable lines per major chapter. Everything else must feel precise, natural, inevitable.

4.5 NATURAL CADENCE OVER PERFECT CADENCE
Allow uneven sentence lengths, practical transitions, quieter connective prose.
Historical reality is slightly irregular -- the prose must preserve that irregularity.

# Section 5 -- VOICE (for the ear)

5.1 WRITE FOR THE EAR
Prose must remain breathable. Avoid overly packed clauses, uninterrupted analytical density,
literary compression without pause points.
Every 2-4 sentences, the prose must create a natural place where an editor could cut away visually.

5.2 BUILD BREATH INTO SENTENCES
Narration must be performable. Alternate: short observation -> medium explanation -> longer reflective sentence.
Good flow: short -> medium -> long -> short
Bad flow:  long -> long -> long -> long

# Section 6 -- IMAGE (writing for cinema)

6.1 VISUALS MUST HAVE SPACE TO EXIST
Do not narrate over every visual moment. Sometimes the image carries the meaning.
Do not resolve meaning faster than the audience can absorb the scene.

6.2 WRITE IN SHOTS, NOT PARAGRAPHS
Each narration segment must imply framing, movement, scale, texture, edit rhythm.
GOOD: "Bekas pukulan itu masih terlihat di Aswan." -- stone texture, close-up, hand movement, dust.

6.3 SILENCE IS PART OF THE NARRATION
After major reveals, discoveries, scale comparisons, physical residue -> allow space.
The audience must occasionally remain alone with the image.

6.4 THE NARRATION MUST FEEL EDITABLE
An editor must be able to cut lines, insert footage, expand visuals -- without damaging prose structure.

6.5 WRITE FOR IMAGE + SOUND + VOICE TOGETHER
Some meaning must remain available for sound design, editing, cinematography.
The film becomes stronger when narration leaves room for itself.

# Section 7 -- PACING

7.1 INFORMATION DENSITY MUST PULSE
After dense explanation -> return to bodies, objects, labor, weather, sound, landscape.

7.2 EVERY MAJOR SECTION NEEDS A DIFFERENT ENERGY SHAPE
Some sections must feel: procedural / observational / intimate / uncertain / expansive /
exhausted / logistical / forensic / silent.

7.3 RECURRING MOTIFS RETURN QUIETLY
Maintain 3-5 recurring motifs (objects, materials, sensory elements) across the full project.
Identify them by chapter 2. Let them return without announcement. Never name them as motifs.

# Section 8 -- STRUCTURE

8.1 LET OBJECTS CARRY THE WEIGHT (paragraph endings)
End paragraphs on objects, gestures, labor, weather, unfinished actions -- not abstract conclusions.
Better endings: "tali rotan yang mulai basah" / "abu dapur yang tertinggal" / "suara nekara yang memantul di lembah"

8.2 SCALE MUST EXPAND QUIETLY
Move from object -> village -> coastline -> migration -> civilization. Gradually.
Do not announce "Implikasinya bagi umat manusia..." Let scale emerge naturally.

8.3 ENDINGS MUST LEAVE RESIDUE
End on: unresolved object, physical trace, remaining silence, changed landscape, practical consequence.
The audience must feel: history continues beyond the frame.

LAYER 0 FINAL RULE:
The writer must disappear behind the reality.
If the writing draws attention to itself -> simplify it.
If a sentence sounds composed -> it is overwritten.
If a paragraph could be screenshotted as wisdom -> break it.

=== LAYER OMEGA-PLUS -- BROADCAST FAILURE PREVENTION ===

Layer Omega-Plus runs DURING generation -- not after.
GENERATE AS IF NO REVISION PASS WILL EVER OCCUR. FIRST OUTPUT MUST BE BROADCAST-FINAL.

# F1 -- TEXT INTEGRITY (no truncation, no artifacts)
Every paragraph: no clipped words, no malformed fragments, no broken sentence continuity.
Hard rule: if a line cannot be read aloud smoothly on first pass -> regenerate it.

# F2 -- CINEMATIC EXPOSITION FILTER (no textbook narration)
Forbidden: "Menurut penelitian...", "Analisis menunjukkan...", "Hal ini membuktikan...",
"Berdasarkan data tersebut...", "Penemuan ini menjelaskan bahwa..."
Replacement: artifact/observation -> implication -> widening mystery
REJECT:   "Analisis isotop menunjukkan para pekerja mendapat nutrisi tinggi."
GENERATE: "Jejak kimia pada tulang mereka menunjukkan sesuatu yang tak terduga. Orang-orang ini tidak diperlakukan seperti tenaga sekali pakai."

# F3 -- MOMENTUM BRIDGE ENFORCER (no abrupt transitions)
Whenever moving between discovery -> interpretation / evidence -> implication / person -> system /
answer -> new mystery -> insert a narrative bridge.
Bridge types (rotate, never repeat in adjacent chapters):
  - "Namun jawaban itu justru membuka pertanyaan yang lebih besar."
  - "Jika manusianya kini mulai terlihat, maka teka-teki berikutnya terletak pada caranya."
  - "Dan di sanalah persoalannya meluas melampaui satu monumen."
  - "Dan benda itu mengubah apa yang kita kira kita tahu."
  - "Tetapi di tengah skala itu, ada satu orang yang bekerja sendirian."
Track bridge type used per chapter -- never adjacent-repeat.

# F4 -- RESONANCE ENDING PROTOCOL (no informational chapter endings)
Never end with: summary / factual recap / thesis restatement / academic conclusion.
End in one of these modes:
  - Reflective scale expansion: object -> civilization -> human capability
  - Quiet unresolved awe: leave one implication suspended
  - Historical inversion: new evidence changes old assumptions
  - Philosophical afterglow: suggest significance without declaring it
Target: quiet intellectual wonder. History has become larger, stranger, and more human.

# F5 -- PERFORMANCE AUDIT (measured, not self-attested)
Before finalizing each chapter, calculate:
  Sentence distribution: short (<=5w) 20% / medium (6-15w) 55% / long (16-25w) 25% (+/-5%)
  Any sentence > 25 words: split or rewrite -- zero tolerance
  Mirror structures ("bukan X, melainkan Y"): <= 2 per chapter, descriptive only
  Detonator density (isolated short paragraphs): 1 per 700-1,200 words
  Signaling language: zero
Then ask: would a senior narrator stumble? Does rhythm breathe? Would editor need to patch? Does final line linger?
If any metric violated or any answer is "no" -> regenerate affected section.

=== DOCUMENTARY VOICEOVER ENGINEERING ===

Narration is written for EARS first. Output is plain prose -- no production markers.

1. PERFORMANCE RHYTHM
Write with: breath spacing, silence points, tonal descent, escalation, pause architecture.
Every sentence must be speakable in one breath or with deliberate pausing.

2. MICRO-SENTENCE IMPACT
After dense information passages, deploy short sentences as psychological percussion.
TECHNIQUE: [Long analytical sentence]. [Short sentence. Stops everything.]
EXAMPLE: "Sistem irigasi itu melibatkan lebih dari empat ribu pekerja selama tiga musim tanam.
          Tidak ada satu pun nama mereka yang tercatat."
The short sentence lands BECAUSE the preceding sentence was long.

3. VISUAL EDIT SPACE
Narration must leave room for image and music. Do not suffocate the edit.
Insert: silence windows (end a thought, let image carry), atmospheric pauses (one-sentence
breathing space for the editor), visual handoff moments (describe what camera can show, then stop).

4. SILENCE POINTS (write into prose structure, not as markers)
After named discoveries, before structural transitions, after object reveals, at chapter closings:
allow white space -- a short sentence or paragraph break that gives editor and composer room.
Double paragraph break = 1.5-2 second silence window. Use after major mystery beats.

5. VISUAL TRIGGER LINES (3-5 per chapter)
Each chapter must contain lines that immediately cue a visual cut:
  Material residue: "Bekas pukulan itu masih terlihat di Aswan."
  Located object:   "Papirus itu kini tersimpan di Museum Mesir, Kairo."
  Specific gesture: "Senter di tangannya menyapu dinding."
  Located action:   "Tahun 1990. Di pinggiran dataran Giza, seekor kuda tersandung."
Editor must read the line and know what shot to cut to.

6. SENTENCE WAVEFORM TARGET
  Short impact lines (<=5w):    20% (+/-5%)
  Medium analysis lines (6-15w): 55% (+/-5%)
  Long reflective lines (16-25w): 25% (+/-5%)
Flat distribution = narrator fatigue. Waveform = endurance.

7. PROPER NOUN GLOSS (listener cannot pause to look up)
First mention of any proper noun or technical term: add inline micro-gloss of 2-5 words.
  WRONG: "Wadi al-Jarf"
  RIGHT: "Wadi al-Jarf -- pelabuhan kuno di tepi Laut Merah"
First mention of person: Name + role in 2-4 words. Subsequent: last name only, no re-gloss.

8. READ-TIME AUDIT
Calculate estimated runtime at 130 WPM per chapter. State explicitly.
Over runtime -> compress textbook lines first (F2), never cinematic ones.
Under runtime -> add atmospheric reset paragraphs, never add thesis.

FINAL CHECK before each paragraph:
  Can a narrator speak this in one breath? (max 25 words per sentence)
  Three consecutive profound sentences? -> make one plain or atmospheric
  Unfamiliar proper noun? -> gloss it inline
  Two aphorisms in one paragraph? -> remove the weaker one
  Does paragraph ending land on object/gesture/labor/weather -- not abstraction?
  Does this chapter have 3-5 visual trigger lines?
  Does the final line linger after silence?

=== MASTER EXECUTION DIRECTIVE ===

Generate each narration as if no revision pass will ever occur.
The first output must already be broadcast-final.

LAYER 0 makes the writing honest.
LAYER OMEGA-PLUS makes the writing broadcastable on first generation.

FORBIDDEN: Textbook exposition. Universalizing framing ("kecenderungan manusia", "setiap peradaban").
Self-aware narrator. Mirror structures as default rhythm. Quotable-line stacking.
Detonator overuse. Abstract chapter closings. Any bracketed production markup in output.

REQUIRED: Scene before thesis. Material reality first. Writer disappears. Evidence tiering inline.
Bridge tracking across chapters. Sentence waveform. Resonance endings. Visual trigger lines.
Editable cinematic rhythm. Plain prose output -- no markers, no scaffolding, no stage directions.
""",
}

VIDEO_SCRIPT_MODIFIER = """
=== VO GENERATION MODE — WRITE FOR EARS, NOT EYES ===

This chapter is being written DIRECTLY as documentary narration.
Do not write prose first and convert later.
Write each sentence as if a narrator is about to speak it aloud — once, without re-read.

=======================================================================
LAYER 1 — CONTENT (HARARI/DIAMOND RULES — already in style guide above)
These are active. What you say must follow the Big History framework.
This layer is not repeated here. Apply it silently.
=======================================================================

=======================================================================
LAYER 2 — DELIVERY (VO ENGINEERING — how you say it)
These rules govern sentence construction, rhythm, and breath architecture.
=======================================================================

── RULE 1: ONE IDEA PER BREATH ──
A narrator breathes every 15–20 words.
Any clause over 25 words must become two sentences.
Do NOT lower the register — shorten the breath unit, not the vocabulary.
  WRONG: "Ibu petani Hemudu? Dua tahun sudah cukup." (conversational)
  RIGHT: "Ibu petani Hemudu menyapih di usia dua tahun. Jarak kelahiran memendek." (formal, two breaths)

── RULE 2: SENTENCE HIERARCHY — MANDATORY RATIO ──
For every 2–3 profound/aphoristic sentences, insert 1 of the following:
  PLAIN: simple declarative, no metaphor, just fact. Resets the ear.
  BREATHING: one observation, present tense or slow rhythm, atmospheric.
  OBSERVATIONAL: concrete sensory detail — what you SEE or HEAR, not what it MEANS.

Max ONE aphoristic/quotable sentence per paragraph.
Never two consecutive profound sentences. The second one cancels the first.

── RULE 3: RHYTHM VARIATION ──
Vary length deliberately: LONG → SHORT → LONG → SHORT → VERY SHORT.
The very short sentence (3–7 words) is the detonator. Place it after buildup.
  EXAMPLE: "Sistem irigasi itu melibatkan lebih dari empat ribu pekerja selama tiga musim tanam.
            Tidak ada satu pun nama mereka yang tercatat."

── RULE 4: [BEAT] MARKERS ──
After every major revelation or emotional peak: insert [BEAT].
The sentence before [BEAT] must be SHORT — the landing strip, not the runway.
[BEAT] signals: pause here, let image carry, cut to visual.

── RULE 5: ANCHOR LINES ──
Every chapter must have 3–5 ANCHOR lines.
Criteria: under 12 words, standalone quotable, paradox or reversal structure, emotionally irreversible.
Mark each with [ANCHOR].
Anchor lines are VERBATIM — do not edit them during revision. Build everything else around them.

── RULE 6: PROPER NOUN GLOSS ──
A listener cannot pause to google. They hear it once.
For every proper noun or technical term on first mention: add an inline micro-gloss of 2–5 words.
  WRONG: "situs Hemudu, delta Sungai Yangtze"
  RIGHT: "situs Hemudu — desa kuno di delta Sungai Yangtze"
  WRONG: "haplogroup Q dan P"
  RIGHT: "penanda genetik yang disebut haplogroup Q dan P"

For person names — first mention only: [Name], [role in 2–4 words].
  RIGHT: "Robert Blust, linguis dari Universitas Hawaii"
  WRONG: "seorang linguis bernama Robert Blust"
Subsequent mentions in same chapter: last name only, no re-gloss.

── RULE 7: CITATION INTEGRATION ──
Academic citation must speak naturally, not parenthetically.
  BANNED: "...sebagaimana dikemukakan Blust (1999, AO 34:2)..."
  RIGHT: "...Robert Blust, dalam empat dekade rekonstruksi leksikalnya, menemukan..."

── RULE 8: CHAPTER CLOSING ──
The final [ANCHOR] of each chapter must:
  a) Not echo or restate the paragraph immediately before it
  b) Open a door — temporal shift, new image, or consequence that pulls forward

── RULE 9: NO ECHO CLOSING ──
  WRONG: paragraph says "ketiadaan" → anchor says "Yang tersisa hanyalah ketiadaan."
  RIGHT: paragraph says "ketiadaan" → anchor says something entirely new that opens the next scene.

── RULE 10: PARAGRAPH SIZE ──
Maximum 4–5 sentences per paragraph for VO.
White space is pacing. Use it.
After a data-dense paragraph: mandatory 1–2 sentence atmospheric reset before continuing.

── RULE 11: FORBIDDEN IN VO ──
- Complex nested clauses with 3+ subordinate levels
- Abstract tangents that cannot be visualized by the listener
- Two aphorisms in the same paragraph
- Anchor echoing the paragraph before it
- Dropping register to conversational when breaking long sentences

=======================================================================
FINAL TEST before writing each paragraph:
  □ Can a narrator speak this in one breath? (max 25 words per sentence)
  □ Three consecutive profound sentences? → make one plain
  □ Unfamiliar proper noun? → gloss it inline
  □ Two aphorisms in one paragraph? → remove the weaker one
  □ Did register drop to conversational? → restore formal tone
  □ Is the chapter's final line an ANCHOR that opens the next chapter?
  □ ANCHOR count for this chapter: 3–5?
=======================================================================
"""


def get_style_rules(style: str, video_mode: bool = False) -> str:
    style_lower = style.lower()
    rules = ""
    for key in STYLE_RULES:
        if key in style_lower:
            rules = STYLE_RULES[key]
            break
    if not rules:
        rules = STYLE_RULES.get("creative non-fiction", "")
    if video_mode:
        # VIDEO_SCRIPT_MODIFIER now contains full VO delivery engineering rules.
        # Append AFTER content rules so delivery layer wraps content layer.
        rules += VIDEO_SCRIPT_MODIFIER
    return rules


def get_generation_preamble(video_mode: bool = False) -> str:
    """Return a short preamble injected at the TOP of every chapter generation prompt.
    For VO mode: signals the model to write for ears from sentence one.
    For normal mode: signals standard written prose.
    """
    if video_mode:
        return (
            "CRITICAL: You are writing DOCUMENTARY NARRATION — text that will be spoken aloud "
            "by a narrator, heard once, and felt immediately. "
            "Do NOT write written prose and convert it. "
            "Write each sentence as spoken language at formal/literary register. "
            "Apply [ANCHOR] and [BEAT] markers as instructed in the VO rules below.\n\n"
        )
    return ""



# ---------------------------------------------------------------------------
# RAG context endpoint — called by server.js Google path
# ---------------------------------------------------------------------------
@app.post("/rag/context")
async def rag_context(body: dict):
    """
    Retrieve Gutenberg passages for a topic and return formatted context block.
    Called by server.js before Google API chapter generation.
    """
    if not RAG_AVAILABLE:
        return {"ok": False, "context_text": "", "sources": [], "passages": 0}

    topic   = (body.get("topic")   or "").strip()
    style   = (body.get("style")   or "epic").strip()
    top_k   = int(body.get("top_k") or 5)

    try:
        from moat.gutenberg.rag_narration import get_narration_context
        from moat.gutenberg.style_rag_config import get_style_config as _get_cfg
        rag_style = _rag_style(style)
        _cfg = _get_cfg(rag_style) if rag_style is not None else {
            "style_filter": None, "structure_filter": None,
            "min_quality": 3, "top_k": top_k, "query_instruction": None}
        ctx = await get_narration_context(
            topic=topic,
            style=_cfg.get("style_filter"),
            structure=_cfg.get("structure_filter"),
            min_quality=_cfg.get("min_quality", 3),
            top_k=top_k,
            query_instruction=_cfg.get("query_instruction"),
            prefer_source=os.environ.get("RAG_PREFER_SOURCE") or None,
        )
        _passages = ctx.get("passages", [])
        return {
            "ok":           True,
            "context_text": ctx.get("context_text", ""),
            "sources":      ctx.get("sources", []),
            "passages":     len(_passages),
            "passage_ids":  [(p.get("passage_id") or p.get("id"))
                             for p in _passages
                             if (p.get("passage_id") or p.get("id"))],
        }
    except Exception as exc:
        return {"ok": False, "context_text": "", "sources": [], "passages": 0, "error": str(exc)}


@app.post("/narasi/outline")
async def narasi_outline(body: dict,
                         user: CurrentUser = Depends(get_current_user)):
    """Generate or revise a narrative outline with chapter weights."""
    import traceback as _tb
    try:
        return await _narasi_outline_impl(body)
    except HTTPException:
        raise
    except Exception as e:
        _tb.print_exc()
        raise HTTPException(500, f"{type(e).__name__}: {e}")


# Maps a language code to a display name; otherwise passes the value through
# AS-IS so any language/register the LLM supports works (jv, su, "Darija Maroko",
# "Basa Jawa Krama", …). Mirrors resolve_language() in rag_narration.py.
_NARASI_LANG_NAMES = {
    "id": "Bahasa Indonesia", "en": "English",
    "jv": "Basa Jawa (Javanese)", "su": "Basa Sunda (Sundanese)",
    "ms": "Bahasa Melayu (Malay)", "ban": "Basa Bali (Balinese)",
    "min": "Baso Minangkabau", "ar": "العربية (Arabic)",
    "zh": "中文 (Chinese)", "ja": "日本語 (Japanese)", "ko": "한국어 (Korean)",
    "es": "Español (Spanish)", "fr": "Français (French)", "de": "Deutsch (German)",
    "nl": "Nederlands (Dutch)", "pt": "Português (Portuguese)",
    "hi": "हिन्दी (Hindi)", "th": "ภาษาไทย (Thai)", "vi": "Tiếng Việt (Vietnamese)",
    "tl": "Tagalog (Filipino)",
}


def _resolve_narasi_lang(language: str) -> str:
    if not language:
        return "Bahasa Indonesia"
    return _NARASI_LANG_NAMES.get(language.strip().lower(), language.strip())


async def _narasi_outline_impl(body: dict):
    action = body.get("action", "outline")
    model = (body.get("model") or "gemini-2.5-flash").strip()
    topic = (body.get("topic") or "").strip()
    style = (body.get("style") or "storytelling").strip()
    language = (body.get("language") or "id").strip()
    word_min = int(body.get("word_min") or 4000)
    word_max = int(body.get("word_max") or 4500)
    chap_count = int(body.get("chap_count") or 5)
    revise_instruction = (body.get("revise_instruction") or "").strip()
    current_outline = (body.get("current_outline") or "").strip()
    outline_for_brief = (body.get("outline") or "").strip()
    client = make_client(model)
    lang_label = _resolve_narasi_lang(language)
    # Resolve tenant + user UUID once for usage logging (auth dependency set the context)
    _ou_ctx = _tenant_ctx.get()
    _ou_tenant = _ou_ctx.tenant_id or None
    _ou_user = (await _resolve_user_uuid(_ou_ctx.tenant_id, _ou_ctx.user_id)) if _ou_ctx.user_id else None

    video_mode_outline = bool(body.get("video_mode", False))
    vo_note = (
        "\n\nVO GENERATION MODE — ACTIVE: This outline will be used to generate DOCUMENTARY NARRATION. "
        "For each chapter description, specify:\n"
        "  (a) the recommended OPENING TYPE (A/B/C/D/E/F/G from the style guide) — vary across chapters\n"
        "  (b) one suggested ANCHOR line concept (short, paradoxical, standalone quotable)\n"
        "  (c) which technical terms or proper nouns will need micro-gloss on first mention\n"
        "Include these VO notes in the description field, separated by a pipe | character.\n"
    ) if video_mode_outline else ""

    if action == "brief":
        vo_brief_note = (
            "\n- For VO mode: describe the breath architecture — where the narrator should accelerate "
            "vs. slow down, and which chapter's emotional peak should carry the longest silence."
        ) if video_mode_outline else ""
        user = (
            f"You are writing a {style} narrative titled: \"{topic}\"\n"
            f"Language: {lang_label}\n\nOutline:\n{outline_for_brief}\n\n"
            f"Write a concise NARRATIVE BRIEF (max 300 words) covering:\n"
            f"- Overall tone, voice, and emotional arc\n"
            f"- Key themes and recurring motifs\n"
            f"- How chapters should connect and flow into each other\n"
            f"- Any characters, time periods, or concepts that appear across multiple chapters"
            + vo_brief_note + "\n\n"
            f"Write the brief in {lang_label}. Return ONLY the brief text, no headings, no markdown."
        )
        resp = client.chat.completions.create(model=model, messages=[{"role": "user", "content": user}],
                                              max_tokens=1000, stream=False)
        await _log_narasi_usage(_ou_tenant, _ou_user, model, resp)
        return {"ok": True, "brief": resp.choices[0].message.content.strip()}

    if revise_instruction and current_outline:
        user = (
            f"Revise this narrative outline for: \"{topic}\"\nStyle: {style} | Language: {lang_label}\n"
            f"Word range: {word_min}-{word_max} words total\n\nCURRENT OUTLINE:\n{current_outline}\n\n"
            f"REVISION INSTRUCTIONS:\n{revise_instruction}\n\n"
            f"Apply the revision. Redistribute word counts so total stays {word_min}-{word_max} words, "
            f"heavier chapters get more words.\n\n"
            f"Return ONLY a valid JSON object with:\n"
            f"  \"chapters\": array, each with: \"id\" (string), \"title\" (in {lang_label}), "
            f"\"description\" (1-2 sentences), \"words\" (integer)\n"
            f"  \"outline_text\": full outline as clean markdown\nNo fences, no explanation."
            + vo_note
        )
    else:
        user = (
            f"Create a detailed narrative outline for a {style} narrative titled: \"{topic}\"\n"
            f"OUTPUT LANGUAGE: {lang_label}. ALL chapter titles, descriptions, and outline_text "
            f"MUST be written in {lang_label}.\n"
            f"Language: {lang_label} | Total words: {word_min}-{word_max} | Chapters: exactly {chap_count}\n\n"
            f"WORD WEIGHT RULES -- CRITICAL:\n"
            f"- Do NOT divide words equally across chapters\n"
            f"- Chapters with deeper/complex/climactic topics get MORE words\n"
            f"- Intro and conclusion get FEWER words\n"
            f"- Total must sum to a value between {word_min} and {word_max}\n\n"
            f"Return ONLY a valid JSON object with:\n"
            f"  \"chapters\": array of exactly {chap_count} objects, each with:\n"
            f"    \"id\": chapter number as string\n"
            f"    \"title\": chapter title in {lang_label}\n"
            f"    \"description\": 1-2 sentence summary, written in {lang_label}\n"
            f"    \"words\": integer word count weighted by topical depth\n"
            f"  \"outline_text\": the full outline as clean markdown\n"
            f"No markdown fences, no explanation. Just the JSON."
            + vo_note
        )

    max_tok = max(4000, chap_count * 600 + 2000)
    resp = client.chat.completions.create(model=model, messages=[{"role": "user", "content": user}], max_tokens=max_tok,
                                          stream=False)
    await _log_narasi_usage(_ou_tenant, _ou_user, model, resp)
    raw = resp.choices[0].message.content.strip()
    raw = _re.sub(r"^```(?:json)?\s*", "", raw, flags=_re.MULTILINE)
    raw = _re.sub(r"\s*```\s*$", "", raw, flags=_re.MULTILINE).strip()

    def _clean_json(s):
        # strip trailing commas before ] or }
        return _re.sub(r",([\s\r\n]*[}\]])", r"\1", s).strip()

    def _extract_chapters(data):
        if isinstance(data, list): return data
        if isinstance(data, dict):
            for key in ("chapters", "outline", "bab"):
                if isinstance(data.get(key), list): return data[key]
            for v in data.values():
                if isinstance(v, dict):
                    for key in ("chapters", "outline", "bab"):
                        if isinstance(v.get(key), list): return v[key]
                if isinstance(v, list) and v and isinstance(v[0], dict) and "id" in v[0]:
                    return v
        return None

    def _try_parse(s):
        try:
            d = json.loads(_clean_json(s))
            ch = _extract_chapters(d)
            if ch:
                ot = d.get("outline_text", "") if isinstance(d, dict) else ""
                return {"ok": True, "chapters": ch, "outline_text": ot}
        except Exception:
            pass
        return None

    result = _try_parse(raw)
    if not result:
        m = _re.search(r"\{[\s\S]+\}", raw)
        if m: result = _try_parse(m.group())
    if not result:
        m = _re.search(r"\[[\s\S]+\]", raw)
        if m: result = _try_parse(m.group())
    if not result:
        raise HTTPException(500, f"Tidak bisa parse outline -- raw: {raw[:400]}")

    # ── ENFORCE word count — never trust AI ──
    _chs = result["chapters"]
    if _chs:
        _total = sum(int(c.get("words", 0)) for c in _chs)
        if _total < word_min or _total > word_max:
            _ratio = word_min / _total if _total > 0 else 1
            for c in _chs:
                c["words"] = max(50, int(int(c.get("words", 0)) * _ratio))
            _diff = word_min - sum(int(c.get("words", 0)) for c in _chs)
            if _diff:
                max(_chs, key=lambda c: c.get("words", 0))["words"] += _diff
        result["chapters"] = _chs

    # ── Fallback outline_text if AI left it empty ──
    if not result.get("outline_text", "").strip():
        _ot = []
        for c in result["chapters"]:
            _ot.append(f"## Bab {c.get('id','??')}: {c.get('title','')}")
            _ot.append(f"{c.get('description','')}")
            _ot.append(f"*Target: {c.get('words',0)} kata*\n")
        result["outline_text"] = "\n".join(_ot)

    # ── Moat capture: store the outline as a creative-chain artifact ──
    # (topic → outline draft). Best-effort, never blocks the response.
    try:
        _octx = _tenant_ctx.get()
        _outline_user = (await _resolve_user_uuid(_octx.tenant_id, _octx.user_id)) if _octx.user_id else None
        await db.save_outline(
            _octx.tenant_id or None,
            _outline_user,
            topic, style, language, chap_count,
            result.get("outline_text", ""),
            result.get("chapters", []),
            model)
        await _log_narasi_usage(_octx.tenant_id, _outline_user, model, resp)
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("save_outline/usage failed (non-fatal): %s", _e)

    return result


# BEFORE
"""
@app.post("/narasi/generate")
async def narasi_generate(body: dict):
    model    = (body.get("model")    or "gemini-2.5-flash").strip()
    topic    = (body.get("topic")    or "").strip()
    style    = (body.get("style")    or "storytelling").strip()
    language = (body.get("language") or "id").strip()
    chapters = body.get("chapters")  or []
    ...
"""
@app.post("/narasi/generate")
async def narasi_generate(body: dict,
                          user: CurrentUser = Depends(get_current_user)):
    # Resolve tenant/user from auth (reliable) and pass into the background task.
    _tenant = user.tenant_id
    _user   = await _resolve_user_uuid(user.tenant_id, user.user_id)

    chapters = body.get("chapters") or []
    topic    = (body.get("topic") or "").strip()
    job_id = (body.get("pre_job_id") or str(uuid.uuid4())[:8])[:16]

    # Create the jobs-table row up front so polling can see it immediately.
    try:
        await db.create_narasi_job(_tenant, _user, job_id, topic, len(chapters))
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("create_narasi_job failed (non-fatal): %s", _e)
    await rc.set_progress(job_id, "Memulai narasi...")

    # Spawn the actual generation on the main loop; return the id immediately.
    asyncio.create_task(_narasi_generate_impl(body, job_id, _tenant, _user))
    return {"ok": True, "job_id": job_id, "status": "started"}


async def _narasi_generate_impl(body: dict, job_id: str, _narasi_tenant, _narasi_user):
    model    = (body.get("model")    or "gemini-2.5-flash").strip()
    topic    = (body.get("topic")    or "").strip()
    style    = (body.get("style")    or "storytelling").strip()
    language = (body.get("language") or "id").strip()
    chapters = body.get("chapters")  or []
    use_rag  = bool(body.get("use_rag", False)) and RAG_AVAILABLE 
    brief = (body.get("brief") or "").strip()
    outline = (body.get("outline") or "").strip()
    lang_label = _resolve_narasi_lang(language)
    tmp_dir = Path(f"/app/data/narasi_temp/{job_id}")
    tmp_dir.mkdir(parents=True, exist_ok=True)
    client = make_client(model)
    errors = []

    # Cross-container cancel lives in Redis (cancel:narasi_{job_id}).
    # cancel_ev is kept only for the local auto-cancel path below.
    cancel_ev = threading.Event()

    auto_cancelled = False
    bab1_words = None
    previous_chapters = []   # accumulate generated text for inter-chapter context
    # Resolve internal jobs.id (UUID) once for usage logging — narasi uses the
    # external 8-char id everywhere, but usage_logs.job_id FKs to jobs.id.
    _narasi_job_uuid = None
    try:
        _jrow = await db.get_job_by_external(_narasi_tenant, job_id)
        _narasi_job_uuid = _jrow.get("id") if _jrow else None
    except Exception:
        _narasi_job_uuid = None
    for i, chapter in enumerate(chapters):
        # Check cancel before each chapter (local auto-cancel OR Redis flag)
        if cancel_ev.is_set() or await rc.is_cancelled(f"narasi_{job_id}"):
            errors.append({"id": "cancelled", "error": "Job cancelled by user"})
            break

        chap_id = chapter.get("id", "?")
        chap_title = chapter.get("title", "")
        chap_desc = chapter.get("description", "")
        word_target = int(chapter.get("words") or 400)
        word_min = int(word_target * 0.9)
        word_max = int(word_target * 1.1)
        try:
            video_mode = bool(body.get("video_mode", False))
            style_rules = get_style_rules(style, video_mode)
            preamble    = get_generation_preamble(video_mode)

            # RAG: retrieve Gutenberg passages for this chapter's topic
            rag_context_text = ""
            import logging as _log_rag
            _rag_log = _log_rag.getLogger("rag_narration")
            if use_rag:
                from moat.gutenberg.rag_narration import get_narration_context, build_rag_prompt
                rag_style = _rag_style(style)
                try:
                    from moat.gutenberg.style_rag_config import get_style_config as _get_style_cfg
                    if rag_style is not None:
                        _cfg = _get_style_cfg(rag_style)
                    else:
                        # Broad genre (narrative non-fiction etc.) — no style filter
                        _cfg = {"style_filter": None, "structure_filter": None,
                                "min_quality": 3, "top_k": 5, "query_instruction": None}
                except (ImportError, KeyError):
                    _cfg = {"style_filter": rag_style, "structure_filter": None,
                            "min_quality": 3, "top_k": 5, "query_instruction": None}

                # Use chapter-specific query, not global topic for every chapter
                rag_query = f"{chap_title}: {chap_desc}" if chap_desc else chap_title

                _rag_ctx = await get_narration_context(
                    topic=rag_query,
                    style=_cfg["style_filter"],
                    structure=_cfg["structure_filter"],
                    min_quality=_cfg["min_quality"],
                    top_k=_cfg["top_k"],
                    query_instruction=_cfg.get("query_instruction"),
                    prefer_source=os.environ.get("RAG_PREFER_SOURCE") or None,
                )
                rag_context_text = _rag_ctx.get("context_text", "")
                _n_passages = len(_rag_ctx.get("passages", []))
                if rag_context_text:
                    _rag_log.info(
                        "generate_rag_narration: bab=%s topic=%r style=%s passages=%d",
                        chap_id, rag_query[:40], rag_style, _n_passages,
                    )
                else:
                    _rag_log.warning(
                        "[RAG] skipped: bab=%s passages=0 style_filter=%r topic=%r — "
                        "Qdrant returned empty (filter too strict or embedding mismatch)",
                        chap_id, rag_style, rag_query[:40],
                    )
            else:
                _rag_log.info(
                    "[RAG] skipped: bab=%s use_rag=False rag_available=%s",
                    chap_id, RAG_AVAILABLE,
                )

            # Build "previously written" context to prevent cross-chapter repetition
            prev_context = ""
            if previous_chapters:
                # Include last 2 chapters max to stay within context window
                recent = previous_chapters[-2:]
                prev_lines = []
                for pc in recent:
                    # Truncate each to ~300 words to save tokens
                    words = pc["text"].split()
                    snippet = " ".join(words[:300]) + ("…" if len(words) > 300 else "")
                    prev_lines.append(f"[Bab {pc['id']}: {pc['title']}]\n{snippet}")
                prev_context = (
                    "PREVIOUSLY WRITTEN CHAPTERS (do NOT repeat ideas, facts, phrases, or metaphors from these):\n"
                    + "\n\n".join(prev_lines)
                    + "\n\n"
                )

            user = (
                    preamble
                    + (rag_context_text + "\n" if rag_context_text else "")
                    + prev_context
                    + f"OUTPUT LANGUAGE: {lang_label}. Write the ENTIRE chapter ONLY in {lang_label}. "
                      f"Any references/context above may be in another language — do NOT mirror them; "
                      f"produce the chapter fully in {lang_label}.\n\n"
                    + f'You are writing Chapter {chap_id} of a {style} narrative titled: "{topic}"\n'
                    + f"Language: {lang_label}\n\n"
                    + style_rules + "\n"
                    + (f"NARRATIVE BRIEF:\n{brief}\n\n" if brief else "")
                    + (f"FULL OUTLINE:\n{outline}\n\n" if outline else "")
                    + f"THIS CHAPTER:\n  Title: {chap_title}\n  Summary: {chap_desc}\n"
                      f"  Target: {word_target} words (range: {word_min}–{word_max})\n\n"
                    + ("Write EXACTLY {word_target} words (count carefully). ".format(word_target=word_target)
                       + ("Include [ANCHOR] and [BEAT] markers — these do NOT count toward the word target.\n"
                          if video_mode else "\n"))
                    + f"Write ONLY in {lang_label}. "
                    + "Do NOT include chapter title/number in output.\n"
                      "Return ONLY the chapter body text. No headings, no markdown, no meta-commentary."
            )
            # Resolve model alias so MODEL_MAX_TOKENS lookup works correctly
            resolved_model = MODELS.get(model, model)
            ceiling = MODEL_MAX_TOKENS.get(resolved_model, DEFAULT_MAX_TOKENS)

            base_tokens = int(word_max * WORDS_TO_TOKENS_NARASI * 1.2) + 1500

            # Gemini 2.5 Pro/Flash via LaoZhang's OpenAI-compatible relay counts
            # thinking tokens against max_tokens (unlike the Google SDK which keeps
            # them in a separate budget).  Without extra headroom the model exhausts
            # the budget on thinking and truncates the actual chapter text.
            # Claude *non-thinking* variants have a hard relay ceiling of ~4096.
            is_claude_thinking = "thinking" in resolved_model
            is_claude_plain    = resolved_model.startswith("claude") and not is_claude_thinking
            thinking_overhead  = THINKING_TOKEN_OVERHEAD if resolved_model in THINKING_MODELS_NARASI else 0

            if is_claude_plain:
                # Plain Claude via LaoZhang: relay caps output at 4096
                safe_max = min(4096, max(4000, base_tokens))
            else:
                safe_max = min(ceiling, max(8000, base_tokens + thinking_overhead))

            resp = client.chat.completions.create(
                model=resolved_model, messages=[{"role": "user", "content": user}],
                max_tokens=safe_max, stream=False
            )
            choice = resp.choices[0]
            text = choice.message.content or ""
            text = text.strip()
            finish = getattr(choice, "finish_reason", "unknown")
            import logging as _log
            _log.warning(f"[narasi] bab {chap_id} finish_reason={finish} words={len(text.split())} model={model}")
            await _log_narasi_usage(_narasi_tenant, _narasi_user, model, resp, job_id=_narasi_job_uuid)
            # Task 4 + Tingkat 4: live progress. current = chapters done so far (i+1).
            _msg = f"Menulis bab {i+1}/{len(chapters)}: {chap_title}"[:200]
            await rc.set_progress(job_id, _msg)
            try:
                await db.update_narasi_progress(_narasi_tenant, job_id, i+1, len(chapters), _msg)
            except Exception as _e:
                import logging as _lg; _lg.getLogger("narasi").warning("update_narasi_progress failed (non-fatal): %s", _e)
            # Retry once if response is empty or too short
            if len(text.split()) < 50:
                _log.warning(f"[narasi] bab {chap_id} EMPTY -- retrying")
                resp2 = client.chat.completions.create(
                    model=resolved_model, messages=[{"role": "user", "content": user}],
                    max_tokens=safe_max, stream=False
                )
                text = (resp2.choices[0].message.content or "").strip()
                await _log_narasi_usage(_narasi_tenant, _narasi_user, model, resp2, job_id=_narasi_job_uuid)
            (tmp_dir / f"{chap_id}.txt").write_text(
                f"## Bab {chap_id}: {chap_title}\n\n{text}\n", encoding="utf-8")
            # Accumulate for inter-chapter context
            previous_chapters.append({"id": chap_id, "title": chap_title, "text": text})

            # ── Step 1.2: persist chapter to narasi_chapters (DB = source of truth) ──
            try:
                _retrieved_ids = []
                if use_rag and rag_context_text:
                    _retrieved_ids = [
                        (p.get("passage_id") or p.get("id"))
                        for p in (_rag_ctx.get("passages") or [])
                        if (p.get("passage_id") or p.get("id"))
                    ]
                await db.save_narasi_chapter(
                    _narasi_tenant, _narasi_job_uuid, i, text,
                    len(text.split()), user, _retrieved_ids,
                    version=1, approved=False)
            except Exception as _e:
                _log.warning("save_narasi_chapter failed (non-fatal): %s", _e)

            # ── Fix 5: moat capture (WS-G Task 5) — store generated narration ──
            try:
                _rag_result = {
                    "rag_used": bool(use_rag and rag_context_text),
                    "sources": None,
                    "passages": (_rag_ctx.get("passages") if (use_rag and rag_context_text) else None),
                    "prompt_used": user,
                    "narration": text,
                }
                _moat_sid = await db.save_moat_session(
                    _narasi_tenant or None,
                    _narasi_user or None,
                    topic, style, _rag_result,
                    model, 0, 0, 0)
                # Stash for the review/save step to attach a correction pair
                (tmp_dir / f"{chap_id}.moat").write_text(str(_moat_sid), encoding="utf-8")
            except Exception as _e:
                _log.warning("moat capture (generate) failed (non-fatal): %s", _e)
            # Auto-cancel if first chapter still < 50 words after retry
            if i == 0:
                bab1_words = len(text.split())
                if bab1_words < 50:
                    auto_cancelled = True
                    cancel_ev.set()
            # Small delay between chapters to avoid rate limiting
            await asyncio.sleep(1)
        except Exception as e:
            errors.append({"id": chap_id, "error": str(e)})
            (tmp_dir / f"{chap_id}.txt").write_text(
                f"## Bab {chap_id}: {chap_title}\n\n<!-- ERROR bab {chap_id}: {e} -->\n",
                encoding="utf-8")

    # Clean up cross-container cancel flag.
    await rc.clear_cancel(f"narasi_{job_id}")
    cancelled = cancel_ev.is_set()

    # ── Task 4: terminal status to jobs table (best-effort) ──
    try:
        if cancelled:
            await db.finish_narasi_job(_narasi_tenant, job_id, "cancelled",
                                       error="Dibatalkan oleh user")
        elif errors and len(errors) >= len(chapters):
            await db.finish_narasi_job(_narasi_tenant, job_id, "error",
                                       error=str(errors[:3]))
        else:
            _stitched_md = "\n\n".join(
                f"## Bab {pc['id']}: {pc['title']}\n\n{pc['text']}"
                for pc in previous_chapters
            )
            await db.finish_narasi_job(_narasi_tenant, job_id, "done", result={
                "chapters": len(chapters),
                "errors": errors,
                "auto_cancelled": auto_cancelled,
                "bab1_words": bab1_words,
                "tmp_dir": str(tmp_dir),
                "markdown": _stitched_md,          # DB = source of truth for combined output
            })
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("finish_narasi_job failed (non-fatal): %s", _e)
    await rc.delete_progress(job_id)
    return  # background task — return value unused


@app.post("/narasi/persist")
async def narasi_persist(body: dict,
                         user: CurrentUser = Depends(get_current_user)):
    """Persist a Google-path (Node) narration run into Postgres: create/find the
    jobs row, write every chapter to narasi_chapters, store stitched markdown in
    jobs.result_payload. Node generates with the user's Google key, then calls
    this so all narasi_chapters writes go through database.py."""
    _tenant = user.tenant_id
    _user   = await _resolve_user_uuid(user.tenant_id, user.user_id)
    job_id  = (body.get("job_id") or str(uuid.uuid4())[:8])[:16]
    topic   = (body.get("topic") or "").strip()
    style   = (body.get("style") or "storytelling").strip()
    chapters = body.get("chapters") or []   # [{index, content, source_prompt, retrieved_ids, word_count, id?, title?}]

    # Idempotent: reuse the job row if one already exists for this external id.
    _row = None
    try:
        _row = await db.get_job_by_external(_tenant, job_id)
    except Exception:
        _row = None
    if not _row:
        try:
            await db.create_narasi_job(_tenant, _user, job_id, topic, len(chapters))
            _row = await db.get_job_by_external(_tenant, job_id)
        except Exception as _e:
            import logging as _lg; _lg.getLogger("narasi").warning("persist create_narasi_job failed: %s", _e)
    _job_uuid = (_row or {}).get("id")

    saved, md_parts = 0, []
    for ch in chapters:
        try:
            _idx  = int(ch.get("index", saved))
            _text = ch.get("content") or ""
            _wc   = int(ch.get("word_count") or len(_text.split()))
            _ids  = list(ch.get("retrieved_ids") or [])
            await db.save_narasi_chapter(
                _tenant, _job_uuid, _idx, _text, _wc,
                ch.get("source_prompt") or "", _ids,
                version=1, approved=False)
            # ── moat capture + usage logging (parity with the LaoZhang path) ──
            try:
                _model = ch.get("model") or "gemini-2.5-flash"
                _ti = int(ch.get("tokens_in") or 0)
                _to = int(ch.get("tokens_out") or 0)
                _cost = _calc_cost(_model, _ti, _to)
                await db.save_moat_session(
                    _tenant or None, _user, topic, style,
                    {"rag_used": bool(ch.get("rag_used")), "sources": None,
                     "passages": _ids, "prompt_used": ch.get("source_prompt") or "",
                     "narration": _text},
                    _model, _ti, _to, _cost)
                await db.log_usage(_tenant, _user, _model, "narasi", _ti, _to, _cost,
                                   job_id=_job_uuid, provider="gemini")
            except Exception as _e2:
                import logging as _lg; _lg.getLogger("narasi").warning("persist moat/usage chapter %s failed (non-fatal): %s", ch.get("index"), _e2)
            md_parts.append(f"## Bab {ch.get('id', _idx)}: {ch.get('title','')}\n\n{_text}")
            saved += 1
        except Exception as _e:
            import logging as _lg; _lg.getLogger("narasi").warning("persist chapter %s failed: %s", ch.get("index"), _e)

    try:
        await db.finish_narasi_job(_tenant, job_id, "done", result={
            "chapters": saved, "source": "google",
            "markdown": "\n\n".join(md_parts),
        })
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("persist finish failed: %s", _e)
    return {"ok": True, "job_id": job_id, "chapters_saved": saved}


@app.post("/narasi/outline/persist")
async def narasi_outline_persist(body: dict,
                                 user: CurrentUser = Depends(get_current_user)):
    """Persist a Google-path (Node) outline into narasi_outlines. Mirror of the
    LaoZhang path's inline db.save_outline so both providers capture the outline
    (research → outline moat artifact). Node calls this after action=outline."""
    _tenant = user.tenant_id
    _user   = await _resolve_user_uuid(user.tenant_id, user.user_id)
    try:
        _chapters = body.get("chapters") or []
        oid = await db.save_outline(
            _tenant or None, _user,
            (body.get("topic") or "").strip(),
            (body.get("style") or "storytelling").strip(),
            (body.get("language") or "id").strip(),
            int(body.get("chap_count") or len(_chapters)),
            body.get("outline_text") or "",
            _chapters,
            body.get("model") or "gemini-2.5-flash")
        return {"ok": True, "outline_id": oid}
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("outline persist failed (non-fatal): %s", _e)
        return {"ok": False, "error": str(_e)}


@app.get("/narasi/jobs")
async def narasi_jobs(user: CurrentUser = Depends(get_current_user)):
    """List the tenant's recent reopenable narasi jobs (Step 1.3 read-back UI)."""
    try:
        return {"ok": True, "jobs": await db.list_narasi_jobs(user.tenant_id, limit=15)}
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("list narasi jobs failed (non-fatal): %s", _e)
        return {"ok": False, "jobs": [], "error": str(_e)}


@app.post("/narasi/rate")
async def narasi_rate(body: dict, user: CurrentUser = Depends(get_current_user)):
    """Record a 1-5 star rating for a chapter → approvals (Step 1.4 moat signal)."""
    chapter_id = (body.get("chapter_id") or "").strip()
    try:
        rating = int(body.get("rating") or 0)
    except Exception:
        rating = 0
    if not chapter_id or rating < 1 or rating > 5:
        return {"ok": False, "error": "chapter_id + rating (1-5) required"}
    _user = await _resolve_user_uuid(user.tenant_id, user.user_id)
    try:
        aid = await db.save_approval(user.tenant_id, _user, chapter_id, rating)
        return {"ok": True, "approval_id": aid, "approved": rating >= 4}
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("rate failed (non-fatal): %s", _e)
        return {"ok": False, "error": str(_e)}


@app.post("/narasi/rate-all")
async def narasi_rate_all(body: dict, user: CurrentUser = Depends(get_current_user)):
    """Rate EVERY chapter of a job with the same 1-5 value ('beri rating narasi')."""
    job_id = (body.get("job_id") or "").strip()
    try:
        rating = int(body.get("rating") or 0)
    except Exception:
        rating = 0
    if not job_id or rating < 1 or rating > 5:
        return {"ok": False, "error": "job_id + rating (1-5) required"}
    _user = await _resolve_user_uuid(user.tenant_id, user.user_id)
    try:
        _row = await db.get_job_by_external(user.tenant_id, job_id)
        if not _row or not _row.get("id"):
            return {"ok": False, "error": "job not found"}
        n = await db.save_approval_all(user.tenant_id, _user, _row["id"], rating)
        return {"ok": True, "rated": n, "approved": rating >= 4}
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("rate-all failed (non-fatal): %s", _e)
        return {"ok": False, "error": str(_e)}


@app.get("/narasi/chapters/{job_id}")
async def narasi_chapters_list(job_id: str, user: CurrentUser = Depends(get_current_user)):
    """Chapters of a job with their latest rating, for the rating UI (Step 1.4)."""
    try:
        _row = await db.get_job_by_external(user.tenant_id, job_id)
        if not _row or not _row.get("id"):
            return {"ok": True, "chapters": []}
        return {"ok": True, "chapters": await db.get_chapters_for_rating(user.tenant_id, _row["id"])}
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("chapters list failed (non-fatal): %s", _e)
        return {"ok": False, "chapters": [], "error": str(_e)}


@app.get("/narasi/status/{job_id}")
async def narasi_status(job_id: str,
                        user: CurrentUser = Depends(get_current_user)):
    """Merge live Redis progress (fast) with the jobs-table row (authoritative).
    Frontend polls this to drive the per-chapter checkbox table."""
    _tenant = user.tenant_id or db._DEV_TENANT_ID
    row = None
    try:
        row = await db.get_job_by_external(_tenant, job_id)
    except Exception:
        row = None
    live = await rc.get_progress(job_id)
    if not row:
        return {"job_id": job_id, "status": "processing" if live else "unknown",
                "progress": live or "", "current": 0, "total": 0, "found": False}
    return {"job_id": job_id,
            "status": row.get("status"),
            "progress": live or row.get("progress_message") or "",
            "current": row.get("progress_current") or 0,
            "total": row.get("progress_total") or 0,
            "result": row.get("result_payload"),
            "error": row.get("error_message"),
            "found": True}


@app.post("/narasi/cancel/{job_id}")
async def narasi_cancel(job_id: str, user: CurrentUser = Depends(get_current_user)):
    """Signal the running narasi job to stop after the current chapter finishes.
    Tenant-scoped: the job must belong to the caller's tenant (RLS)."""
    _row = await db.get_job_by_external(user.tenant_id, job_id)
    if not _row:
        raise HTTPException(404, "job not found")
    await rc.set_cancel(f"narasi_{job_id}")
    return {"ok": True, "status": "cancel_requested", "job_id": job_id}


@app.post("/narasi/review")
async def narasi_review(body: dict, user: CurrentUser = Depends(get_current_user)):
    """Non-streaming editorial review — same pattern as narasi/generate. Auth'd +
    tenant-scoped so the capture (usage + moat) is isolated per tenant."""
    model = (body.get("model") or "gemini-2.5-flash").strip()
    # Rules come from the SERVER persona (by style), never from the client.
    # Backward-compat: only use server persona when a style is actually sent.
    _style = (body.get("style") or "").strip()
    system = ((_review_persona_for(_style).get("system") if _style else body.get("system")) or "You are a helpful editorial assistant.").strip()
    message = (body.get("message") or "").strip()
    max_tokens = int(body.get("max_tokens") or 16000)
    if not message:
        raise HTTPException(400, "message required")

    # Resolve model alias
    resolved = MODELS.get(model, model)
    ceiling = MODEL_MAX_TOKENS.get(resolved, DEFAULT_MAX_TOKENS)
    safe_max = min(max_tokens, ceiling)

    client = make_client(model)
    try:
        resp = client.chat.completions.create(
            model=resolved,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": message},
            ],
            max_tokens=safe_max,
            temperature=0,
            stream=False,
        )
        choice = resp.choices[0]
        text = (choice.message.content or "").strip()
        finish = getattr(choice, "finish_reason", "unknown")
        usage = getattr(resp, "usage", None)
        in_tok = getattr(usage, "prompt_tokens", 0) if usage else 0
        out_tok = getattr(usage, "completion_tokens", 0) if usage else 0
        print(f"[review] model={resolved} finish={finish} in={in_tok} out={out_tok} max={safe_max} temp=0", flush=True)
        # Step 1.5: capture editorial review — usage + the review text as a moat artifact.
        try:
            _rtenant = user.tenant_id
            _ruser = await _resolve_user_uuid(user.tenant_id, user.user_id)
            await _log_narasi_usage(_rtenant, _ruser, model, resp)
            await db.save_moat_session(
                _rtenant, _ruser, (body.get("topic") or "editorial_review"), "editorial_review",
                {"rag_used": False, "sources": None, "passages": None,
                 "prompt_used": (system + "\n\n" + message)[:8000], "narration": text},
                model, in_tok, out_tok, _calc_cost(model, in_tok, out_tok))
        except Exception as _ce:
            import logging as _lg; _lg.getLogger("narasi").warning("review capture failed (non-fatal): %s", _ce)
        return {"ok": True, "text": text, "finish_reason": finish, "usage": {"input": in_tok, "output": out_tok}}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Fix 5: moat capture (WS-G Task 5) — correction pair on edited save ────────
# Call this when the user SAVES an edited narration. The frontend sends the
# original generated text (or the .moat session id from generate) and the final
# edited text; we store the diff as a training pair.
@app.post("/narasi/save-edit/{job_id}")
async def narasi_save_edit(job_id: str, body: dict,
                           user: CurrentUser = Depends(get_current_user)):
    """Persist a human edit as a correction pair. Non-fatal: never blocks save."""
    chap_id        = str(body.get("chap_id") or "")
    original_text  = body.get("original_text") or ""
    corrected_text = body.get("corrected_text") or ""
    style_label    = body.get("style") or None
    topic          = body.get("topic") or None
    duration_min   = body.get("duration_minutes")
    language       = body.get("language") or None

    # Resolve the moat_session id written by generate (if present)
    moat_sid = body.get("moat_session_id")
    if not moat_sid and chap_id:
        try:
            moat_sid = (Path(f"/app/data/narasi_temp/{job_id}") /
                        f"{chap_id}.moat").read_text(encoding="utf-8").strip()
        except Exception:
            moat_sid = None

    if not (original_text and corrected_text):
        return {"ok": False, "reason": "original_text and corrected_text required"}

    try:
        # Resolve the Clerk user id → users.id UUID (save_correction_pair casts
        # user_id to UUID; passing the raw Clerk id silently failed the capture).
        _user = await _resolve_user_uuid(user.tenant_id, user.user_id)
        pair = await db.save_correction_pair(
            moat_sid, user.tenant_id or db._DEV_TENANT_ID, _user,
            original_text, corrected_text,
            style_label, topic, duration_min, language)
        return {"ok": True, "quality_tier": pair.get("quality_tier"),
                "edit_ratio": pair.get("edit_ratio")}
    except Exception as e:
        _logging.getLogger("moat").warning("correction capture failed (non-fatal): %s", e)
        return {"ok": False, "reason": str(e)}


@app.post("/narasi/stitch/{job_id}")
async def narasi_stitch(job_id: str, body: dict,
                        user: CurrentUser = Depends(get_current_user)):
    """Read a job's narration back. Step 1.3: the DB is the source of truth
    (jobs.result_payload markdown → narasi_chapters), so an old job survives a
    redeploy. The temp dir is only a fast cache; we never regenerate."""
    _tenant    = user.tenant_id
    style      = (body.get("style") or "storytelling").strip()
    language   = (body.get("language") or "id").strip()
    lang_label = _resolve_narasi_lang(language)

    body_text = ""
    # ── 1. DB = source of truth: stored stitched markdown, else narasi_chapters ──
    try:
        _row = await db.get_job_by_external(_tenant, job_id)
        if _row:
            _payload = _row.get("result_payload") or {}
            if isinstance(_payload, str):
                try: _payload = json.loads(_payload)
                except Exception: _payload = {}
            _stored_md = ((_payload or {}).get("markdown") or "")
            if _stored_md.strip():
                body_text = _stored_md
            elif _row.get("id"):
                _chs = await db.get_narasi_chapters(_tenant, _row["id"])
                if _chs:
                    body_text = "\n\n".join(
                        f"## Bab {int(c['chapter_index']) + 1}\n\n{c.get('content', '')}"
                        for c in _chs)
    except Exception as _e:
        import logging as _lg; _lg.getLogger("narasi").warning("stitch DB read failed (non-fatal): %s", _e)

    # ── 2. fallback: temp-dir cache (fast; gone after redeploy) ──
    if not body_text.strip():
        tmp_dir = Path(f"/app/data/narasi_temp/{job_id}")
        if tmp_dir.exists():
            import re as _re
            files = sorted(tmp_dir.glob("*.txt"),
                           key=lambda f: [int(c) if c.isdigit() else c for c in _re.split(r'(\d+)', f.stem)])
            body_text = "\n".join(f.read_text(encoding="utf-8") for f in files)

    if not body_text.strip():
        raise HTTPException(404, f"Job {job_id} not found")

    total_words = len(body_text.split())
    markdown = (f"> **Gaya:** {style} | **Bahasa:** {lang_label} | **{total_words} kata**\n\n---\n\n"
                + body_text)
    return {"ok": True, "markdown": markdown, "total_words": total_words}



@app.post("/script/tts")
async def script_to_tts(body: dict):
    """
    Transform a raw script into TTS-ready text with emotion/intonation tags.
    Preserves the original language. Enriches with contextual tone tags.
    """
    script_text = (body.get("script") or "").strip()
    model = (body.get("model") or "gemini-2.5-flash").strip()
    if not script_text:
        raise HTTPException(400, "script is required")

    client = make_client(model)


    system = (
"You are an elite cinematic documentary TTS script editor and narration-performance architect. "
"Your ONLY responsibility is to transform raw narration into a performance-ready voiceover transcript "
"through precision pacing, intonation shaping, cinematic paragraph architecture, and acoustic readability engineering. "
"You are NOT a writer. You are NOT allowed to alter authorship. "
"You must NEVER add, remove, summarize, paraphrase, reinterpret, modernize, simplify, or rewrite ANY original wording.\n\n"

"CORE PHILOSOPHY:\n"
"The narration must sound like lived thought unfolding in real time.\n"
"Not recitation.\n"
"Not acting.\n"
"Not performance poetry.\n"
"The voice must feel intellectually alive.\n"
"Calm, precise, observant, cinematic, and emotionally restrained.\n\n"

"PRIMARY OBJECTIVE:\n"
"Transform flat historical narration into premium cinematic documentary voiceover that sounds like:\n"
"- Netflix historical documentary\n"
"- Premium History Channel narration\n"
"- PBS cinematic nonfiction\n"
"- Harari-style reflective macro-history\n"
"- BBC prestige documentary\n"
"- High-end literary audiobook essay narration\n\n"

"THE NARRATION MUST FEEL:\n"
"- Thoughtful but forward-moving\n"
"- Intelligent without sounding performative\n"
"- Controlled, never melodramatic\n"
"- Atmospheric without becoming theatrical\n"
"- Emotionally restrained but psychologically alive\n"
"- Human and breathing, never robotic\n"
"- Cinematic without sounding scripted\n"
"- Dense with meaning, but effortless to follow\n\n"

"ABSOLUTE NON-NEGOTIABLE RULES:\n"
"- Copy EVERY original word VERBATIM\n"
"- NEVER add new wording\n"
"- NEVER remove wording\n"
"- NEVER replace wording\n"
"- NEVER paraphrase wording\n"
"- NEVER simplify wording\n"
"- NEVER reorder sentences\n"
"- NEVER modify punctuation unless absolutely required for TTS breathing\n"
"- NEVER explain tags or pacing decisions\n"
"- NEVER insert commentary\n"
"- ONLY insert approved tags, pauses, pacing cues, and paragraph breaks\n"
"- Preserve all formatting, dates, terminology, names, and quotations exactly\n"
"- Preserve authorial intelligence and literary rhythm\n\n"

"CRITICAL PHILOSOPHY OF DOCUMENTARY VOICEOVER:\n"
"Documentary narration is written for ears first, eyes second.\n"
"Listeners cannot rewind cognition in real time.\n"
"Therefore pacing must regulate comprehension invisibly.\n"
"Your job is not emotional decoration.\n"
"Your job is cognitive choreography.\n\n"

"COGNITIVE LOAD MANAGEMENT:\n"
"- Dense information must be rhythmically decompressed\n"
"- After analytical passages, create acoustic breathing room\n"
"- Large ideas must land before the next idea begins\n"
"- Do not allow uninterrupted conceptual compression\n"
"- Paragraph breaks should function like invisible editing cuts\n"
"- Long sentences must remain acoustically navigable\n"
"- Avoid consecutive high-density paragraphs without relief\n"
"- The audience must never feel mentally trapped\n\n"

"PARAGRAPH KINETICS:\n"
"- Paragraphs are units of emotional and intellectual momentum\n"
"- Break paragraphs based on shifts in:\n"
"  * scale\n"
"  * tension\n"
"  * revelation\n"
"  * causality\n"
"  * perspective\n"
"  * sensory grounding\n"
"  * philosophical implication\n"
"- Short paragraphs create propulsion\n"
"- Medium paragraphs create flow\n"
"- Long paragraphs should feel immersive and wave-like\n"
"- Never create static pacing\n"
"- Never allow rhythm flattening\n\n"

"SILENCE ARCHITECTURE:\n"
"- Pauses must feel earned\n"
"- Silence is used for realization, scale, contradiction, grief, or implication\n"
"- Never overuse pauses\n"
"- Avoid theatrical silence\n"
"- Pause markers should feel invisible to the listener\n"
"- Short pauses are preferred over dramatic pauses\n"
"- Use pauses primarily after:\n"
"  * historical reversals\n"
"  * existential implications\n"
"  * scale expansion\n"
"  * devastating factual contrast\n"
"  * emotionally irreversible statements\n\n"

"EMOTIONAL GOVERNANCE:\n"
"- Emotion must emerge from facts, not vocal theatrics\n"
"- Historical revelations should feel discovered, not announced\n"
"- Tragedy should remain restrained\n"
"- Awe should be rare and fully earned\n"
"- Mystery should generate forward pull, not horror\n"
"- Reflection should feel observational, not sentimental\n"
"- Philosophical moments should slightly slow without becoming grandiose\n"
"- Human moments should become tactile, intimate, and quieter\n"
"- Never oversell emotion already present in the writing\n"
"- The narrator must trust the material\n\n"

"ANTI-MELODRAMA SAFEGUARDS:\n"
"- Do NOT over-tag emotional beats\n"
"- Do NOT turn historical narration into movie-trailer narration\n"
"- Do NOT create artificial epicness\n"
"- Do NOT sustain intensity too long\n"
"- Constant gravitas destroys gravitas\n"
"- Not every paragraph deserves emotional emphasis\n"
"- Allow neutral informational passages to remain neutral\n"
"- Contrast creates emotional legitimacy\n\n"

"VOICE PERFORMANCE LOGIC:\n"
"- The narrator should sound like an elite historian thinking aloud\n"
"- Slightly lowered voice for dangerous truths or civilizational implications\n"
"- Slight acceleration during narrative movement\n"
"- Slight deceleration during philosophical realization\n"
"- Human details should narrow intimacy naturally\n"
"- Never sound shocked unless the fact itself is genuinely staggering\n"
"- Avoid repetitive emotional cadence patterns\n"
"- The narration must feel dynamically alive across long runtimes\n\n"

"SENTENCE ENERGY HIERARCHY:\n"
"- Not every sentence may sound profound\n"
"- Preserve contrast between:\n"
"  * functional sentences\n"
"  * atmospheric sentences\n"
"  * analytical sentences\n"
"  * impact sentences\n"
"- Protect major lines by surrounding them with restraint\n"
"- If everything sounds monumental, nothing feels monumental\n\n"

"DOCUMENTARY CADENCE ENGINEERING:\n"
"- The narration should feel edited, not written\n"
"- Tags should create acoustic realism, not visible decoration\n"
"- The transcript must still read naturally as prose\n"
"- Avoid rhythmic predictability\n"
"- Vary cadence density across sections\n"
"- Maintain invisible momentum at all times\n"
"- Forward pull is mandatory\n\n"

"TAGGING PHILOSOPHY:\n"
"- Tag emotional intention, not sentence topic\n"
"- Use the minimum number of tags necessary\n"
"- Tags should shape performance invisibly\n"
"- Understatement is preferred over emphasis\n"
"- Restraint creates authority\n"
"- Repetition weakens emotional credibility\n\n"

"ALLOWED STRUCTURAL INSERTIONS:\n"
"- EXACTLY one leading tag per paragraph\n"
"- Optional inline pacing cues\n"
"- Optional cinematic pause markers\n"
"- Optional paragraph restructuring for acoustic readability\n\n"

"AVAILABLE LEADING TAGS:\n"
"[cold open] [information] [reflection] [revelation] [quiet confidence]\n"
"[deadpan] [wonder] [weight] [intimacy] [urgency] [melancholy]\n"
"[gravity] [tenderness] [disbelief] [sadness] [joy] [fear]\n"
"[awe] [suspense] [nostalgia] [determination] [warmth]\n"
"[solemnity] [measured] [observational] [contemplative]\n"
"[historical weight] [quiet realization] [measured disbelief]\n"
"[cognitive shift] [existential reflection] [low reflective]\n"
"[building curiosity] [slow emphasis] [firm transition]\n"
"[soft landing] [narrative acceleration] [controlled intensity]\n"
"[quietly stunned] [literary pause] [measured tension]\n"
"[documentary cadence] [human detail] [philosophical reflection]\n\n"

"OPTIONAL INLINE CUES:\n"
"(pause short)\n"
"(pause beat)\n"
"(slower)\n"
"(lower voice)\n"
"(slight emphasis)\n"
"(measured)\n"
"(hushed)\n"
"(accelerating slightly)\n"
"(softly)\n"
"(firmly)\n\n"

"ADVANCED PERFORMANCE RULES:\n"
"- Inline cues must remain sparse and strategic\n"
"- Never stack multiple cues excessively\n"
"- Avoid emotional redundancy between tags and cues\n"
"- Protect aphoristic lines with acoustic space\n"
"- Use pacing variation to prevent listener fatigue\n"
"- Major revelations require simplification of surrounding cadence\n"
"- Dense historical information should alternate with sensory grounding\n"
"- Every 30–60 seconds of narration should contain some acoustic variation\n"
"- The narration should remain sustainable across hours of listening\n\n"

"FORBIDDEN OUTPUT BEHAVIORS:\n"
"- No markdown\n"
"- No explanations\n"
"- No analysis\n"
"- No commentary\n"
"- No summaries\n"
"- No section labels added by you\n"
"- No emotional overacting\n"
"- No repetitive tag spam\n"
"- No artificial cinematic excess\n"
"- No audiobook fantasy cadence\n"
"- No motivational-speaker rhythm\n"
"- No YouTube clickbait tone\n\n"

"FINAL PERFORMANCE STANDARD:\n"
"The final transcript should sound like:\n"
"- a world-class documentary narrator inside a perfectly edited historical film\n"
"- a historian discovering meaning while speaking\n"
"- an intelligent human voice carrying the weight of evidence\n"
"- calm authority under emotional restraint\n"
"- cinematic realism rather than performance\n\n"

"OUTPUT FORMAT:\n"
"- Paragraphs separated by EXACTLY one blank line (\\n\\n)\n"
"- NEVER use single line breaks between paragraphs\n"
"- EVERY paragraph begins with EXACTLY ONE leading tag\n"
"- Preserve original wording perfectly\n"
"- Return ONLY the transformed transcript\n"
"- No markdown\n"
"- No explanations\n"
"- No extra text"
    )


    input_para_count = len([p for p in script_text.split("\n\n") if p.strip()])
    user = (
        f"Input has {input_para_count} paragraphs. You may split them further but never merge.\n\n"
        f"Transform this script into a tagged TTS transcript:\n\n{script_text}"
    )

    resp = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": user},
        ],
        max_tokens=32000,
        stream=False,
    )
    result = resp.choices[0].message.content.strip()
    result = _re.sub(r"^```[^\n]*\n?", "", result, flags=_re.MULTILINE)
    result = _re.sub(r"\n?```$", "", result, flags=_re.MULTILINE).strip()

    paragraphs = [p.strip() for p in result.split("\n\n") if p.strip()]
    return {"ok": True, "transcript": result, "paragraphs": paragraphs, "count": len(paragraphs)}


@app.post("/flow/storyboard")
async def flow_storyboard(
        req: FlowStoryboardRequest,
        x_image_api_key: Optional[str] = Header(None, alias="X-Image-API-Key"),
):
    """
    Use Gemini to parse a script into cinematic scenes.
    Optionally generate a storyboard image per scene in parallel.
    """
    system_prompt = (
        "You are a professional film director and cinematographer. "
        "Break scripts into detailed, visually rich scene descriptions suitable "
        "for AI video generation. "
        "IMPORTANT: Output ALL JSON field values in English, regardless of the script's language."
    )

    client = make_client(req.chat_model)

    # ── Auto scene count: ask AI how many scenes this script warrants ──
    if req.auto_scene_count or req.scene_count == 0:
        auto_resp = client.chat.completions.create(
            model=req.chat_model or "gemini-2.5-flash",
            messages=[
                {"role": "system", "content": "You are a professional film director. Analyze scripts and determine the optimal number of cinematic scenes needed to visualize the narrative."},
                {"role": "user", "content": (
                    f"Read this script carefully and determine the optimal number of cinematic scenes "                    f"needed to visualize it fully. Consider: narrative beats, location changes, time jumps, "                    f"emotional shifts, and key visual moments. Return ONLY a single integer between 4 and 50. "                    f"No explanation.\n\nScript:\n{req.script}"
                )},
            ],
            max_tokens=10,
            stream=False,
        )
        raw_count = auto_resp.choices[0].message.content.strip()
        try:
            total = max(4, min(50, int(_re.search(r"\d+", raw_count).group())))
        except Exception:
            total = 12  # safe fallback
    else:
        total = max(1, int(req.scene_count))

    def _gen_batch(count: int, offset: int):
        """Generate `count` scenes (positions offset+1 .. offset+count of `total`)
        as one independent AI call with its own token budget."""
        start, end = offset + 1, offset + count
        if total > count:
            scope = (
                f"You are breaking a script into exactly {total} cinematic scenes total "
                f"(numbered 1 to {total}). Generate ONLY scenes {start} to {end} "
                f"({count} scenes) as a coherent part of that full sequence -- they must "
                f"flow naturally from the overall narrative. "
            )
        else:
            scope = f"Break the following script/story into exactly {count} cinematic scenes. "
        style_clause = f" Visual render style: {req.image_style}." if req.image_style else ""
        user_prompt = (
            f"{scope}Use a {req.style} narrative style.{style_clause}\n"
            f"ALL JSON field values must be written in English regardless of the script language.\n\n"
            f"Script:\n{req.script}\n\n"
            f"For each scene return a JSON object with exactly these keys:\n"
            f'  "title": short scene title in English (5-8 words)\n'
            f'  "description": rich English visual prompt for AI image/video generation (60-90 words).'
            f' Read the script carefully and faithfully extract the SPECIFIC location, geography,'
            f' time period, and environmental details described.'
            f' Give equal weight to the landscape, environment, and atmosphere as to any human subject --'
            f' for historical, geographic, or nature-focused scenes the landscape IS the primary subject.'
            f' Cover: specific setting and environment, any human presence and action,'
            f' lighting quality and direction, color palette, mood and atmosphere'
            f'{(", rendered in " + req.image_style + " visual style") if req.image_style else ""}.\n'
            f'  "camera": English technical camera note -- shot size (extreme close-up/close-up/medium/wide/extreme wide),'
            f' angle (eye-level/low/high/overhead), movement (static/slow push-in/dolly out/handheld/crane rise),'
            f' lens character (wide-angle/telephoto compression/shallow focus/anamorphic).\n'
            f'  "audio": English structured sound design prompt -- list specific sounds foreground to background'
            f' in order of prominence with layer tags [FG] [MID] [SCORE] [BG].\n'
            f'  "duration": integer seconds (5 or 8)\n'
            f'  "start_kalimat": the exact opening sentence or phrase (8-15 words) from the ORIGINAL script text'
            f' that this scene is directly based on. Copy verbatim from the script, preserving the original language.\n\n'
            f"Return ONLY a valid JSON array of exactly {count} scene objects. "
            f"No markdown, no explanation."
        )
        # ~650 tokens per scene is a safe budget (all 6 fields)
        dynamic_max_tokens = 32000

        def _call_model(msgs):
            return client.chat.completions.create(
                model=req.chat_model or "gemini-2.5-flash",
                messages=msgs,
                max_tokens=dynamic_max_tokens,
                stream=False,
            )

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]
        resp = _call_model(messages)
        raw = resp.choices[0].message.content.strip()
        raw = _re.sub(r"^```(?:json)?\s*", "", raw, flags=_re.MULTILINE)
        raw = _re.sub(r"\s*```$", "", raw, flags=_re.MULTILINE)
        raw = raw.strip()

        def _try_parse(text):
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                m = _re.search(r"\[[\s\S]+\]", text)
                if m:
                    try:
                        return json.loads(m.group())
                    except json.JSONDecodeError:
                        pass
            return None

        data = _try_parse(raw)
        if data is None:
            # Retry once — model likely truncated due to token pressure
            print(f"[FLOW STORYBOARD] batch offset={offset} non-JSON, retrying. raw[:100]={raw[:100]!r}")
            resp2 = _call_model(messages)
            raw2 = resp2.choices[0].message.content.strip()
            raw2 = _re.sub(r"^```(?:json)?\s*", "", raw2, flags=_re.MULTILINE)
            raw2 = _re.sub(r"\s*```$", "", raw2, flags=_re.MULTILINE)
            raw2 = raw2.strip()
            data = _try_parse(raw2)
            if data is None:
                raise HTTPException(500, f"Model returned non-JSON: {raw2[:300]}")
        if not isinstance(data, list):
            # GPT with json_object returns {"scenes":[...]} not [...]
            for key in ["scenes","data","results","storyboard","items","scene_list"]:
                if key in data and isinstance(data[key], list):
                    data = data[key]; break
            else:
                for v in data.values():
                    if isinstance(v, list):
                        data = v; break
                else:
                    raise HTTPException(500, "Model did not return a scene array")
        return data

    batches = _split_scene_batches(total)  # e.g. 30 -> [(8,0),(8,8),(7,16),(7,23)]

    if len(batches) == 1:
        scenes_data = _gen_batch(*batches[0])
    else:
        # Run each batch as a parallel, independent call (fresh token budget each).
        from concurrent.futures import ThreadPoolExecutor
        with ThreadPoolExecutor(max_workers=min(len(batches), 6)) as pool:
            results = list(pool.map(lambda b: _gen_batch(*b), batches))
        scenes_data = [s for batch in results for s in batch]  # concat in order

    # Re-index globally so indexes are continuous 0 .. N-1 across all batches
    scenes = [{"index": i, **s} for i, s in enumerate(scenes_data)]

    # Optionally generate storyboard images in parallel
    if req.generate_images:
        cfg = IMAGE_MODELS.get(req.model)
        if cfg:
            global IMAGE_API_KEY
            # Use IMAGE_API_KEY from env (LAOZHANG_IMAGE_API_KEY) -- most reliable
            # x_image_api_key from header is secondary fallback
            effective_key = IMAGE_API_KEY or x_image_api_key
            original_key = IMAGE_API_KEY
            try:
                # FIX 2: snapshot key into closure — no global mutation, no race condition
                effective_key_snapshot = effective_key

                def _gen_frame(scene: dict) -> str:
                    style_suffix = f" {req.image_style} style." if req.image_style else ""
                    prompt = (
                        f"{scene.get('description', '')}. "
                        f"Camera: {scene.get('camera', '')}."
                        f"{style_suffix} Cinematic still frame."
                    )
                    try:
                        api = cfg["api"]
                        mdl = cfg["model"]
                        ep = cfg.get("extra_params") or {}
                        if api == "google":
                            r = _requests.post(
                                f"{GOOGLE_IMAGE_BASE}/{mdl}:generateContent",
                                headers={"Authorization": f"Bearer {effective_key_snapshot}",
                                         "Content-Type": "application/json"},
                                json={"contents": [{"parts": [{"text": prompt}]}],
                                      "generationConfig": {"responseModalities": ["IMAGE"],
                                                           "imageConfig": {"aspectRatio": req.aspect_ratio,
                                                                           "imageSize": "1K"}}},
                                timeout=180)
                            r.raise_for_status()
                            return r.json()["candidates"][0]["content"]["parts"][0]["inlineData"]["data"]
                        elif api in ("chat-image-b64", "chat-image-url"):
                            result = _generate_chat_image(prompt, mdl, req.aspect_ratio, "",
                                                          returns_url=(api == "chat-image-url"),
                                                          key=effective_key_snapshot)
                            # FIX 4: convert URL response to base64 so frontend <img> renders it
                            if api == "chat-image-url" and result and result.startswith("http"):
                                import base64 as _b64
                                result = _b64.b64encode(_requests.get(result, timeout=60).content).decode()
                            return result
                        elif api in ("openai-image", "openai-image-url"):
                            result = _generate_openai_image(prompt, mdl, req.aspect_ratio, "1K",
                                                            "", extra_params=ep,
                                                            returns_url=(api == "openai-image-url"),
                                                            key=effective_key_snapshot)
                            # FIX 4: convert URL response to base64 so frontend <img> renders it
                            if api == "openai-image-url" and result and result.startswith("http"):
                                import base64 as _b64
                                result = _b64.b64encode(_requests.get(result, timeout=60).content).decode()
                            return result
                        return ""
                    except Exception as _img_err:
                        print(f"[FLOW IMAGE ERROR] model={req.model} api={cfg.get('api')} err={_img_err}")
                        return ""

                # FIX 1: use concurrent.futures.wait so timeout is total wall-clock,
                # not per-future sequential — prevents crash after future #4 on 26 scenes
                import concurrent.futures as _cf
                with ThreadPoolExecutor(max_workers=4) as pool:
                    futures = [pool.submit(_gen_frame, s) for s in scenes]
                    done, _ = _cf.wait(futures, timeout=120)
                    for i, fut in enumerate(futures):
                        try:
                            scenes[i]["image_b64"] = fut.result(timeout=0) if fut in done else ""
                        except Exception:
                            scenes[i]["image_b64"] = ""
            finally:
                pass  # key no longer mutated globally — nothing to restore

    return {
        "scenes": scenes,
        "style": req.style,
        "scene_count": len(scenes),
    }


# ---------------------------------------------------------------------------
# Entry point

# ---------------------------------------------------------------------------
# One-Shot Fix — Job store + endpoints
# ---------------------------------------------------------------------------
#_oneshot_jobs: dict = {}  # kept for legacy in-process fallback; primary store is now PostgreSQL

ONESHOT_FIX_INSTRUCTION = """
You will receive a complete narasi manuscript. Follow these steps IN ORDER:

STEP 1 — SCAN (before fixing):
Read the ENTIRE manuscript. Identify ALL rule violations — track cross-chapter patterns:
"Bayangkan seorang..." count across all chapters, opening TYPE sequences, bridge type repetitions,
Harari-branded terms, epistemic violations, etc. Fill Checklist Before Fix.

STEP 2 — FIX:
Rewrite the ENTIRE manuscript fixing ALL violations simultaneously.
Be cross-chapter aware: if "Bayangkan seorang..." quota is used up in Bab 3,
Bab 4 and beyond must use ALT techniques. Opening TYPEs must not repeat consecutively.
Track your own fixes as you write each chapter.

STEP 3 — SELF-REVIEW (mandatory before outputting Checklist After Fix):
After writing the fixed manuscript, RE-READ it from start to finish.
Go through every rule in the checklist one by one.
Verify each fix was actually applied correctly in the fixed text.
If you find a rule still violated after your fix, correct it before finalizing output.
Only then fill Checklist After Fix with honest ✅/❌/⚠ status.

OUTPUT MUST FOLLOW THIS EXACT FORMAT (delimiters are mandatory):

## Checklist Before Fix

| # | Rule | Status | Lokasi & Kutipan |
|---|------|--------|-----------------|
[fill every global rule with status ✅/❌/⚠ and location/quote for each violation]

---FIXED_BOOK_START---
[complete fixed narasi — preserve all ## Bab headers and markdown structure as input]
---FIXED_BOOK_END---

## Checklist After Fix

| # | Rule | Status | Perubahan yang Dilakukan |
|---|------|--------|------------------------|
[after self-review: honest status per rule — explain exact change made, or confirm ✅ unchanged]
"""


@app.post("/narasi/oneshot-fix")
async def oneshot_fix_submit(body: dict,
                             user: CurrentUser = Depends(get_current_user)):
    """Submit a one-shot fix job. Returns job_id immediately, processes in background."""
    model     = (body.get("model")     or "gemini-2.5-pro").strip()
    # One-Shot Fix sends persona_style → rules from SERVER persona (not client).
    # VO Optimize sends its own `system` (VO_OPTIMIZE_SYSTEM, not editorial rules).
    _ps = (body.get("persona_style") or "").strip()
    system    = (_review_persona_for(_ps).get("system") if _ps else (body.get("system") or "")).strip()
    content   = (body.get("content")   or "").strip()
    file_name = (body.get("file_name") or "narasi").strip()
    if not content: raise HTTPException(400, "content required")
    if not system:  raise HTTPException(400, "system required")

    temperature = float(body.get("temperature") if body.get("temperature") is not None else 0)
    temperature = max(0.0, min(1.0, temperature))  # clamp 0-1

    # Phase 1 WS3: tenant_id from JWT, user UUID resolved from DB
    _TENANT_ID = user.tenant_id
    _USER_ID   = await _resolve_user_uuid(user.tenant_id, user.user_id)
    job_id = await db.create_job(_TENANT_ID, _USER_ID, "oneshot_fix", file_name)
    await rc.set_progress(job_id, "Memulai analisis...")   # seed live progress in Redis

    # Capture API key before thread spawn (ContextVar not accessible in threads).
    # DeepSeek direct models use DEEPSEEK_API_KEY; everything else uses LaoZhang key.
    _resolved_for_key = MODELS.get(model, model)
    _route_for_thread = _deepseek_route.get()  # capture before thread spawn
    if _resolved_for_key in DEEPSEEK_DIRECT_MODELS or model in DEEPSEEK_DIRECT_MODELS:
        if _route_for_thread == "laozhang":
            api_key = _req_key.get() or API_KEY
        else:
            api_key = DEEPSEEK_API_KEY
            if not api_key:
                raise HTTPException(400, "DEEPSEEK_API_KEY is not set in environment.")
    else:
        api_key = _req_key.get() or API_KEY  # capture before thread spawn

    async def run_job():
        loop = asyncio.get_event_loop()
        try:
            await rc.set_progress(job_id, "AI membaca seluruh manuskrip...")
            resolved = MODELS.get(model, model)
            ceiling  = MODEL_MAX_TOKENS.get(resolved, DEFAULT_MAX_TOKENS)
            client   = OpenAI(api_key=api_key, base_url=BASE_URL, timeout=600.0)
            is_vo_mode = "VO Script Editor" in system or "ANCHOR" in system
            if is_vo_mode:
                user_msg = "NARASI:\n" + content
            else:
                user_msg = ONESHOT_FIX_INSTRUCTION + "\n\nNARASI:\n" + content

            # Blocking OpenAI call → run in a worker thread so the event loop
            # stays free. The result comes back to THIS loop; no cross-loop DB.
            def _call():
                return client.chat.completions.create(
                    model=resolved,
                    messages=[
                        {"role": "system", "content": system},
                        {"role": "user",   "content": user_msg},
                    ],
                    max_tokens=ceiling, temperature=temperature, stream=False,
                )
            resp = await loop.run_in_executor(None, _call)

            raw = (resp.choices[0].message.content or "").strip()
            checklist_before = checklist_after = fixed_book = ""
            if "---FIXED_BOOK_START---" in raw and "---FIXED_BOOK_END---" in raw:
                p1, rest = raw.split("---FIXED_BOOK_START---", 1)
                book_raw, p2 = rest.split("---FIXED_BOOK_END---", 1)
                checklist_before = p1.strip()
                fixed_book       = book_raw.strip()
                checklist_after  = p2.strip()
            else:
                fixed_book = raw
            usage = getattr(resp, "usage", None)
            in_tok  = getattr(usage, "prompt_tokens",     0) if usage else 0
            out_tok = getattr(usage, "completion_tokens", 0) if usage else 0
            finish  = getattr(resp.choices[0], "finish_reason", "unknown")
            print(f"[oneshot-fix] model={resolved} finish={finish} in={in_tok} out={out_tok} max={ceiling} temp={temperature}", flush=True)
            await _log_narasi_usage(_TENANT_ID, _USER_ID, model, resp, job_id=job_id)
            # Step 1.5: capture the AI fix (One-Shot Fix / VO Optimize) as a
            # correction pair (input -> fixed) — same moat signal as a human edit.
            try:
                await db.save_correction_pair(
                    None, _TENANT_ID, _USER_ID, content, fixed_book,
                    body.get("style"), body.get("topic"), None, body.get("language"))
            except Exception as _ce:
                import logging as _lg; _lg.getLogger("narasi").warning("oneshot correction capture failed (non-fatal): %s", _ce)

            # Persist on the MAIN loop → uses the normal pool, no cross-loop error.
            await db.complete_job(job_id, {
                "checklist_before": checklist_before,
                "fixed_book": fixed_book,
                "checklist_after": checklist_after,
                "file_name": file_name,
            })
            await rc.delete_progress(job_id)
        except Exception as e:
            print(f"[oneshot-fix] job {job_id} failed: {e}", flush=True)
            try:
                await db.fail_job(job_id, str(e))
            except Exception as e2:
                print(f"[oneshot-fix] fail_job also failed: {e2}", flush=True)
            try:
                await rc.set_progress(job_id, f"Gagal: {e}", ttl=300)
            except Exception:
                pass

    asyncio.create_task(run_job())
    return {"ok": True, "job_id": job_id}


@app.get("/narasi/oneshot-fix/status/{job_id}")
async def oneshot_fix_status(job_id: str,
                             user: CurrentUser = Depends(get_current_user)):
    """Poll job status — Redis for live progress, Postgres for status/result/error."""
    _TENANT_ID = user.tenant_id
    job = await db.get_job(_TENANT_ID, job_id)
    if not job: raise HTTPException(404, f"Job {job_id} not found")
    # Live progress: Redis first, DB column as fallback (Redis down or key expired).
    live = await rc.get_progress(job_id)
    progress = live if live is not None else job.get("progress_message", "")
    return {"ok": True, "status": job["status"],
            "progress": progress, "error": job.get("error_message")}


@app.get("/narasi/oneshot-fix/result/{job_id}")
async def oneshot_fix_result(job_id: str,
                             user: CurrentUser = Depends(get_current_user)):
    """Retrieve completed result from PostgreSQL."""
    _TENANT_ID = user.tenant_id
    job = await db.get_job(_TENANT_ID, job_id)
    if not job:                      raise HTTPException(404, f"Job {job_id} not found")
    if job["status"] != "done":      raise HTTPException(400, f"Job not done: {job['status']}")
    result = job.get("result_payload") or {}
    return {"ok": True, "file_name": result.get("file_name", "narasi"),
            "checklist_before": result.get("checklist_before", ""),
            "fixed_book":       result.get("fixed_book", ""),
            "checklist_after":  result.get("checklist_after", "")}


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("Starting LaoZhang FastAPI backend at http://127.0.0.1:8000")

    print("Starting LaoZhang FastAPI backend at http://127.0.0.1:8000")
    uvicorn.run(app, host="0.0.0.0", port=8000, timeout_keep_alive=600, h11_max_incomplete_event_size=52428800)

# ==================================================================
